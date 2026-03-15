from PyQt6.QtWidgets import (QWidget, QLabel, QVBoxLayout, QHBoxLayout, QTableWidget, 
                             QTableWidgetItem, QHeaderView, QComboBox, QPushButton, 
                             QScrollArea, QMessageBox, QFileDialog)
from PyQt6.QtGui import QFont, QColor, QBrush
from PyQt6.QtCore import Qt
import datetime
from datetime import datetime as dt_datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import base64

# Set matplotlib backend before importing matplotlib components
import matplotlib
matplotlib.use('Qt5Agg')
import matplotlib.pyplot as plt
plt.ioff()

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import sys, os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'src')))

from src.settings import SettingsController
from src.rent_tracker import RentTracker
from assets.Logger import Logger
logger = Logger()

class ComprehensiveTenantAnalysisTab(QWidget):
    """Comprehensive analysis tab showing aggregated data across all tenants"""
    
    def __init__(self, rent_tracker=None):
        super().__init__()
        logger.debug("ComprehensiveTenantAnalysisTab", "Initializing ComprehensiveTenantAnalysisTab")
        self.apply_theme()
        self.rent_tracker = rent_tracker if rent_tracker else RentTracker()
        self.detailed_view = True  # Track view mode
        self.init_ui()
        logger.info("ComprehensiveTenantAnalysisTab", "ComprehensiveTenantAnalysisTab initialized")
    
    def init_ui(self):
        """Initialize the UI"""
        logger.debug("ComprehensiveTenantAnalysisTab", "Initializing UI")
        # Main layout - will contain header and scrollable content
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Header section (title + refresh button - not scrollable)
        header_layout = QVBoxLayout()
        header_layout.setContentsMargins(24, 24, 24, 0)
        header_layout.setSpacing(16)
        
        # Title
        title = QLabel('Comprehensive Tenant Analysis')
        title.setFont(QFont('Segoe UI', 20, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignLeft)
        header_layout.addWidget(title)
        
        # Controls layout
        controls_layout = QHBoxLayout()
        controls_layout.setSpacing(16)
        
        # Refresh button
        refresh_btn = QPushButton('🔄 Refresh')
        refresh_btn.setStyleSheet('''
            QPushButton {
                background-color: #4caf50;
                color: white;
                border: none;
                padding: 8px 16px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 6px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        ''')
        refresh_btn.clicked.connect(self.refresh_all)
        controls_layout.addWidget(refresh_btn)
        
        # Export to Excel button
        export_btn = QPushButton('📊 Export to Excel')
        export_btn.setStyleSheet('''
            QPushButton {
                background-color: #2196f3;
                color: white;
                border: none;
                padding: 8px 16px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 6px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #0b7dda;
            }
        ''')
        export_btn.clicked.connect(self.export_to_excel)
        controls_layout.addWidget(export_btn)
        
        controls_layout.addStretch()
        
        header_layout.addLayout(controls_layout)
        
        header_widget = QWidget()
        header_widget.setLayout(header_layout)
        main_layout.addWidget(header_widget)
        
        # Scrollable content area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet('''
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                border: none;
                background-color: #f0f0f0;
                width: 12px;
                margin: 0px 0px 0px 0px;
            }
            QScrollBar::handle:vertical {
                background-color: #c0c0c0;
                border-radius: 6px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #a0a0a0;
            }
        ''')
        
        # Content widget for scrolling
        content_widget = QWidget()
        content_layout = QVBoxLayout()
        content_layout.setContentsMargins(24, 24, 24, 24)
        content_layout.setSpacing(24)
        
        # Aggregated Charts Section
        charts_title = QLabel('Aggregated Metrics')
        charts_title.setFont(QFont('Segoe UI', 14, QFont.Weight.Bold))
        content_layout.addWidget(charts_title)
        content_layout.addSpacing(12)
        
        # Create aggregated charts
        chart_layout = QHBoxLayout()
        chart_layout.setSpacing(20)
        
        try:
            # Create figure and canvas for delinquency chart
            self.delinq_fig = Figure(figsize=(5, 5), dpi=100)
            self.delinq_chart = FigureCanvas(self.delinq_fig)
            self.delinq_chart.setMinimumHeight(400)
            
            # Create figure and canvas for monthly chart
            self.month_fig = Figure(figsize=(5, 5), dpi=100)
            self.month_chart = FigureCanvas(self.month_fig)
            self.month_chart.setMinimumHeight(400)
            
            # Create figure and canvas for status chart
            self.status_fig = Figure(figsize=(5, 5), dpi=100)
            self.status_chart = FigureCanvas(self.status_fig)
            self.status_chart.setMinimumHeight(400)
            
            chart_layout.addWidget(self.delinq_chart)
            chart_layout.addWidget(self.month_chart)
            chart_layout.addWidget(self.status_chart)
        except Exception as e:
            logger.error("ComprehensiveTenantAnalysisTab", f"Error creating charts: {e}")
            chart_layout.addWidget(QLabel("Charts unavailable"))
        
        content_layout.addLayout(chart_layout)
        content_layout.addSpacing(24)
        
        # Payment History Line Graphs Section
        payment_title_layout = QHBoxLayout()
        payment_title_layout.setSpacing(20)
        
        payment_title = QLabel('Payment History Trends')
        payment_title.setFont(QFont('Segoe UI', 14, QFont.Weight.Bold))
        payment_title_layout.addWidget(payment_title)
        
        # Add checkbox for detailed view toggle
        from PyQt6.QtWidgets import QCheckBox
        self.detailed_view_checkbox = QCheckBox('Show All Months (Detailed View)')
        self.detailed_view_checkbox.setChecked(True)
        self.detailed_view_checkbox.stateChanged.connect(self.on_view_mode_changed)
        payment_title_layout.addWidget(self.detailed_view_checkbox)
        payment_title_layout.addStretch()
        
        content_layout.addLayout(payment_title_layout)
        content_layout.addSpacing(12)
        
        # Create scroll area for payment history charts
        payment_scroll = QScrollArea()
        payment_scroll.setWidgetResizable(True)
        payment_scroll.setMinimumHeight(550)
        payment_scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                width: 10px;
            }
            QScrollBar::handle:vertical {
                background: #888;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical:hover {
                background: #555;
            }
        """)
        
        # Create container widget for payment charts
        payment_container = QWidget()
        payment_chart_layout = QHBoxLayout()
        payment_chart_layout.setSpacing(20)
        payment_chart_layout.setContentsMargins(0, 0, 0, 0)
        
        try:
            # Create figure and canvas for cumulative payments chart
            self.cumulative_fig = Figure(figsize=(7, 5), dpi=100)
            self.cumulative_chart = FigureCanvas(self.cumulative_fig)
            self.cumulative_chart.setMinimumHeight(450)
            self.cumulative_chart.setMinimumWidth(600)
            
            # Create figure and canvas for monthly rent vs paid chart
            self.monthly_trend_fig = Figure(figsize=(7, 5), dpi=100)
            self.monthly_trend_chart = FigureCanvas(self.monthly_trend_fig)
            self.monthly_trend_chart.setMinimumHeight(450)
            self.monthly_trend_chart.setMinimumWidth(600)
            
            payment_chart_layout.addWidget(self.cumulative_chart)
            payment_chart_layout.addWidget(self.monthly_trend_chart)
        except Exception as e:
            logger.error("ComprehensiveTenantAnalysisTab", f"Error creating payment charts: {e}")
            payment_chart_layout.addWidget(QLabel("Payment charts unavailable"))
        
        payment_container.setLayout(payment_chart_layout)
        payment_scroll.setWidget(payment_container)
        content_layout.addWidget(payment_scroll)
        content_layout.addSpacing(24)
        
        # Individual Tenant Metrics
        individual_title = QLabel('Individual Tenant Breakdown')
        individual_title.setFont(QFont('Segoe UI', 14, QFont.Weight.Bold))
        content_layout.addWidget(individual_title)
        content_layout.addSpacing(12)
        
        # Summary table for each tenant
        self.summary_table = QTableWidget()
        self.summary_table.setColumnCount(8)
        self.summary_table.setHorizontalHeaderLabels([
            'Tenant Name',
            'Status',
            'Rent Amount',
            'Delinquency',
            'Overpayment',
            'Service Credit',
            'This Month: Expected',
            'This Month: Paid'
        ])
        self.summary_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.summary_table.verticalHeader().setVisible(False)
        self.summary_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.summary_table.setStyleSheet('font-size: 12px;')
        self.summary_table.setMinimumHeight(180)
        self.summary_table.setMaximumHeight(250)
        content_layout.addWidget(self.summary_table)
        content_layout.addSpacing(24)
        
        # Monthly Breakdown for Each Tenant
        monthly_title = QLabel('Monthly Payment Status by Tenant')
        monthly_title.setFont(QFont('Segoe UI', 14, QFont.Weight.Bold))
        content_layout.addWidget(monthly_title)
        content_layout.addSpacing(12)
        
        # Monthly detail table
        self.monthly_table = QTableWidget()
        self.monthly_table.setColumnCount(7)
        self.monthly_table.setHorizontalHeaderLabels([
            'Tenant Name',
            'Year-Month',
            'Expected Rent',
            'Amount Paid',
            'Balance',
            'Status',
            'Notes'
        ])
        self.monthly_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.monthly_table.verticalHeader().setVisible(False)
        self.monthly_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.monthly_table.setStyleSheet('font-size: 11px;')
        self.monthly_table.setMinimumHeight(200)
        self.monthly_table.setMaximumHeight(300)
        content_layout.addWidget(self.monthly_table)
        content_layout.addSpacing(24)
        
        # Summary Statistics
        stats_title = QLabel('Summary Statistics')
        stats_title.setFont(QFont('Segoe UI', 14, QFont.Weight.Bold))
        content_layout.addWidget(stats_title)
        content_layout.addSpacing(12)
        
        self.stats_layout = QHBoxLayout()
        self.stats_layout.setSpacing(16)
        content_layout.addLayout(self.stats_layout)
        content_layout.addSpacing(24)
        
        content_widget.setLayout(content_layout)
        scroll_area.setWidget(content_widget)
        
        main_layout.addWidget(scroll_area)
        self.setLayout(main_layout)
        self.refresh_all()
    
    def refresh_all(self):
        """Refresh all data and charts"""
        logger.debug("ComprehensiveTenantAnalysisTab", "Refreshing all data and charts")
        try:
            tenants = self.rent_tracker.get_all_tenants()
            self.update_aggregated_charts(tenants)
            self.update_payment_history_charts(tenants)
            self.update_summary_table(tenants)
            self.update_monthly_table(tenants)
            self.update_statistics(tenants)
            logger.info("ComprehensiveTenantAnalysisTab", f"Refreshed data for {len(tenants)} tenants")
        except Exception as e:
            logger.error("ComprehensiveTenantAnalysisTab", f"Failed to refresh data: {str(e)}")
            QMessageBox.critical(self, "Error", f"Failed to refresh data: {str(e)}")
    
    def on_view_mode_changed(self, state):
        """Handle view mode toggle"""
        logger.debug("ComprehensiveTenantAnalysisTab", f"View mode changed to {'detailed' if state else 'summary'}")
        self.detailed_view = self.detailed_view_checkbox.isChecked()
        # Refresh charts with new view mode
        try:
            tenants = self.rent_tracker.get_all_tenants()
            self.update_payment_history_charts(tenants)
        except Exception as e:
            logger.error("ComprehensiveTenantAnalysisTab", f"Error refreshing charts: {e}")
    
    def update_aggregated_charts(self, tenants):
        """Update the aggregated charts showing all tenant data combined"""
        logger.debug("ComprehensiveTenantAnalysisTab", f"Updating aggregated charts for {len(tenants)} tenants")
        active_tenants = [t for t in tenants if getattr(t, 'account_status', 'active').lower() == 'active']
        
        if not active_tenants:
            return
        
        now = datetime.date.today()
        
        # Calculate aggregated delinquency data
        total_expected_due = 0
        total_delinq = sum(t.delinquency_balance for t in active_tenants)
        future_payment_credits = 0
        
        for tenant in active_tenants:
            if hasattr(tenant, 'months_to_charge') and tenant.months_to_charge:
                for year, month in tenant.months_to_charge:
                    month_date = datetime.date(year, month, 1)
                    if month_date <= now:
                        effective_rent = self.rent_tracker.get_effective_rent(tenant, year, month)
                        total_expected_due += effective_rent
                    elif month_date > now and hasattr(tenant, 'payment_history'):
                        month_key = f"{year}-{month:02d}"
                        future_payments = sum(p.get('amount', 0) for p in tenant.payment_history 
                                            if p.get('payment_month') == month_key)
                        if future_payments > 0:
                            effective_rent = self.rent_tracker.get_effective_rent(tenant, year, month)
                            future_payment_credits += min(future_payments, effective_rent)
        
        adjusted_delinq = max(total_delinq - future_payment_credits, 0)
        delinq_data = [max(adjusted_delinq, 0), max(total_expected_due - adjusted_delinq, 0)]
        
        # Update delinquency chart
        try:
            self.delinq_fig.clear()
            ax1 = self.delinq_fig.add_subplot(111)
            
            if sum(delinq_data) == 0:
                ax1.text(0.5, 0.5, 'No Data', ha='center', va='center', fontsize=14)
                ax1.set_title('All Tenants: Delinquency vs Expected', fontsize=12, pad=20)
            else:
                colors = ['#ff5050', '#0078d4']
                wedges, texts, autotexts = ax1.pie(delinq_data, 
                                            autopct='%1.1f%%',
                                            colors=colors,
                                            startangle=90,
                                            textprops={'fontsize': 10})
                for autotext in autotexts:
                    autotext.set_fontsize(10)
                    autotext.set_color('white')
                    autotext.set_weight('bold')
                
                # Create legend with better positioning
                ax1.legend(['Delinquency', 'Paid/Expected'], 
                          loc='upper right',
                          fontsize=10,
                          framealpha=0.95)
                
                ax1.set_title(f'All Tenants: Delinquency vs Expected\n${adjusted_delinq:.0f} / ${total_expected_due:.0f}',
                             fontsize=12, pad=20)
            
            self.delinq_fig.tight_layout(pad=1.5)
            self.delinq_chart.draw()
        except Exception as e:
            logger.error("ComprehensiveTenantAnalysisTab", f"Error updating delinquency chart: {e}")
        
        # Calculate current month data
        month_expected = 0
        month_paid = 0
        
        for tenant in active_tenants:
            should_charge = False
            if hasattr(tenant, 'months_to_charge'):
                for month_entry in tenant.months_to_charge:
                    if len(month_entry) >= 2 and month_entry[0] == now.year and month_entry[1] == now.month:
                        should_charge = True
                        break
            
            if should_charge:
                effective_rent = self.rent_tracker.get_effective_rent(tenant, now.year, now.month)
                month_expected += effective_rent
                
                actual_paid = 0
                if hasattr(tenant, 'payment_history'):
                    month_key = f"{now.year}-{now.month:02d}"
                    for payment in tenant.payment_history:
                        if payment.get('payment_month') == month_key:
                            actual_paid += payment.get('amount', 0)
                
                paid_toward_expected = min(actual_paid, effective_rent)
                month_paid += paid_toward_expected
        
        month_data = [max(month_paid, 0), max(month_expected - month_paid, 0)]
        
        # Update monthly chart
        try:
            self.month_fig.clear()
            ax2 = self.month_fig.add_subplot(111)
            
            if sum(month_data) == 0:
                ax2.text(0.5, 0.5, 'No Data', ha='center', va='center', fontsize=14)
                ax2.set_title('All Tenants: This Month', fontsize=12, pad=20)
            else:
                colors = ['#0078d4', '#ffb300']
                wedges, texts, autotexts = ax2.pie(month_data,
                                            autopct='%1.1f%%',
                                            colors=colors,
                                            startangle=90,
                                            textprops={'fontsize': 10})
                for autotext in autotexts:
                    autotext.set_fontsize(10)
                    autotext.set_color('white')
                    autotext.set_weight('bold')
                
                # Create legend with better positioning
                ax2.legend(['Paid', 'Outstanding'],
                          loc='upper right',
                          fontsize=10,
                          framealpha=0.95)
                
                ax2.set_title(f'All Tenants: This Month\n${month_paid:.0f} / ${month_expected:.0f}',
                             fontsize=12, pad=20)
            
            self.month_fig.tight_layout(pad=1.5)
            self.month_chart.draw()
        except Exception as e:
            logger.error("ComprehensiveTenantAnalysisTab", f"Error updating monthly chart: {e}")
        
        # Account status chart
        active_count = len([t for t in tenants if getattr(t, 'account_status', 'active').lower() == 'active'])
        inactive_count = len([t for t in tenants if getattr(t, 'account_status', 'active').lower() != 'active'])
        status_data = [active_count, inactive_count]
        
        try:
            self.status_fig.clear()
            ax3 = self.status_fig.add_subplot(111)
            
            if sum(status_data) == 0:
                ax3.text(0.5, 0.5, 'No Data', ha='center', va='center', fontsize=14)
                ax3.set_title('Account Status Distribution', fontsize=12, pad=20)
            else:
                colors = ['#4caf50', '#999999']
                wedges, texts, autotexts = ax3.pie(status_data,
                                            autopct='%1.1f%%',
                                            colors=colors,
                                            startangle=90,
                                            textprops={'fontsize': 10})
                for autotext in autotexts:
                    autotext.set_fontsize(10)
                    autotext.set_color('white')
                    autotext.set_weight('bold')
                
                # Create legend with better positioning
                ax3.legend(['Active', 'Inactive'],
                          loc='upper right',
                          fontsize=10,
                          framealpha=0.95)
                
                ax3.set_title(f'Account Status\n{active_count} Active / {inactive_count} Inactive',
                             fontsize=12, pad=20)
            
            self.status_fig.tight_layout(pad=1.5)
            self.status_chart.draw()
        except Exception as e:
            logger.error("ComprehensiveTenantAnalysisTab", f"Error updating status chart: {e}")
    
    def update_summary_table(self, tenants):
        logger.debug("ComprehensiveTenantAnalysisTab", f"Updating summary table with {len(tenants)} tenants")
        """Update the tenant summary table"""
        self.summary_table.setRowCount(0)
        
        now = datetime.date.today()
        
        for tenant in sorted(tenants, key=lambda t: t.name):
            row = self.summary_table.rowCount()
            self.summary_table.insertRow(row)
            
            # Calculate this month's expected and paid
            month_expected = 0
            month_paid = 0
            
            if hasattr(tenant, 'months_to_charge'):
                for month_entry in tenant.months_to_charge:
                    if len(month_entry) >= 2 and month_entry[0] == now.year and month_entry[1] == now.month:
                        month_expected = self.rent_tracker.get_effective_rent(tenant, now.year, now.month)
                        break
            
            if hasattr(tenant, 'payment_history'):
                for payment in tenant.payment_history:
                    if payment.get('year') == now.year and payment.get('month') == now.month:
                        month_paid += payment.get('amount', 0)
            
            # Add cells
            cells = [
                tenant.name,
                getattr(tenant, 'account_status', 'Active'),
                f"${tenant.rent_amount:.2f}",
                f"${tenant.delinquency_balance:.2f}",
                f"${getattr(tenant, 'overpayment_credit', 0.0):.2f}",
                f"${getattr(tenant, 'service_credit', 0.0):.2f}",
                f"${month_expected:.2f}",
                f"${month_paid:.2f}"
            ]
            
            for col, cell_text in enumerate(cells):
                item = QTableWidgetItem(cell_text)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                
                # Color code delinquency
                if col == 3 and tenant.delinquency_balance > 0:
                    item.setBackground(QColor('#ffcccc'))
                elif col == 4 and getattr(tenant, 'overpayment_credit', 0.0) > 0:
                    item.setBackground(QColor('#ccffcc'))
                
                self.summary_table.setItem(row, col, item)
    
    def update_monthly_table(self, tenants):
        logger.debug("ComprehensiveTenantAnalysisTab", f"Updating monthly table with {len(tenants)} tenants")
        """Update the monthly detail table showing payment status for each tenant and month"""
        self.monthly_table.setRowCount(0)
        
        now = datetime.date.today()
        
        for tenant in sorted(tenants, key=lambda t: t.name):
            if hasattr(tenant, 'months_to_charge') and tenant.months_to_charge:
                for month_entry in tenant.months_to_charge:
                    if len(month_entry) >= 2:
                        year, month = month_entry[0], month_entry[1]
                        month_date = datetime.date(year, month, 1)
                        
                        # Only show months that are due or recently due
                        if month_date <= now:
                            row = self.monthly_table.rowCount()
                            self.monthly_table.insertRow(row)
                            
                            expected_rent = self.rent_tracker.get_effective_rent(tenant, year, month)
                            
                            # Calculate paid for this month
                            amount_paid = 0
                            if hasattr(tenant, 'payment_history'):
                                month_key = f"{year}-{month:02d}"
                                for payment in tenant.payment_history:
                                    if payment.get('payment_month') == month_key:
                                        amount_paid += payment.get('amount', 0)
                            
                            balance = expected_rent - amount_paid
                            
                            # Determine status
                            if amount_paid >= expected_rent:
                                status = "Paid"
                                status_color = QColor('#ccffcc')
                            elif amount_paid > 0:
                                status = "Partial"
                                status_color = QColor('#ffffcc')
                            else:
                                status = "Unpaid"
                                status_color = QColor('#ffcccc')
                            
                            cells = [
                                tenant.name,
                                f"{year}-{month:02d}",
                                f"${expected_rent:.2f}",
                                f"${amount_paid:.2f}",
                                f"${balance:.2f}",
                                status,
                                ""
                            ]
                            
                            for col, cell_text in enumerate(cells):
                                item = QTableWidgetItem(cell_text)
                                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                                
                                if col == 5:  # Status column
                                    item.setBackground(status_color)
                                
                                self.monthly_table.setItem(row, col, item)
    
    def update_statistics(self, tenants):
        logger.debug("ComprehensiveTenantAnalysisTab", "Updating statistics")
        """Update summary statistics"""
        # Clear existing stats
        while self.stats_layout.count():
            widget = self.stats_layout.takeAt(0).widget()
            if widget:
                widget.deleteLater()
        
        if not tenants:
            return
        
        # Calculate statistics
        total_rent = sum(t.rent_amount for t in tenants)
        total_delinq = sum(t.delinquency_balance for t in tenants)
        total_overpay = sum(getattr(t, 'overpayment_credit', 0.0) for t in tenants)
        total_service = sum(getattr(t, 'service_credit', 0.0) for t in tenants)
        active_tenants = len([t for t in tenants if getattr(t, 'account_status', 'active').lower() == 'active'])
        
        stats = [
            ('Total Tenants', str(len(tenants))),
            ('Active Tenants', str(active_tenants)),
            ('Total Monthly Rent', f"${total_rent:.2f}"),
            ('Total Delinquency', f"${total_delinq:.2f}"),
            ('Total Overpayment', f"${total_overpay:.2f}"),
            ('Total Service Credit', f"${total_service:.2f}")
        ]
        
        for label, value in stats:
            stat_widget = QWidget()
            stat_layout = QVBoxLayout()
            stat_layout.setContentsMargins(15, 15, 15, 15)
            stat_layout.setSpacing(12)
            
            label_text = QLabel(label)
            label_text.setFont(QFont('Segoe UI', 11))
            label_text.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            value_text = QLabel(value)
            value_text.setFont(QFont('Segoe UI', 16, QFont.Weight.Bold))
            value_text.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            stat_layout.addWidget(label_text)
            stat_layout.addWidget(value_text)
            stat_widget.setLayout(stat_layout)
            stat_widget.setStyleSheet('''
                QWidget {
                    background-color: #f0f0f0;
                    border: 1px solid #ddd;
                    border-radius: 8px;
                }
            ''')
            stat_widget.setMinimumHeight(100)
            
            self.stats_layout.addWidget(stat_widget)
    
    def update_payment_history_charts(self, tenants):
        logger.debug("ComprehensiveTenantAnalysisTab", f"Updating payment history charts for {len(tenants)} tenants")
        """Update the payment history line charts"""
        active_tenants = [t for t in tenants if getattr(t, 'account_status', 'active').lower() == 'active']
        
        if not active_tenants:
            return
        
        # Update cumulative payments chart
        try:
            self.cumulative_fig.clear()
            ax1 = self.cumulative_fig.add_subplot(111)
            
            all_dates = set()
            tenant_payments = {}
            delinquency_by_date = {}
            earliest_date = None
            latest_date = None
            
            now = datetime.date.today()
            
            # Collect all payment dates and cumulative data per tenant
            for tenant in active_tenants:
                if hasattr(tenant, 'payment_history') and tenant.payment_history:
                    cumulative = 0
                    payments_by_date = {}
                    
                    for payment in sorted(tenant.payment_history, key=lambda p: p.get('date', '')):
                        date_str = payment.get('date', '')
                        if date_str:
                            try:
                                payment_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
                                all_dates.add(payment_date)
                                cumulative += payment.get('amount', 0)
                                payments_by_date[payment_date] = cumulative
                            except ValueError:
                                pass
                    
                    if payments_by_date:
                        tenant_payments[tenant.name] = payments_by_date
            
            # Calculate earliest and latest rental dates
            for tenant in active_tenants:
                if hasattr(tenant, 'months_to_charge') and tenant.months_to_charge:
                    for month_entry in tenant.months_to_charge:
                        if len(month_entry) >= 2:
                            year, month = month_entry[0], month_entry[1]
                            month_date = datetime.date(year, month, 1)
                            
                            if earliest_date is None or month_date < earliest_date:
                                earliest_date = month_date
                            if latest_date is None or month_date > latest_date:
                                latest_date = month_date
            
            # Calculate cumulative expected rent and delinquency per tenant by MONTH
            if all_dates:
                sorted_dates = sorted(all_dates)
                tenant_delinquency = {}  # {tenant_name: {date: delinquency_amount}}
                
                now = datetime.date.today()
                
                # Calculate per-tenant delinquency BY MONTH
                for tenant in active_tenants:
                    tenant_name = tenant.name
                    tenant_delinquency[tenant_name] = {}
                    
                    # Pre-calculate all months this tenant should pay rent for
                    tenant_months_set = set()
                    month_rent_amounts = {}  # {(year, month): rent_amount}
                    
                    if hasattr(tenant, 'months_to_charge'):
                        for month_entry in tenant.months_to_charge:
                            if len(month_entry) >= 2:
                                year, month = month_entry[0], month_entry[1]
                                month_key = (year, month)
                                tenant_months_set.add(month_key)
                                
                                # Check if there's a monthly exception (override) for this month
                                month_str = f"{year}-{month:02d}"
                                if hasattr(tenant, 'monthly_exceptions') and month_str in tenant.monthly_exceptions:
                                    # Use effective rent if there's an override
                                    month_rent_amounts[month_key] = self.rent_tracker.get_effective_rent(tenant, year, month)
                                else:
                                    # Use base rent if no override for this month
                                    month_rent_amounts[month_key] = tenant.rent_amount if hasattr(tenant, 'rent_amount') else 0
                    
                    # Sort months chronologically
                    sorted_months = sorted(list(tenant_months_set))
                    
                    # Build payment history by month
                    payments_by_month = {}  # {(year, month): total_paid_in_month}
                    if hasattr(tenant, 'payment_history'):
                        for payment in tenant.payment_history:
                            payment_month = payment.get('payment_month', '')
                            if payment_month and len(payment_month) == 7:  # YYYY-MM format
                                try:
                                    year, month = int(payment_month[:4]), int(payment_month[5:7])
                                    month_key = (year, month)
                                    if month_key not in payments_by_month:
                                        payments_by_month[month_key] = 0
                                    payments_by_month[month_key] += payment.get('amount', 0)
                                except (ValueError, IndexError):
                                    pass
                    
                    # For each month, calculate delinquency: expected - paid
                    for year, month in sorted_months:
                        month_date = datetime.date(year, month, 1)
                        
                        # Only calculate delinquency for months that have passed
                        if month_date <= now:
                            month_key = (year, month)
                            expected_rent = month_rent_amounts.get(month_key, 0)
                            paid_this_month = payments_by_month.get(month_key, 0)
                            
                            # Delinquency = expected - paid (for this month only)
                            month_delinquency = max(expected_rent - paid_this_month, 0)
                            tenant_delinquency[tenant_name][month_date] = month_delinquency

            
            if all_dates:
                sorted_dates = sorted(all_dates)
                
                # Get unique tenant names from both payments and delinquency data
                all_tenant_names = set(tenant_payments.keys()) | set(tenant_delinquency.keys())
                colors = plt.cm.tab10(range(len(all_tenant_names)))
                color_map = {name: colors[i] for i, name in enumerate(sorted(all_tenant_names))}
                
                for tenant_name in sorted(all_tenant_names):
                    color = color_map[tenant_name]
                    
                    # Plot payment line if this tenant has payments
                    if tenant_name in tenant_payments:
                        all_payment_dates = sorted(tenant_payments[tenant_name].keys())
                        
                        if self.detailed_view:
                            # Detailed: show all payments
                            display_payment_dates = all_payment_dates
                        else:
                            # Summary: filter to show every 2-3 payments to keep clean
                            display_payment_dates = [d for i, d in enumerate(all_payment_dates) if i % 2 == 0]
                        
                        payment_amounts = [tenant_payments[tenant_name][d] for d in display_payment_dates]
                        
                        if display_payment_dates:
                            ax1.plot(display_payment_dates, payment_amounts, marker='o', label=tenant_name, color=color, linewidth=2, markersize=6)
                    
                    # Plot delinquency line for this tenant - filter based on view mode
                    if tenant_name in tenant_delinquency:
                        if self.detailed_view:
                            # Detailed: show all months
                            delinq_dates = sorted(tenant_delinquency[tenant_name].keys())
                        else:
                            # Summary: show every 2nd month
                            all_delinq_dates = sorted(tenant_delinquency[tenant_name].keys())
                            delinq_dates = [d for i, d in enumerate(all_delinq_dates) if i % 2 == 0]
                        
                        delinq_amounts = [tenant_delinquency[tenant_name][d] for d in delinq_dates]
                        
                        if delinq_amounts:
                            ax1.plot(delinq_dates, delinq_amounts, marker='x', 
                                    label=f'{tenant_name} Delinquency', color=color, linewidth=2, 
                                    markersize=7, linestyle='--', alpha=0.7)
                
                # Add current month line
                now = datetime.date.today()
                ax1.axvline(x=now, color='red', linestyle='--', linewidth=2, alpha=0.6, label='Current Month')
                
                # Set x-axis limits to show full timeline from earliest to latest rental date
                if earliest_date and latest_date:
                    ax1.set_xlim(earliest_date, latest_date)
                    
                    # Generate month labels for x-axis
                    current = earliest_date
                    month_dates = []
                    month_labels = []
                    
                    if self.detailed_view:
                        # Show every month
                        step = 1
                    else:
                        # Show every 2-3 months in summary
                        step = 2
                    
                    label_count = 0
                    while current <= latest_date:
                        if label_count % step == 0:
                            month_dates.append(current)
                            month_labels.append(current.strftime('%Y-%m'))
                        # Move to next month
                        if current.month == 12:
                            current = datetime.date(current.year + 1, 1, 1)
                        else:
                            current = datetime.date(current.year, current.month + 1, 1)
                        label_count += 1
                    
                    if month_dates:
                        ax1.set_xticks(month_dates)
                        ax1.set_xticklabels(month_labels, rotation=45, ha='right')
                
                ax1.set_xlabel('Date', fontsize=11)
                ax1.set_ylabel('Amount ($)', fontsize=11)
                ax1.set_title('Cumulative Payment History & Delinquency Per Tenant', fontsize=12, fontweight='bold')
                ax1.legend(loc='best', fontsize=9, framealpha=0.95)
                ax1.grid(True, alpha=0.3)
                ax1.tick_params(axis='x', rotation=45)
            else:
                ax1.text(0.5, 0.5, 'No Payment Data', ha='center', va='center', fontsize=14)
                ax1.set_title('Cumulative Payment History Per Tenant', fontsize=12, fontweight='bold')
            
            self.cumulative_fig.tight_layout()
            self.cumulative_chart.draw()
        except Exception as e:
            logger.error("ComprehensiveTenantAnalysisTab", f"Error updating cumulative payment chart: {e}")
            import traceback
            traceback.print_exc()
        
        # Update monthly rent vs paid chart (trend over all tenants)
        try:
            self.monthly_trend_fig.clear()
            ax2 = self.monthly_trend_fig.add_subplot(111)
            
            month_stats = {}  # {month_key: {'expected': X, 'paid': Y}}
            
            now = datetime.date.today()
            current_month_key = f"{now.year}-{now.month:02d}"
            
            # Calculate expected and paid for each month from months_to_charge
            earliest_date = None
            latest_date = None
            
            for tenant in active_tenants:
                if hasattr(tenant, 'months_to_charge') and tenant.months_to_charge:
                    for month_entry in tenant.months_to_charge:
                        if len(month_entry) >= 2:
                            year, month = month_entry[0], month_entry[1]
                            month_date = datetime.date(year, month, 1)
                            
                            # Track earliest and latest dates
                            if earliest_date is None or month_date < earliest_date:
                                earliest_date = month_date
                            if latest_date is None or month_date > latest_date:
                                latest_date = month_date
                            
                            month_key = f"{year}-{month:02d}"
                            
                            if month_key not in month_stats:
                                month_stats[month_key] = {'expected': 0, 'paid': 0}
                            
                            # Check if there's a monthly exception (override) for this month
                            if hasattr(tenant, 'monthly_exceptions') and month_key in tenant.monthly_exceptions:
                                # Use effective rent if there's an override
                                rent = self.rent_tracker.get_effective_rent(tenant, year, month)
                            else:
                                # Use base rent if no override for this month
                                rent = tenant.rent_amount if hasattr(tenant, 'rent_amount') else 0
                            
                            month_stats[month_key]['expected'] += rent
                            
                            # Calculate paid for this month
                            if hasattr(tenant, 'payment_history'):
                                for payment in tenant.payment_history:
                                    if payment.get('payment_month') == month_key:
                                        month_stats[month_key]['paid'] += payment.get('amount', 0)
            
            if month_stats and earliest_date and latest_date:
                sorted_months = sorted(month_stats.keys())
                
                # Prepare display data based on view mode
                if self.detailed_view:
                    # Detailed view: show all months
                    display_months = sorted_months
                    expected_amounts = [month_stats[m]['expected'] for m in display_months]
                    paid_amounts = [month_stats[m]['paid'] for m in display_months]
                else:
                    # Summary view: aggregate every 3 months (quarterly)
                    display_months = []
                    expected_amounts = []
                    paid_amounts = []
                    
                    for i in range(0, len(sorted_months), 3):
                        # Get the month label from the 3rd month in the group
                        end_idx = min(i + 3, len(sorted_months))
                        quarter_months = sorted_months[i:end_idx]
                        
                        # Use the last month in the quarter as the label
                        display_months.append(quarter_months[-1])
                        
                        # Sum expected and paid for all months in this quarter
                        expected_sum = sum(month_stats[m]['expected'] for m in quarter_months)
                        paid_sum = sum(month_stats[m]['paid'] for m in quarter_months)
                        
                        expected_amounts.append(expected_sum)
                        paid_amounts.append(paid_sum)
                
                # Find current month index for line
                if self.detailed_view:
                    current_idx = sorted_months.index(current_month_key) if current_month_key in sorted_months else len(sorted_months) // 2
                else:
                    # For summary view, find which quarter the current month is in
                    current_idx = 0
                    if current_month_key in sorted_months:
                        month_position = sorted_months.index(current_month_key)
                        current_idx = month_position // 3
                
                x_pos = range(len(display_months))
                ax2.plot(x_pos, expected_amounts, marker='o', label='Expected Rent', color='#ff7f0e', linewidth=2, markersize=6)
                ax2.plot(x_pos, paid_amounts, marker='s', label='Actual Paid', color='#2ca02c', linewidth=2, markersize=6)
                
                # Add current month line
                if current_month_key in sorted_months:
                    ax2.axvline(x=current_idx, color='red', linestyle='--', linewidth=2, alpha=0.6, label='Current Month')
                
                ax2.set_xlabel('Month', fontsize=11)
                ax2.set_ylabel('Amount ($)', fontsize=11)
                ax2.set_title('Expected vs Actual Rent Payments', fontsize=12, fontweight='bold')
                ax2.set_xticks(x_pos)
                ax2.set_xticklabels(display_months, rotation=45, ha='right')
                ax2.legend(loc='upper left', fontsize=9, framealpha=0.95)
                ax2.grid(True, alpha=0.3)
                
                # Show full timeline with some padding
                ax2.set_xlim(-0.5, len(display_months) - 0.5)
            else:
                ax2.text(0.5, 0.5, 'No Data', ha='center', va='center', fontsize=14)
                ax2.set_title('Expected vs Actual Rent Payments', fontsize=12, fontweight='bold')
            
            self.monthly_trend_fig.tight_layout()
            self.monthly_trend_chart.draw()
        except Exception as e:
            logger.error("ComprehensiveTenantAnalysisTab", f"Error updating monthly trend chart: {e}")
            import traceback
            traceback.print_exc()
    
    def export_to_excel(self):
        logger.debug("ComprehensiveTenantAnalysisTab", "Starting Excel export")
        """Export tenant analysis data to Excel workbook"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Excel Report",
                f"Tenant_Analysis_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            from openpyxl import Workbook
            from openpyxl.worksheet.table import Table, TableStyleInfo
            
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            tenants = self.rent_tracker.get_all_tenants()
            active_tenants = [t for t in tenants if getattr(t, 'account_status', 'active').lower() == 'active']
            
            # Create Dashboard sheet
            self._create_dashboard_sheet(wb, active_tenants)
            
            # Create individual tenant sheets
            for tenant in active_tenants:
                self._create_tenant_sheet(wb, tenant)
            
            wb.save(file_path)
            QMessageBox.information(self, "Success", f"Report exported successfully to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export report: {str(e)}")
            logger.error("ComprehensiveTenantAnalysisTab", f"Export error: {e}")
            import traceback
            traceback.print_exc()
    
    def _create_dashboard_sheet(self, wb, tenants):
        """Create dashboard sheet with summary and detailed tenant table"""
        ws = wb.create_sheet("Dashboard", 0)
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        section_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        section_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # Title
        ws[f'A{row}'] = 'PROPERTY MANAGEMENT DASHBOARD'
        ws[f'A{row}'].font = Font(size=16, bold=True)
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        # Report generation date
        ws[f'A{row}'] = f'Report Generated: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws[f'A{row}'].font = Font(size=10, bold=True)
        ws.merge_cells(f'A{row}:G{row}')
        row += 2
        
        # Summary section
        ws[f'A{row}'] = 'SUMMARY STATISTICS'
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        # Calculate totals
        total_expected = sum(getattr(t, 'rent_amount', 0) for t in tenants)
        total_delinquency = sum(getattr(t, 'delinquency_balance', 0) for t in tenants)
        total_overpayment = sum(getattr(t, 'overpayment_credit', 0.0) for t in tenants)
        total_service = sum(getattr(t, 'service_credit', 0.0) for t in tenants)
        
        now = datetime.date.today()
        month_key = f"{now.year}-{now.month:02d}"
        total_paid_this_month = 0
        total_owed_this_month = 0
        
        for tenant in tenants:
            # Calculate owed this month
            expected_rent = self.rent_tracker.get_effective_rent(tenant, now.year, now.month)
            total_owed_this_month += expected_rent
            
            # Calculate paid this month
            if hasattr(tenant, 'payment_history'):
                for payment in tenant.payment_history:
                    if payment.get('payment_month') == month_key:
                        total_paid_this_month += payment.get('amount', 0)
        
        # Calculate remaining owed for current month
        remaining_owed = total_owed_this_month - total_paid_this_month
        
        metrics = [
            ('Total Active Tenants:', len(tenants)),
            ('Total Monthly Expected Rent:', f"${total_expected:,.2f}"),
            ('Total Owed This Month:', f"${total_owed_this_month:,.2f}"),
            ('Total Paid This Month:', f"${total_paid_this_month:,.2f}"),
            ('Remaining Owed This Month:', f"${max(0, remaining_owed):,.2f}"),
            ('Total Delinquency Balance:', f"${total_delinquency:,.2f}"),
            ('Total Overpayment Credit:', f"${total_overpayment:,.2f}"),
            ('Total Service Credit:', f"${total_service:,.2f}"),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].border = border
            row += 1
        
        row += 2
        
        # Tenant summary table
        ws[f'A{row}'] = 'TENANT SUMMARY'
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        headers = ['Tenant Name', 'Status', 'Rent Amount', 'Delinquency', 'Overpayment', 'Service Credit', 'Paid This Month']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        row += 1
        for tenant in tenants:
            ws.cell(row=row, column=1).value = tenant.name
            ws.cell(row=row, column=1).border = border
            
            ws.cell(row=row, column=2).value = getattr(tenant, 'account_status', 'Active')
            ws.cell(row=row, column=2).border = border
            
            rent_amt = getattr(tenant, 'rent_amount', 0)
            ws.cell(row=row, column=3).value = rent_amt
            ws.cell(row=row, column=3).number_format = '$#,##0.00'
            ws.cell(row=row, column=3).border = border
            
            delinq = getattr(tenant, 'delinquency_balance', 0)
            ws.cell(row=row, column=4).value = delinq
            ws.cell(row=row, column=4).number_format = '$#,##0.00'
            ws.cell(row=row, column=4).border = border
            if delinq > 0:
                ws.cell(row=row, column=4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            overpay = getattr(tenant, 'overpayment_credit', 0.0)
            ws.cell(row=row, column=5).value = overpay
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            ws.cell(row=row, column=5).border = border
            if overpay > 0:
                ws.cell(row=row, column=5).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            service = getattr(tenant, 'service_credit', 0.0)
            ws.cell(row=row, column=6).value = service
            ws.cell(row=row, column=6).number_format = '$#,##0.00'
            ws.cell(row=row, column=6).border = border
            
            paid_this_month = 0
            if hasattr(tenant, 'payment_history'):
                for payment in tenant.payment_history:
                    if payment.get('payment_month') == month_key:
                        paid_this_month += payment.get('amount', 0)
            ws.cell(row=row, column=7).value = paid_this_month
            ws.cell(row=row, column=7).number_format = '$#,##0.00'
            ws.cell(row=row, column=7).border = border
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 18
        
        # Create a dedicated area for charts (rows 50+) with hidden data
        chart_data_row = 50
        chart_position_row = 20  # Position charts starting at row 20
        
        try:
            from openpyxl.chart import PieChart, BarChart, Reference
            
            # ===== CHART DATA PREPARATION (will be hidden) =====
            
            # Chart 1 Data: Tenant Status Breakdown
            status_data_row = chart_data_row
            ws[f'L{status_data_row}'] = "Status"
            ws[f'M{status_data_row}'] = "Count"
            status_data_row += 1
            status_data_start = status_data_row
            
            delinquent_count = sum(1 for t in tenants if getattr(t, 'delinquency_balance', 0) > 0)
            healthy_count = sum(1 for t in tenants if getattr(t, 'delinquency_balance', 0) == 0)
            
            if delinquent_count > 0:
                ws[f'L{status_data_row}'] = "Delinquent"
                ws[f'M{status_data_row}'] = delinquent_count
                status_data_row += 1
            
            if healthy_count > 0:
                ws[f'L{status_data_row}'] = "Paid Up"
                ws[f'M{status_data_row}'] = healthy_count
                status_data_row += 1
            
            # Chart 2 Data: Delinquency by Tenant
            delinq_data_row = status_data_row + 2
            ws[f'L{delinq_data_row}'] = "Tenant"
            ws[f'M{delinq_data_row}'] = "Delinquency"
            delinq_data_row += 1
            delinq_data_start = delinq_data_row
            
            for tenant in tenants:
                delinq = getattr(tenant, 'delinquency_balance', 0)
                ws[f'L{delinq_data_row}'] = tenant.name
                ws[f'M{delinq_data_row}'] = delinq
                ws[f'M{delinq_data_row}'].number_format = '$#,##0.00'
                delinq_data_row += 1
            
            # Chart 3 Data: Owed vs Paid This Month
            payment_data_row = delinq_data_row + 2
            ws[f'L{payment_data_row}'] = "Metric"
            ws[f'M{payment_data_row}'] = "Amount"
            payment_data_row += 1
            payment_data_start = payment_data_row
            
            ws[f'L{payment_data_row}'] = "Owed"
            ws[f'M{payment_data_row}'] = total_owed_this_month
            ws[f'M{payment_data_row}'].number_format = '$#,##0.00'
            payment_data_row += 1
            
            ws[f'L{payment_data_row}'] = "Paid"
            ws[f'M{payment_data_row}'] = total_paid_this_month
            ws[f'M{payment_data_row}'].number_format = '$#,##0.00'
            payment_data_row += 1
            
            # ===== CREATE CHARTS =====
            
            # Chart 1: Pie Chart - Status Breakdown
            if status_data_start < status_data_row:
                pie = PieChart()
                pie.title = "Tenant Status Breakdown"
                pie.style = 10
                pie.height = 8
                pie.width = 12
                
                labels = Reference(ws, min_col=12, min_row=status_data_start, max_row=status_data_row-1)
                data = Reference(ws, min_col=13, min_row=status_data_start-1, max_row=status_data_row-1)
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                ws.add_chart(pie, f'A{chart_position_row}')
            
            # Chart 2: Bar Chart - Delinquency by Tenant
            if delinq_data_start < delinq_data_row:
                bar = BarChart()
                bar.type = "col"
                bar.style = 10
                bar.title = "Delinquency by Tenant"
                bar.y_axis.title = "Amount"
                bar.height = 8
                bar.width = 12
                
                data = Reference(ws, min_col=13, min_row=delinq_data_start-1, max_row=delinq_data_row-1)
                labels = Reference(ws, min_col=12, min_row=delinq_data_start, max_row=delinq_data_row-1)
                bar.add_data(data, titles_from_data=True)
                bar.set_categories(labels)
                ws.add_chart(bar, f'H{chart_position_row}')
            
            # Chart 3: Bar Chart - Owed vs Paid
            bar2 = BarChart()
            bar2.type = "col"
            bar2.style = 10
            bar2.title = "This Month: Owed vs Paid"
            bar2.y_axis.title = "Amount"
            bar2.height = 8
            bar2.width = 12
            
            data = Reference(ws, min_col=13, min_row=payment_data_start-1, max_row=payment_data_row-1)
            labels = Reference(ws, min_col=12, min_row=payment_data_start, max_row=payment_data_row-1)
            bar2.add_data(data, titles_from_data=True)
            bar2.set_categories(labels)
            ws.add_chart(bar2, f'A{chart_position_row+12}')
            
            # Hide the chart data columns
            ws.column_dimensions['L'].hidden = True
            ws.column_dimensions['M'].hidden = True
            
        except Exception as e:
            logger.error("ComprehensiveTenantAnalysisTab", f"Error creating charts: {e}")
            import traceback
            traceback.print_exc()
    
    def _sanitize_sheet_name(self, name):
        """Sanitize sheet name by removing invalid Excel characters: : [ ] * ? / \\"""
        invalid_chars = [':', '[', ']', '*', '?', '/', '\\']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '')
        return sanitized[:31]  # Excel sheet name limit is 31 chars
    
    def _calculate_overpayment_created(self, tenant, payment):
        """Calculate if this payment created an overpayment and return the amount"""
        try:
            payment_month = payment.get('payment_month', '')
            if not payment_month:
                return 0.0
            
            # Don't count credit usage as creating overpayment
            if payment.get('is_credit_usage', False):
                return 0.0
            
            year, month = payment_month.split('-')
            year, month = int(year), int(month)
            
            # Get expected rent for that month
            expected_rent = self.rent_tracker.get_effective_rent(tenant, year, month)
            
            # Calculate total paid for that month up to and including this payment
            total_paid = 0.0
            payment_date = payment.get('date', '')
            
            for p in getattr(tenant, 'payment_history', []):
                if p.get('payment_month') == payment_month:
                    # Only include payments up to this payment's date
                    if p.get('date', '') <= payment_date:
                        total_paid += p.get('amount', 0.0)
            
            # If total paid exceeds expected rent, calculate overpayment
            if total_paid > expected_rent:
                # Check if previous payments already caused overpayment
                previous_total = total_paid - payment.get('amount', 0.0)
                if previous_total >= expected_rent:
                    # Previous payments already covered it, no new overpayment
                    return 0.0
                else:
                    # This payment created overpayment
                    return total_paid - expected_rent
            
            return 0.0
        except Exception:
            return 0.0
        header_font = Font(bold=True, color="FFFFFF", size=12)
        section_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        section_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        credit_usage_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow for credit usage rows
        
        # Status color mapping
        status_colors = {
            'Paid in Full': "C6EFCE", 'Overpayment': "92D050", 'Partial Payment': "FFEB9C",
            'Not Paid': "FFC7CE", 'Due Soon': "FFD580", 'Not Due': "D3D3D3",
            'Delinquent': "FF0000", 'N/A': "CCCCCC"
        }
        
        # Color mapping for details
        details_colors = {
            'Overpaid': "92D050", 'remaining': "FFD580", 'Fully paid': "C6EFCE"
        }
        
        row = 1
        
        # Add report generation date at the top
        ws.merge_cells(f'A{row}:H{row}')
        date_cell = ws[f'A{row}']
        date_cell.value = f"Report Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.font = Font(bold=True, size=11)
        date_cell.alignment = Alignment(horizontal='left')
        row += 1
        
        # Add disclaimer at the top
        disclaimer_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws.merge_cells(f'A{row}:H{row}')
        disclaimer_cell = ws[f'A{row}']
        disclaimer_cell.value = "ℹ️ NOTE: Rows highlighted in light yellow are Overpayment Credit or Service Credit usage. These do NOT represent new rent payments and are not included in total payment calculations, as they were already counted when the original overpayment occurred."
        disclaimer_cell.fill = disclaimer_fill
        disclaimer_cell.font = Font(italic=True, size=9)
        disclaimer_cell.alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[row].height = 40
        row += 2
        
        # Helper for section headers
        def add_section_header(title):
            nonlocal row
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws[f'A{row}']
            cell.value = title
            cell.fill = section_fill
            cell.font = section_font
            row += 1
        
        # Helper for adding row with formatting
        def add_row(label, value, width=25):
            nonlocal row
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # ===== TENANT INFORMATION =====
        add_section_header("TENANT INFORMATION")
        add_row("Name", tenant.name)
        add_row("Tenant ID", getattr(tenant, 'tenant_id', 'N/A'))
        
        # Format rental period
        rental_period = getattr(tenant, 'rental_period', {})
        if isinstance(rental_period, dict):
            start_date = rental_period.get('start_date', 'N/A')
            end_date = rental_period.get('end_date', 'N/A')
            period_str = f"{start_date} - {end_date}"
        else:
            period_str = str(rental_period)
        add_row("Rental Period", period_str)
        
        add_row("Rent Amount", getattr(tenant, 'rent_amount', 0))
        ws[f'B{row-1}'].number_format = '$#,##0.00'
        
        add_row("Deposit Amount", getattr(tenant, 'deposit_amount', 0))
        ws[f'B{row-1}'].number_format = '$#,##0.00'
        
        add_row("Due Day", getattr(tenant, 'rent_due_date', 'N/A'))
        
        # Format contact info
        contact_info = getattr(tenant, 'contact_info', {})
        if isinstance(contact_info, dict):
            contact_str = ''
            if contact_info.get('phone'):
                contact_str += f"Phone: {contact_info.get('phone')}"
            if contact_info.get('email'):
                if contact_str:
                    contact_str += " | "
                contact_str += f"Email: {contact_info.get('email')}"
            contact_str = contact_str or 'Not provided'
        else:
            contact_str = str(contact_info) if contact_info else 'Not provided'
        add_row("Contact", contact_str)
        
        # Notes
        notes = getattr(tenant, 'notes', '')
        if isinstance(getattr(tenant, '_notes_list', None), list):
            notes = '; '.join(tenant._notes_list)
        add_row("Notes", notes or '')
        row += 1
        
        # ===== CURRENT STATUS & CALCULATIONS =====
        add_section_header("CURRENT STATUS & CALCULATIONS")
        
        today = datetime.date.today()
        current_month_key = f"{today.year}-{today.month:02d}"
        
        # Get effective rent
        effective_rent = self.rent_tracker.get_effective_rent(tenant, today.year, today.month)
        
        # Get monthly/yearly override
        monthly_override = None
        yearly_override = None
        if hasattr(tenant, 'monthly_exceptions') and tenant.monthly_exceptions:
            monthly_override = tenant.monthly_exceptions.get(current_month_key)
            if monthly_override is None:
                yearly_override = tenant.monthly_exceptions.get(str(today.year))
        
        delinquency_balance = float(getattr(tenant, 'delinquency_balance', 0.0) or 0.0)
        overpayment_credit = float(getattr(tenant, 'overpayment_credit', 0.0) or 0.0)
        service_credit = float(getattr(tenant, 'service_credit', 0.0) or 0.0)
        net_due = delinquency_balance - overpayment_credit
        
        add_row("Current Month", today.strftime('%B %Y'))
        add_row("Base Rent", getattr(tenant, 'rent_amount', 0))
        ws[f'B{row-1}'].number_format = '$#,##0.00'
        
        override_val = f"{monthly_override:.2f}" if monthly_override is not None else (f"{yearly_override:.2f}" if yearly_override is not None else "None")
        add_row("Active Override", override_val)
        if monthly_override is not None or yearly_override is not None:
            ws[f'B{row-1}'].number_format = '$#,##0.00'
        
        add_row("Effective Rent", effective_rent)
        ws[f'B{row-1}'].number_format = '$#,##0.00'
        
        add_row("Delinquency Balance", delinquency_balance)
        delinq_cell = ws[f'B{row-1}']
        delinq_cell.number_format = '$#,##0.00'
        if delinquency_balance > 0:
            delinq_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        add_row("Overpayment Credit", overpayment_credit)
        overpay_cell = ws[f'B{row-1}']
        overpay_cell.number_format = '$#,##0.00'
        if overpayment_credit > 0:
            overpay_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        add_row("Service Credit", service_credit)
        ws[f'B{row-1}'].number_format = '$#,##0.00'
        
        add_row("Net Amount Due", abs(net_due) if net_due != 0 else 0.00)
        net_cell = ws[f'B{row-1}']
        net_cell.number_format = '$#,##0.00'
        if net_due > 0:
            net_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif net_due < 0:
            net_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        row += 1
        
        # ===== PAYMENT HISTORY =====
        add_section_header("PAYMENT HISTORY (Sorted by Date Received)")
        
        # Sort payments by date
        payments = sorted(
            getattr(tenant, 'payment_history', []) or [],
            key=lambda p: dt_datetime.strptime(p.get('date', '1970-01-01'), '%Y-%m-%d')
        )
        
        # Column headers
        headers = ['Date Received', 'Amount', 'Type', 'For Month', 'Status', 'Details', 'Overpayment Created', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1
        
        # Add payment rows with status and details
        for payment in payments:
            date_received = payment.get('date', '')
            amount = float(payment.get('amount', 0.0))
            ptype = payment.get('type', '')
            pmonth = payment.get('payment_month', '')
            
            # Get status and details
            status = self._get_payment_status(tenant, payment)
            details = self._get_payment_details(tenant, payment)
            notes = payment.get('notes', '')
            is_credit_usage = payment.get('is_credit_usage', False)
            is_credit_row = is_credit_usage or 'Overpayment Credit' in ptype or 'Service Credit' in ptype
            overpayment_created = self._calculate_overpayment_created(tenant, payment)
            
            # Date column
            ws.cell(row=row, column=1).value = dt_datetime.strptime(date_received, '%Y-%m-%d') if date_received else None
            ws.cell(row=row, column=1).number_format = 'MM/DD/YYYY'
            
            # Amount column
            ws.cell(row=row, column=2).value = amount
            ws.cell(row=row, column=2).number_format = '$#,##0.00'
            
            ws.cell(row=row, column=3).value = ptype
            ws.cell(row=row, column=4).value = pmonth
            ws.cell(row=row, column=5).value = status
            ws.cell(row=row, column=6).value = details
            ws.cell(row=row, column=7).value = overpayment_created if overpayment_created > 0 else ''
            ws.cell(row=row, column=7).number_format = '$#,##0.00'
            ws.cell(row=row, column=8).value = notes
            
            # Color entire row if it's credit usage
            if is_credit_row:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = credit_usage_fill
            else:
                # Apply status color (for non-credit rows)
                status_cell = ws.cell(row=row, column=5)
                if status in status_colors:
                    status_cell.fill = PatternFill(
                        start_color=status_colors[status],
                        end_color=status_colors[status],
                        fill_type="solid"
                    )
                    if status in ['Delinquent', 'Not Paid']:
                        status_cell.font = Font(bold=True, color="FFFFFF")
                
                # Apply details color
                details_cell = ws.cell(row=row, column=6)
                if 'Overpaid' in details:
                    details_cell.fill = PatternFill(start_color=details_colors['Overpaid'], end_color=details_colors['Overpaid'], fill_type="solid")
                elif 'remaining' in details:
                    details_cell.fill = PatternFill(start_color=details_colors['remaining'], end_color=details_colors['remaining'], fill_type="solid")
                elif 'Fully paid' in details:
                    details_cell.fill = PatternFill(start_color=details_colors['Fully paid'], end_color=details_colors['Fully paid'], fill_type="solid")
            
            # Apply borders
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
            
            row += 1
        row += 1
        
        # ===== MONTHLY BALANCE SUMMARY =====
        add_section_header("MONTHLY BALANCE SUMMARY")
        
        # Column headers
        headers = ['Month', 'Status', 'Rent Due', 'Total Paid', 'Balance']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1
        
        # Add monthly summaries
        summaries = self._get_monthly_summaries_for_export(tenant)
        for month_key in sorted(summaries.keys()):
            summary = summaries[month_key]
            
            ws.cell(row=row, column=1).value = month_key
            
            status = summary.get('status', '')
            ws.cell(row=row, column=2).value = status
            
            rent_due = float(summary.get('rent_due', 0.0))
            ws.cell(row=row, column=3).value = rent_due
            ws.cell(row=row, column=3).number_format = '$#,##0.00'
            
            total_paid = float(summary.get('total_paid', 0.0))
            ws.cell(row=row, column=4).value = total_paid
            ws.cell(row=row, column=4).number_format = '$#,##0.00'
            
            balance = float(summary.get('balance', 0.0))
            ws.cell(row=row, column=5).value = balance
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            
            # Apply status color
            status_cell = ws.cell(row=row, column=2)
            if status in status_colors:
                status_cell.fill = PatternFill(start_color=status_colors[status], end_color=status_colors[status], fill_type="solid")
                if status in ['Delinquent', 'Not Paid']:
                    status_cell.font = Font(bold=True, color="FFFFFF")
            
            # Color balance column
            balance_cell = ws.cell(row=row, column=5)
            if balance > 0:
                balance_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif balance < 0:
                balance_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 25
    
    def _get_payment_status(self, tenant, payment):
        """Get human-readable status for a payment"""
        payment_month = payment.get('payment_month', '')
        if not payment_month:
            return 'Unknown'
        
        try:
            year, month = payment_month.split('-')
            year, month = int(year), int(month)
            
            # Get expected rent for that month
            expected_rent = self.rent_tracker.get_effective_rent(tenant, year, month)
            
            # Get total paid for that month
            total_paid = self._get_total_paid_for_month(tenant, year, month)
            
            if total_paid >= expected_rent:
                if total_paid > expected_rent:
                    return "Overpayment"
                else:
                    return "Paid in Full"
            elif total_paid > 0:
                return "Partial Payment"
            else:
                return "No Payment"
                
        except Exception:
            return 'Unknown'
    
    def _get_payment_details(self, tenant, payment):
        """Get detailed information about a payment"""
        payment_month = payment.get('payment_month', '')
        
        if not payment_month:
            return 'No details available'
        
        try:
            year, month = payment_month.split('-')
            year, month = int(year), int(month)
            
            expected_rent = self.rent_tracker.get_effective_rent(tenant, year, month)
            total_paid = self._get_total_paid_for_month(tenant, year, month)
            
            if total_paid > expected_rent:
                overpayment = total_paid - expected_rent
                return f"Overpaid by ${overpayment:.2f}"
            elif total_paid < expected_rent:
                shortage = expected_rent - total_paid
                return f"${shortage:.2f} remaining"
            else:
                return "Fully paid"
                
        except Exception:
            return 'Details unavailable'
    
    def _get_total_paid_for_month(self, tenant, year, month):
        """Calculate total paid for a specific month"""
        payment_history = getattr(tenant, 'payment_history', [])
        total = 0.0
        
        for payment in payment_history:
            if payment.get('payment_month') == f"{year}-{month:02d}":
                total += payment.get('amount', 0.0)
        
        return total
    
    def _get_monthly_summaries_for_export(self, tenant):
        """Get monthly summaries for Excel export"""
        from datetime import date, timedelta
        summaries = {}
        payment_history = getattr(tenant, 'payment_history', [])
        
        today = date.today()
        
        # Get rental period boundaries
        rental_start_date = None
        rental_end_date = None
        if hasattr(tenant, 'rental_period') and tenant.rental_period:
            try:
                if isinstance(tenant.rental_period, dict):
                    start_date_str = tenant.rental_period.get('start_date')
                    end_date_str = tenant.rental_period.get('end_date')
                elif isinstance(tenant.rental_period, (list, tuple)) and len(tenant.rental_period) >= 2:
                    start_date_str = tenant.rental_period[0]
                    end_date_str = tenant.rental_period[1]
                else:
                    start_date_str = end_date_str = None
                
                if start_date_str and end_date_str:
                    rental_start_date = date.fromisoformat(start_date_str)
                    rental_end_date = date.fromisoformat(end_date_str)
            except:
                pass
        
        # Get all months from payments
        months_to_include = set()
        for payment in payment_history:
            pmonth = payment.get('payment_month', '')
            if pmonth:
                months_to_include.add(pmonth)
        
        # Add months from rental period
        if rental_start_date and rental_end_date:
            current = rental_start_date
            while current <= rental_end_date:
                month_key = f"{current.year}-{current.month:02d}"
                months_to_include.add(month_key)
                # Move to next month
                if current.month == 12:
                    current = current.replace(year=current.year + 1, month=1)
                else:
                    current = current.replace(month=current.month + 1)
        
        # Calculate summaries for each month
        for month_key in sorted(months_to_include):
            try:
                year, month = map(int, month_key.split('-'))
            except:
                continue
            
            expected_rent = self.rent_tracker.get_effective_rent(tenant, year, month)
            total_paid = self._get_total_paid_for_month(tenant, year, month)
            balance = expected_rent - total_paid
            
            # Determine status
            month_date = date(year, month, 1)
            if month_date <= today and balance > 0:
                status = "Delinquent" if month_date < today.replace(day=1) else "Due Soon"
            elif total_paid >= expected_rent:
                status = "Paid in Full" if total_paid == expected_rent else "Overpayment"
            elif month_date > today:
                status = "Not Due"
            else:
                status = "Not Paid"
            
            summaries[month_key] = {
                'rent_due': expected_rent,
                'total_paid': total_paid,
                'balance': balance,
                'status': status
            }
        
        return summaries
    
    def apply_theme(self):
        """Apply theme settings"""
        settings = SettingsController()
        theme = settings.get_theme()
        if theme == 'dark':
            self.setStyleSheet('''
                QWidget { background-color: #181818; color: #f0f0f0; }
                QTableWidget { background: #232323; color: #f0f0f0; }
                QHeaderView::section { background: #282828; color: #f0f0f0; }
                QLabel { color: #f0f0f0; }
            ''')
