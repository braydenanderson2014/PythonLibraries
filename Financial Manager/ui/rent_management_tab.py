from PyQt6.QtWidgets import QWidget, QVBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QHeaderView, QPushButton, QHBoxLayout, QLineEdit, QDoubleSpinBox, QScrollArea, QComboBox, QMessageBox, QCompleter, QFrame, QGridLayout, QTextEdit, QFileDialog
import csv
from PyQt6.QtCore import Qt, QStringListModel
from src.settings import SettingsController
from src.action_queue import ActionQueue
from src.database_worker import DatabaseOperationManager
from datetime import date, datetime
import calendar
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from PyQt6.QtCore import pyqtSignal
from ui.payment_dialog import PaymentDialog
from ui.rent_modification_dialog import RentModificationDialog
from ui.monthly_override_dialog import MonthlyOverrideDialog
from ui.yearly_override_dialog import YearlyOverrideDialog
from ui.rent_query_dialog import RentQueryDialog
from ui.renew_lease_dialog import RenewLeaseDialog
from ui.action_queue_dialog import ActionQueueDialog
from assets.Logger import Logger
from src.rental_summaries import RentalSummaries

logger = Logger()

class RentManagementTab(QWidget):
    tenant_viewed = pyqtSignal(bool)  # True if tenant is loaded, False otherwise

    def __init__(self, rent_tracker):
        super().__init__()
        logger.debug("RentManagementTab", "Initializing RentManagementTab")
        self.rent_tracker = rent_tracker
        self.action_queue = ActionQueue()  # Initialize action queue
        self.rental_summaries = RentalSummaries(rent_tracker=rent_tracker)  # Initialize rental summaries system
        self.dashboard_tab = None  # Will be set by main window
        self.selected_tenant = None
        self.tenant_info_label = None  # Initialize tenant info widget reference
        self.settings = SettingsController()
        self.layout = QVBoxLayout()
        
        # Initialize payment selection tracking
        self.selected_payment_row = None
        self.selected_payment_data = None
        
        # Initialize references for tenant-specific UI elements
        self.rent_buttons_container = None  # Reference to button container layout
        
        # Tenant Selection Section
        self.create_tenant_selection()
        
        # Default settings UI - initially hidden
        self.defaults_widget = QWidget()
        self.defaults_widget.hide()  # Hide by default
        defaults_layout = QVBoxLayout()
        defaults_layout.addWidget(QLabel('Default Rent Settings'))
        self.default_rent = QDoubleSpinBox()
        self.default_rent.setMaximum(100000)
        self.default_rent.setPrefix('$')
        self.default_rent.setValue(self.settings.get('default_rent_amount'))
        defaults_layout.addWidget(QLabel('Default Rent Amount:'))
        defaults_layout.addWidget(self.default_rent)
        self.default_deposit = QDoubleSpinBox()
        self.default_deposit.setMaximum(100000)
        self.default_deposit.setPrefix('$')
        self.default_deposit.setValue(self.settings.get('default_deposit_amount'))
        defaults_layout.addWidget(QLabel('Default Deposit Amount:'))
        defaults_layout.addWidget(self.default_deposit)
        self.default_due_day = QLineEdit()
        self.default_due_day.setText(str(self.settings.get('default_due_day')))
        defaults_layout.addWidget(QLabel('Default Due Day:'))
        defaults_layout.addWidget(self.default_due_day)
        self.save_defaults_btn = QPushButton('Save Defaults')
        self.save_defaults_btn.clicked.connect(self.save_defaults)
        defaults_layout.addWidget(self.save_defaults_btn)
        self.defaults_widget.setLayout(defaults_layout)
        self.layout.addWidget(self.defaults_widget)
        
        # Wrap the entire layout in a scroll area
        self.content_widget = QWidget()
        self.content_widget.setLayout(self.layout)
        
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidget(self.content_widget)
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.scroll_area.setStyleSheet("")
        
        # Create main layout for the entire tab
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(self.scroll_area)
        self.setLayout(main_layout)

        self.apply_theme()

    def _theme_palette(self):
        return self.settings.get_theme_palette()

    def apply_theme(self):
        palette = self._theme_palette()
        self.setStyleSheet(f'''
            QWidget {{
                background-color: {palette['window_bg']};
                color: {palette['text']};
            }}
            QLineEdit, QComboBox, QDoubleSpinBox, QTextEdit, QTableWidget {{
                background-color: {palette['input_bg']};
                color: {palette['text']};
                border: 1px solid {palette['border']};
                border-radius: 6px;
            }}
            QHeaderView::section {{
                background-color: {palette['header_bg']};
                color: {palette['text']};
                border: none;
                padding: 8px;
                font-weight: bold;
            }}
            QScrollArea {{
                border: none;
                background-color: {palette['window_bg']};
            }}
        ''')

        if hasattr(self, 'scroll_area'):
            self.scroll_area.setStyleSheet(f'''
                QScrollArea {{
                    border: none;
                    background-color: {palette['window_bg']};
                }}
                QScrollArea > QWidget > QWidget {{
                    background-color: {palette['window_bg']};
                }}
            ''')

        if hasattr(self, 'tenant_dropdown'):
            self.tenant_dropdown.setStyleSheet(f'''
                QComboBox {{
                    background-color: {palette['input_bg']};
                    color: {palette['text']};
                    border: 2px solid {palette['accent']};
                    border-radius: 5px;
                    padding: 10px;
                    font-size: 14px;
                    min-height: 20px;
                    padding-right: 25px;
                }}
                QComboBox:focus {{
                    border-color: {palette['accent_hover']};
                }}
                QComboBox::drop-down {{
                    border: none;
                    width: 20px;
                    background-color: {palette['accent']};
                    border-top-right-radius: 5px;
                    border-bottom-right-radius: 5px;
                }}
                QComboBox QAbstractItemView {{
                    background-color: {palette['input_bg']};
                    color: {palette['text']};
                    border: 1px solid {palette['border']};
                    selection-background-color: {palette['accent']};
                    selection-color: {palette['button_text']};
                }}
            ''')

    def refresh_theme(self):
        self.apply_theme()
        if self.selected_tenant:
            refreshed_tenant = self.rent_tracker.tenant_manager.get_tenant(self.selected_tenant.tenant_id) or self.selected_tenant
            self.load_tenant(refreshed_tenant)
        else:
            self.refresh_tenant_dropdown()
    
    def save_tenants_async(self, on_success=None, on_error=None):
        """
        Save tenants asynchronously without progress dialog
        
        Args:
            on_success: Callback when save succeeds
            on_error: Callback when save fails
        """
        # Prevent multiple concurrent save operations
        if hasattr(self, '_saving_in_progress') and self._saving_in_progress:
            logger.debug("RentManagementTab", "Save already in progress, skipping duplicate save request")
            return
        
        self._saving_in_progress = True
        
        def default_success(result):
            self._saving_in_progress = False
            logger.info("RentManagementTab", "Tenants saved successfully")
            if on_success:
                on_success(result)
        
        def default_error(error):
            self._saving_in_progress = False
            logger.error("RentManagementTab", f"Failed to save tenants: {error}")
            if on_error:
                on_error(error)
            else:
                QMessageBox.critical(
                    self,
                    "Save Error",
                    f"Failed to save tenant data:\n\n{error}"
                )
        
        # Execute save operation in background using QThread directly
        # to avoid issues with AnimatedProgressDialog cleanup
        from PyQt6.QtCore import QThread, pyqtSignal
        from PyQt6.QtWidgets import QApplication
        
        class SaveWorkerThread(QThread):
            success = pyqtSignal(object)
            error = pyqtSignal(str)
            
            def __init__(self, save_func):
                super().__init__()
                self.save_func = save_func
            
            def run(self):
                try:
                    result = self.save_func()
                    self.success.emit(result)
                except Exception as e:
                    self.error.emit(str(e))
        
        # Create and start worker thread
        worker = SaveWorkerThread(self.rent_tracker.save_tenants)
        worker.success.connect(default_success)
        worker.error.connect(default_error)
        
        # Keep reference to prevent garbage collection
        if not hasattr(self, '_save_workers'):
            self._save_workers = []
        self._save_workers.append(worker)
        
        # Clean up finished workers
        def cleanup_worker():
            try:
                if worker in self._save_workers:
                    self._save_workers.remove(worker)
            except:
                pass
        
        worker.finished.connect(cleanup_worker)
        worker.start()

    def create_tenant_selection(self):
        """Create tenant selection dropdown with search capabilities"""
        # Tenant selection frame
        selection_frame = QWidget()
        selection_layout = QVBoxLayout()
        palette = self._theme_palette()
        
        # Title
        selection_title = QLabel('Select Tenant')
        selection_title.setStyleSheet(f'''
            font-size: 16px;
            font-weight: bold;
            color: {palette['text']};
            margin: 5px 0;
            background-color: {palette['header_bg']};
            padding: 8px;
            border-radius: 4px;
        ''')
        selection_layout.addWidget(selection_title)
        
        # Tenant dropdown with search
        self.tenant_dropdown = QComboBox()
        self.tenant_dropdown.setEditable(True)  # Allow typing for search
        self.tenant_dropdown.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)  # Don't add typed text as items
        self.tenant_dropdown.activated.connect(self.on_tenant_selected_by_index)  # When item selected from dropdown
        self.tenant_dropdown.currentTextChanged.connect(self.on_tenant_text_changed)  # When text changes
        self.tenant_dropdown.setStyleSheet("")
        
        # Set up completer for search functionality
        self.tenant_completer = QCompleter()
        self.tenant_completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self.tenant_completer.setFilterMode(Qt.MatchFlag.MatchContains)
        self.tenant_completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        self.tenant_dropdown.setCompleter(self.tenant_completer)
        
        # Make sure dropdown shows all items when clicked
        self.tenant_dropdown.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents)
        self.tenant_dropdown.setMaxVisibleItems(10)  # Show up to 10 items in dropdown
        
        selection_layout.addWidget(self.tenant_dropdown)
        
        # Add some spacing
        selection_layout.addSpacing(10)
        
        selection_frame.setLayout(selection_layout)
        self.layout.addWidget(selection_frame)
        
        # Load tenants into dropdown
        self.populate_tenant_dropdown()

    def populate_tenant_dropdown(self):
        """Populate the tenant dropdown and auto-select if only one tenant"""
        try:
            tenants = self.rent_tracker.get_all_tenants()
            
            # Clear existing items
            self.tenant_dropdown.clear()
            
            if not tenants:
                self.tenant_dropdown.addItem("No tenants available")
                self.tenant_dropdown.setEnabled(False)
                return
            
            # Add placeholder
            self.tenant_dropdown.addItem("Select a tenant...")
            
            # Add tenant options
            tenant_names = []
            for tenant in tenants:
                display_name = f"{tenant.name} (ID: {tenant.tenant_id})"
                self.tenant_dropdown.addItem(display_name)
                tenant_names.append(display_name)
            
            # Set up completer with tenant names
            completer_model = QStringListModel(tenant_names)
            self.tenant_completer.setModel(completer_model)
            
            self.tenant_dropdown.setEnabled(True)
            
            # Auto-load if only one tenant
            if len(tenants) == 1:
                # Use a timer to delay auto-selection to ensure everything is initialized
                from PyQt6.QtCore import QTimer
                QTimer.singleShot(100, lambda: self.auto_select_single_tenant(tenants[0]))
                
        except Exception as e:
            print(f"[DEBUG] Error populating tenant dropdown: {e}")
            self.tenant_dropdown.addItem("Error loading tenants")
            self.tenant_dropdown.setEnabled(False)

    def auto_select_single_tenant(self, tenant):
        """Auto-select the single tenant after a short delay"""
        try:
            self.tenant_dropdown.setCurrentIndex(1)  # Select the first (and only) tenant
            self.load_tenant(tenant)
        except Exception as e:
            print(f"[DEBUG] Error in auto-select: {e}")

    def on_tenant_selected(self, selected_text):
        """Handle tenant selection from dropdown"""
        if not selected_text or selected_text in ["Select a tenant...", "No tenants available", "Error loading tenants"]:
            return
            
        try:
            # Extract tenant ID from display name (format: "Name (ID: XXXXXXXX)")
            if " (ID: " not in selected_text:
                return
                
            tenant_id = selected_text.split(" (ID: ")[1].rstrip(")")
            
            # Find tenant by ID
            tenants = self.rent_tracker.get_all_tenants()
            for tenant in tenants:
                if tenant.tenant_id == tenant_id:
                    self.load_tenant(tenant)
                    break
                    
        except Exception as e:
            print(f"[DEBUG] Error selecting tenant: {e}")

    def on_tenant_selected_by_index(self, index):
        """Handle tenant selection by dropdown index (when item is clicked)"""
        if index < 0:
            return
        text = self.tenant_dropdown.itemText(index)
        self.on_tenant_selected(text)

    def on_tenant_text_changed(self, text):
        """Handle text changes in dropdown (for search functionality)"""
        # Only auto-select if text exactly matches an item (not during typing)
        if text and text != "Select a tenant..." and text != "No tenants available":
            index = self.tenant_dropdown.findText(text)
            if index >= 0 and text == self.tenant_dropdown.itemText(index):
                # Exact match found, select it
                self.on_tenant_selected(text)

    def set_dashboard_tab(self, dashboard_tab):
        """Set reference to dashboard tab for chart refresh"""
        self.dashboard_tab = dashboard_tab

    def set_tenant_dashboard(self, tenant_dashboard):
        """Set reference to tenant dashboard for analytics"""
        self.tenant_dashboard = tenant_dashboard
    
    def set_comprehensive_analysis_tab(self, analysis_tab):
        """Set reference to comprehensive analysis tab for chart refresh"""
        self.comprehensive_analysis_tab = analysis_tab

    def refresh_dashboard_charts(self):
        """Refresh dashboard charts after payment changes"""
        if self.dashboard_tab and hasattr(self.dashboard_tab, 'refresh_charts'):
            self.dashboard_tab.refresh_charts()
    
    def refresh_tenant_dashboard(self):
        """Refresh tenant analytics dashboard"""
        if hasattr(self, 'tenant_dashboard') and self.tenant_dashboard and self.selected_tenant:
            self.tenant_dashboard.load_tenant(self.selected_tenant)

    def refresh_monthly_balance_display(self):
        """Refresh the monthly balance display and payment information"""
        if self.selected_tenant:
            # Reload the current tenant to refresh all displays
            self.load_tenant(self.selected_tenant)
            # Also refresh associated dashboard displays
            self.refresh_dashboard_charts()
            self.refresh_tenant_dashboard()

    def refresh_tenant_display(self):
        """Refresh the current tenant display"""
        if self.selected_tenant:
            # Get the most up-to-date tenant data
            updated_tenant = self.rent_tracker.tenant_manager.get_tenant(self.selected_tenant.tenant_id)
            if updated_tenant:
                self.load_tenant(updated_tenant)
            else:
                # If the selected tenant no longer exists (e.g., was deleted), load the next available tenant
                print(f"[WARNING] Could not find tenant {self.selected_tenant.tenant_id} for refresh")
                
                # Try to load the currently selected item in the dropdown
                current_index = self.tenant_dropdown.currentIndex()
                if current_index > 0:  # Skip the placeholder "Select a tenant..."
                    current_text = self.tenant_dropdown.currentText()
                    if current_text and current_text != "Select a tenant..." and current_text != "No tenants available":
                        # Extract tenant ID from the display name "Name (ID: id)"
                        try:
                            tenant_id = current_text.split("(ID: ")[-1].rstrip(")")
                            tenant = self.rent_tracker.tenant_manager.get_tenant(tenant_id)
                            if tenant:
                                self.load_tenant(tenant)
                                return
                        except:
                            pass
                
                # If that didn't work, try to find any available tenant
                all_tenants = self.rent_tracker.get_all_tenants()
                if all_tenants:
                    self.load_tenant(all_tenants[0])
                else:
                    # No tenants left, clear the display
                    self.clear_tenant_widgets()

    def load_tenant_by_id(self, tenant_id):
        """Load a tenant by ID and refresh the display"""
        tenant = self.rent_tracker.tenant_manager.get_tenant(tenant_id)
        if tenant:
            self.load_tenant(tenant)
        else:
            print(f"[ERROR] Could not find tenant with ID: {tenant_id}")

    def update_dropdown_selection(self, tenant):
        """Update the dropdown to show the currently loaded tenant"""
        if not tenant or not hasattr(self, 'tenant_dropdown'):
            return
        
        try:
            # Create the display name that should match what's in the dropdown
            display_name = f"{tenant.name} (ID: {tenant.tenant_id})"
            
            # Find this tenant in the dropdown
            index = self.tenant_dropdown.findText(display_name)
            if index >= 0:
                # Temporarily disconnect signals to avoid triggering selection handlers
                self.tenant_dropdown.activated.disconnect()
                self.tenant_dropdown.currentTextChanged.disconnect()
                
                # Set the selection
                self.tenant_dropdown.setCurrentIndex(index)
                
                # Reconnect signals
                self.tenant_dropdown.activated.connect(self.on_tenant_selected_by_index)
                self.tenant_dropdown.currentTextChanged.connect(self.on_tenant_text_changed)
            else:
                print(f"[DEBUG] Could not find tenant '{display_name}' in dropdown")
        except Exception as e:
            print(f"[DEBUG] Error updating dropdown selection: {e}")

    def format_rental_period(self, rental_period):
        """Format rental period as 'Sept 23, 2023 - Sept 23, 2025'"""
        if not rental_period:
            return 'Not set'
        
        try:
            from datetime import datetime
            start_date_str = end_date_str = None
            
            if isinstance(rental_period, dict):
                # New format: {"start_date": "2025-09-11", "end_date": "2026-09-11", "lease_type": "Fixed Term"}
                start_date_str = rental_period.get('start_date')
                end_date_str = rental_period.get('end_date')
            elif isinstance(rental_period, (list, tuple)) and len(rental_period) >= 2:
                # Old format: ["2025-09-11", "2026-09-11"] or ("2025-09-11", "2026-09-11")
                start_date_str = rental_period[0]
                end_date_str = rental_period[1]
            
            if not start_date_str or not end_date_str:
                return 'Invalid format'
            
            # Parse the ISO dates
            start_date = datetime.fromisoformat(start_date_str)
            end_date = datetime.fromisoformat(end_date_str)
            
            # Format as "Sept 23, 2023 - Sept 23, 2025"
            start_formatted = start_date.strftime("%b %d, %Y")
            end_formatted = end_date.strftime("%b %d, %Y")
            
            return f"{start_formatted} - {end_formatted}"
            
        except Exception as e:
            print(f"[DEBUG] Error formatting rental period: {e}")
            return str(rental_period)
    
    def format_contact_info(self, contact_info):
        """Format contact information with phone and email"""
        if not contact_info:
            return 'Not provided'
        
        if isinstance(contact_info, dict):
            parts = []
            
            # Add phone if available
            phone = contact_info.get('phone', '').strip()
            if phone:
                parts.append(f"📞 {phone}")
            
            # Add email if available
            email = contact_info.get('email', '').strip()
            if email:
                parts.append(f"✉️ {email}")
            
            # Add legacy contact if available and no phone/email
            legacy_contact = contact_info.get('contact', '').strip()
            if legacy_contact and not phone and not email:
                parts.append(legacy_contact)
            
            return ' | '.join(parts) if parts else 'Not provided'
        else:
            # Legacy string contact field
            return str(contact_info) if contact_info else 'Not provided'

    def load_tenant(self, tenant):
        print(f"[DEBUG] load_tenant called for {tenant.name}")
        self.tenant_viewed.emit(True)
        # Update delinquency before displaying
        self.rent_tracker.check_and_update_delinquency()
        self.selected_tenant = tenant
        
        # Update dropdown to reflect the loaded tenant
        self.update_dropdown_selection(tenant)
        
        self.clear_tenant_widgets()
        self.defaults_widget.hide()
        
        # Scroll to top to ensure buttons are visible
        if hasattr(self, 'scroll_area') and self.scroll_area:
            scrollbar = self.scroll_area.verticalScrollBar()
            if scrollbar:
                scrollbar.setValue(0)
                print("[DEBUG] Scrolled to top when loading tenant")
        # Format due date as full date, or N/A if outside rental period
        try:
            due_day = int(tenant.rent_due_date) if tenant.rent_due_date else 1
        except Exception:
            due_day = 1
        today = date.today()
        
        # Handle rental period format differences
        period_start, period_end = None, None
        if tenant.rental_period:
            if isinstance(tenant.rental_period, dict):
                # New format: {"start_date": "2025-09-11", "end_date": "2026-09-11", "lease_type": "Fixed Term"}
                period_start = tenant.rental_period.get('start_date')
                period_end = tenant.rental_period.get('end_date')
            elif isinstance(tenant.rental_period, (list, tuple)) and len(tenant.rental_period) >= 2:
                # Old format: ["2025-09-11", "2026-09-11"] or ("2025-09-11", "2026-09-11")
                period_start = tenant.rental_period[0]
                period_end = tenant.rental_period[1]
        
        show_due = True
        try:
            if period_start and period_end:
                start_dt = date.fromisoformat(period_start)
                end_dt = date.fromisoformat(period_end)
                if today < start_dt or today > end_dt:
                    show_due = False
        except Exception:
            show_due = True
        if show_due:
            try:
                due_date_obj = date(today.year, today.month, due_day)
                due_date_str = due_date_obj.strftime('%B %d, %Y')
            except Exception:
                due_date_str = f"{due_day} (invalid date)"
        else:
            due_date_str = "N/A (Rent period ended)"
            
        # Create a professional tenant info section with proper styling
        from PyQt6.QtWidgets import QFrame, QGridLayout
        palette = self._theme_palette()
        section_header_style = f"font-size: 16px; font-weight: bold; color: {palette['text']}; margin: 5px 0; background-color: {palette['header_bg']}; padding: 8px; border-radius: 4px;"
        
        tenant_info_frame = QFrame()
        tenant_info_frame.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        tenant_info_frame.setLineWidth(2)
        tenant_info_frame.setStyleSheet(f'''
            QFrame {{
                background-color: {palette['panel_bg']};
                border: 2px solid {palette['accent']};
                border-radius: 10px;
                margin: 5px;
            }}
        ''')
        
        grid_layout = QGridLayout()
        grid_layout.setSpacing(5)  # Reduce spacing between items
        grid_layout.setContentsMargins(15, 15, 15, 15)  # Reduce margins
        
        # Tenant header
        status_text = (tenant.account_status or 'active').title()
        header_label = QLabel(f"Tenant Information - {tenant.name} ({status_text})")
        header_label.setStyleSheet(f'''
            font-size: 20px;
            font-weight: bold;
            color: {palette['text']};
            margin-bottom: 15px;
            padding: 10px;
            background-color: {palette['header_bg']};
            border-radius: 5px;
        ''')
        grid_layout.addWidget(header_label, 0, 0, 1, 2)
        
        # Debug: print tenant information
        print(f"[DEBUG] Loading tenant: {tenant.name}")
        print(f"[DEBUG] Contact: {tenant.contact_info}")
        print(f"[DEBUG] Rent: {tenant.rent_amount}")
        print(f"[DEBUG] Status: {tenant.account_status}")
        print(f"[DEBUG] Delinquency Balance: {tenant.delinquency_balance}")
        print(f"[DEBUG] Overpayment Credit: {tenant.overpayment_credit}")
        print(f"[DEBUG] Payment History: {len(tenant.payment_history) if hasattr(tenant, 'payment_history') else 'No payment_history attribute'} entries")
        
        # Create info labels in a grid
        service_credit = getattr(tenant, 'service_credit', 0.0)
        # Format notes: support list or string for backward compatibility
        raw_notes = tenant.notes
        if isinstance(getattr(tenant, '_notes_list', None), list):
            notes_display = '\n'.join(f"• {n}" for n in tenant._notes_list) if tenant._notes_list else 'None'
        else:
            notes_display = raw_notes if raw_notes else 'None'

        info_items = [
            ("Contact:", self.format_contact_info(tenant.contact_info)),
            ("Notes:", notes_display),
            ("Rental Period:", self.format_rental_period(tenant.rental_period)),
            ("Rent Amount:", f"${tenant.rent_amount:.2f}"),
            ("Deposit:", f"${tenant.deposit_amount:.2f}"),
            ("Due Date:", due_date_str),
            ("Status:", tenant.account_status),
            ("Delinquency Balance:", f"${tenant.delinquency_balance:.2f}"),
            ("Overpayment Credit:", f"${tenant.overpayment_credit:.2f}"),
            ("Service Credit:", f"${service_credit:.2f}")
        ]
        
        row = 1
        for label_text, value_text in info_items:
            label = QLabel(label_text)
            label.setStyleSheet("""
                QLabel {
                    font-weight: bold; 
                    color: #ffffff; 
                    font-size: 13px;
                    padding: 8px;
                    background-color: #0056b3;
                    border: 1px solid #004085;
                    border-radius: 4px;
                    min-height: 20px;
                    min-width: 120px;
                }
            """)
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            value = QLabel(str(value_text))
            value.setStyleSheet(f'''
                QLabel {{
                    color: {palette['text']};
                    font-size: 13px;
                    padding: 8px;
                    background-color: {palette['surface_bg']};
                    border: 1px solid {palette['border']};
                    border-radius: 4px;
                    min-height: 20px;
                }}
            ''')
            value.setWordWrap(True)
            value.setAlignment(Qt.AlignmentFlag.AlignLeft)
            
            grid_layout.addWidget(label, row, 0)
            grid_layout.addWidget(value, row, 1)
            row += 1
            
        tenant_info_frame.setLayout(grid_layout)
        
        # Create a scroll area for the tenant information
        scroll_area = QScrollArea()
        scroll_area.setWidget(tenant_info_frame)
        scroll_area.setWidgetResizable(True)
        scroll_area.setFixedHeight(300)  # Reduced height to prevent overlap
        scroll_area.setMaximumHeight(300)  # Ensure it doesn't expand
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setStyleSheet(f'''
            QScrollArea {{
                border: 2px solid {palette['accent']};
                border-radius: 10px;
                background-color: {palette['panel_bg']};
                max-height: 300px;
            }}
            QScrollArea > QWidget > QWidget {{
                background-color: {palette['panel_bg']};
            }}
        ''')
        
        self.tenant_info_label = scroll_area  # Keep reference for clearing
        self.layout.addWidget(scroll_area)
        self.layout.addSpacing(10)  # Add spacing before status section
        
        # Add Current Status Section
        self.add_current_status_section(tenant)
        
        self.layout.addSpacing(10)  # Add spacing after status section
        
        # Rent Management Section
        rent_mgmt_label = QLabel('Rent Management')
        rent_mgmt_label.setStyleSheet(section_header_style)
        self.layout.addWidget(rent_mgmt_label)
        
        # Rent management buttons - organized in a consistent grid
        self.rent_buttons_container = QGridLayout()
        self.rent_buttons_container.setHorizontalSpacing(10)
        self.rent_buttons_container.setVerticalSpacing(10)
        
        # Define consistent button style
        button_style = """
            QPushButton {{
                color: white;
                border: none;
                padding: 8px 16px;
                font-size: 12px;
                font-weight: bold;
                border-radius: 4px;
                min-width: 120px;
                max-width: 160px;
            }}
            QPushButton:hover {{
                background-color: {hover_color};
            }}
            QPushButton:pressed {{
                background-color: {pressed_color};
            }}
        """
        
        # First row: Core rent management functions
        rent_buttons_row1 = QHBoxLayout()
        
        self.modify_rent_btn = QPushButton('Modify Rent')
        self.modify_rent_btn.clicked.connect(self.show_modify_rent_dialog)
        self.modify_rent_btn.setStyleSheet(button_style.format(
            hover_color='#138496',
            pressed_color='#117a8b'
        ).replace('QPushButton {', 'QPushButton { background-color: #17a2b8;'))
        
        self.monthly_override_btn = QPushButton('Monthly Override')
        self.monthly_override_btn.clicked.connect(self.show_monthly_override_dialog)
        self.monthly_override_btn.setStyleSheet(button_style.format(
            hover_color='#e8690a',
            pressed_color='#dc6308'
        ).replace('QPushButton {', 'QPushButton { background-color: #fd7e14;'))
        
        self.yearly_override_btn = QPushButton('Yearly Override')
        self.yearly_override_btn.clicked.connect(self.show_yearly_override_dialog)
        self.yearly_override_btn.setStyleSheet(button_style.format(
            hover_color='#5d378a',
            pressed_color='#543078'
        ).replace('QPushButton {', 'QPushButton { background-color: #6f42c1;'))
        
        self.service_credit_btn = QPushButton('Service Credit')
        self.service_credit_btn.clicked.connect(self.show_service_credit_dialog)
        self.service_credit_btn.setStyleSheet(button_style.format(
            hover_color='#218838',
            pressed_color='#1e7e34'
        ).replace('QPushButton {', 'QPushButton { background-color: #28a745;'))
        self.service_credit_btn_original_style = self.service_credit_btn.styleSheet()
        
        # Add Query System button
        self.query_system_btn = QPushButton('Query System')
        self.query_system_btn.clicked.connect(self.show_query_dialog)
        self.query_system_btn.setStyleSheet(button_style.format(
            hover_color='#138496',
            pressed_color='#11727d'
        ).replace('QPushButton {', 'QPushButton { background-color: #17a2b8;'))
        
        # Add Renew Lease button
        self.renew_lease_btn = QPushButton('Renew Lease')
        self.renew_lease_btn.clicked.connect(self.show_renew_lease_dialog)
        self.renew_lease_btn.setStyleSheet(button_style.format(
            hover_color='#218838',
            pressed_color='#1e7e34'
        ).replace('QPushButton {', 'QPushButton { background-color: #28a745;'))
        self.renew_lease_btn_original_style = self.renew_lease_btn.styleSheet()
        
        # Add Scheduled Actions button
        self.scheduled_actions_btn = QPushButton('Scheduled Actions')
        self.scheduled_actions_btn.clicked.connect(self.show_scheduled_actions)
        self.scheduled_actions_btn.setStyleSheet(button_style.format(
            hover_color='#545b62',
            pressed_color='#454a4f'
        ).replace('QPushButton {', 'QPushButton { background-color: #6c757d;'))
        
        # Schedule Notification Button
        self.schedule_notification_btn = QPushButton('Schedule Notification')
        self.schedule_notification_btn.clicked.connect(self.show_schedule_notification_dialog)
        self.schedule_notification_btn.setStyleSheet(button_style.format(
            hover_color='#e8680f',
            pressed_color='#d15a00'
        ).replace('QPushButton {', 'QPushButton { background-color: #fd7e14;'))

        # Export CSV Button
        self.export_csv_btn = QPushButton('Export CSV')
        self.export_csv_btn.clicked.connect(self.export_rent_management_to_csv)
        self.export_csv_btn.setStyleSheet(button_style.format(
            hover_color='#0b5ed7',
            pressed_color='#0a58ca'
        ).replace('QPushButton {', 'QPushButton { background-color: #0d6efd;'))
        
        # Add Payment Button
        self.add_payment_btn = QPushButton('Add Payment')
        self.add_payment_btn.clicked.connect(self.show_add_payment_dialog)
        self.add_payment_btn.setStyleSheet(button_style.format(
            hover_color='#218838',
            pressed_color='#1e7e34'
        ).replace('QPushButton {', 'QPushButton { background-color: #28a745;'))
        self.add_payment_btn_original_style = self.add_payment_btn.styleSheet()
        
        # Modify Payment Button
        self.modify_payment_btn = QPushButton('Modify Payment')
        self.modify_payment_btn.clicked.connect(self.on_modify_payment_button_clicked)
        self.modify_payment_btn.setEnabled(False)  # Disabled by default
        self.modify_payment_btn.setStyleSheet(button_style.format(
                hover_color='#e0a800',
                pressed_color='#d39e00'
            ).replace('QPushButton {', 'QPushButton { background-color: #ffc107; color: #212529;'))
        self.modify_payment_btn_original_style = self.modify_payment_btn.styleSheet()
        
        # Refund/Reverse Payment Button
        self.add_refund_btn = QPushButton('Refund/Reverse Payment')
        self.add_refund_btn.clicked.connect(self.show_add_refund_dialog)
        self.add_refund_btn.setStyleSheet(button_style.format(
            hover_color='#c82333',
            pressed_color='#bd2130'
        ).replace('QPushButton {', 'QPushButton { background-color: #dc3545; color: #ffffff;'))
        self.add_refund_btn_original_style = self.add_refund_btn.styleSheet()
            
        self.view_summary_btn = QPushButton('View Summary')
        self.view_summary_btn.clicked.connect(self.show_rental_period_summary)
        self.view_summary_btn.setStyleSheet(button_style.format(
            hover_color='#0b5ed7',
            pressed_color='#0a58ca'
        ).replace('QPushButton {', 'QPushButton { background-color: #0d6efd;'))
        
        self.monthly_summary_btn = QPushButton('Monthly Summary')
        self.monthly_summary_btn.clicked.connect(self.show_monthly_summary_dialog)
        self.monthly_summary_btn.setStyleSheet(button_style.format(
            hover_color='#198754',
            pressed_color='#157347'
        ).replace('QPushButton {', 'QPushButton { background-color: #1b5e20;').replace('color: white', 'color: #ffffff'))
        
        self.yearly_summary_btn = QPushButton('Yearly Summary')
        self.yearly_summary_btn.clicked.connect(self.show_yearly_summary_dialog)
        self.yearly_summary_btn.setStyleSheet(button_style.format(
            hover_color='#0d47a1',
            pressed_color='#081a50'
        ).replace('QPushButton {', 'QPushButton { background-color: #1565c0;'))
        
        action_buttons = [
            # Row 0: Primary payment actions + quick summaries (most frequent tasks)
            self.add_payment_btn,
            self.modify_payment_btn,
            self.add_refund_btn,
            self.view_summary_btn,
            self.monthly_summary_btn,
            # Row 1: Rent adjustments
            self.modify_rent_btn,
            self.monthly_override_btn,
            self.yearly_override_btn,
            self.service_credit_btn,
            self.yearly_summary_btn,
            # Row 2: Lease management & tools
            self.renew_lease_btn,
            self.scheduled_actions_btn,
            self.schedule_notification_btn,
            self.export_csv_btn,
            self.query_system_btn,
        ]

        for index, button in enumerate(action_buttons):
            row = index // 5
            column = index % 5
            self.rent_buttons_container.addWidget(button, row, column)

        for column in range(5):
            self.rent_buttons_container.setColumnStretch(column, 1)
        
        self.layout.addLayout(self.rent_buttons_container)
        print(f"[DEBUG] Added rent buttons container to layout")
        self.layout.addSpacing(15)
        
        # Payment history section with better styling (outside scroll area)
        history_label = QLabel('Payment History')
        history_label.setStyleSheet(section_header_style)
        self.payment_table_label = history_label
        self.layout.addWidget(self.payment_table_label)
        self.layout.addSpacing(3)
        self.payment_table = QTableWidget()
        self.payment_table.setColumnCount(7)
        self.payment_table.setHorizontalHeaderLabels(['Date Received', 'Amount', 'Type', 'For Month', 'Status', 'Month Status', 'Notes'])
        # Set fixed widths for some columns and allow horizontal scrolling
        self.payment_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.payment_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.payment_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.payment_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.payment_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.payment_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        self.payment_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)
        self.payment_table.setColumnWidth(0, 100)  # Date
        self.payment_table.setColumnWidth(1, 80)   # Amount
        self.payment_table.setColumnWidth(3, 90)   # For Month
        self.payment_table.setColumnWidth(4, 100)  # Status
        self.payment_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.payment_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.payment_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)  # Make table read-only
        self.payment_table.setStyleSheet(f'''
            QTableWidget {{
                background-color: {palette['input_bg']};
                border: 1px solid {palette['border']};
                border-radius: 8px;
                gridline-color: {palette['border']};
                font-size: 13px;
                color: {palette['text']};
            }}
            QHeaderView::section {{
                background-color: {palette['header_bg']};
                padding: 8px;
                border: none;
                font-weight: bold;
                color: {palette['text']};
            }}
            QTableWidget::item {{
                padding: 8px;
                border-bottom: 1px solid {palette['border']};
            }}
        ''')
        self.payment_table.setAlternatingRowColors(True)
        self.payment_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.payment_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.payment_table.itemSelectionChanged.connect(self.on_payment_selection_changed)
        self.payment_table.setMaximumHeight(200)  # Limit table height
        self.payment_table.setMinimumHeight(120)  # Ensure minimum visibility
        
        # Enable sorting on the table
        self.payment_table.setSortingEnabled(False)  # We'll handle sorting manually
        self.payment_table.horizontalHeader().sectionClicked.connect(self.sort_payment_table)
        
        # Track current sort state (column, ascending/descending)
        self.payment_sort_column = None
        self.payment_sort_ascending = True
        
        # Populate payment table with individual payments
        self.populate_payment_table(tenant)
        
        self.layout.addWidget(self.payment_table)
        
        # Add payment section
        self.add_payment_section(tenant)
        
        # Add tenant management section
        self.add_tenant_management_section(tenant)
        
        # Update button states based on account status
        self.update_button_states(tenant)
        
        # Refresh tenant analytics dashboard
        self.refresh_tenant_dashboard()
    
    def add_tenant_management_section(self, tenant):
        """Add tenant management options (delete, backup, etc.)"""
        from PyQt6.QtWidgets import QHBoxLayout, QPushButton
        
        # Tenant Management Section
        mgmt_label = QLabel('Tenant Management')
        mgmt_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #0056b3; margin: 5px 0; background-color: #f8f9fa; padding: 8px; border-radius: 4px;")
        self.layout.addWidget(mgmt_label)
        
        # Management buttons layout
        mgmt_buttons_layout = QHBoxLayout()
        
        # Backup Tenant Button
        backup_btn = QPushButton('Backup This Tenant')
        backup_btn.clicked.connect(lambda: self.backup_tenant(tenant))
        backup_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 8px 16px;
                font-size: 12px;
                font-weight: bold;
                border-radius: 4px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
            QPushButton:pressed {
                background-color: #1e7e34;
            }
        """)
        mgmt_buttons_layout.addWidget(backup_btn)
        
        # Delete Tenant Button
        delete_btn = QPushButton('Delete Tenant')
        delete_btn.clicked.connect(lambda: self.delete_tenant(tenant))
        delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                padding: 8px 16px;
                font-size: 12px;
                font-weight: bold;
                border-radius: 4px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
            QPushButton:pressed {
                background-color: #bd2130;
            }
        """)
        mgmt_buttons_layout.addWidget(delete_btn)
        mgmt_buttons_layout.addStretch()
        
        self.layout.addLayout(mgmt_buttons_layout)
        self.layout.addSpacing(10)
        
    def update_button_states(self, tenant):
        """Enable/disable buttons based on tenant account status"""
        account_status = getattr(tenant, 'account_status', 'active')
        # Normalize to lowercase for consistent comparison
        account_status = account_status.lower() if account_status else 'active'
        
        # Define button groups
        # Payment buttons - disabled for inactive (expired lease) and terminated
        payment_buttons = [
            ('add_payment_btn', 'add_payment_btn_original_style'),
            ('modify_payment_btn', 'modify_payment_btn_original_style'),
            ('add_refund_btn', 'add_refund_btn_original_style')
        ]
        
        # Service credit buttons - only enabled for active
        service_buttons = [
            ('service_credit_btn', 'service_credit_btn_original_style')
        ]
        
        # Renew lease button - enabled for active and inactive (to renew expired leases)
        renew_lease_buttons = [
            ('renew_lease_btn', 'renew_lease_btn_original_style')
        ]
        
        # Disable payment buttons if inactive or terminated
        is_expired_or_terminated = account_status in ['inactive', 'terminated']
        
        # Enable renew lease for active and inactive (expired leases)
        can_renew = account_status in ['active', 'inactive']
        
        # Define disabled style
        disabled_style = """
            QPushButton {
                background-color: #e9ecef;
                color: #6c757d;
                border: 2px solid #dee2e6;
                border-radius: 4px;
                padding: 8px 16px;
                font-size: 12px;
                font-weight: bold;
                min-width: 120px;
                max-width: 140px;
            }
            QPushButton:hover {
                background-color: #e9ecef;
            }
            QPushButton:pressed {
                background-color: #e9ecef;
            }
        """
        
        # Payment buttons
        for button_name, style_attr in payment_buttons:
            if hasattr(self, button_name):
                button = getattr(self, button_name)
                is_disabled = is_expired_or_terminated
                button.setEnabled(not is_disabled)
                if is_disabled:
                    button.setStyleSheet(disabled_style)
                else:
                    # Restore original style when enabling
                    if hasattr(self, style_attr):
                        original_style = getattr(self, style_attr)
                        button.setStyleSheet(original_style)
        
        # Service credit buttons
        for button_name, style_attr in service_buttons:
            if hasattr(self, button_name):
                button = getattr(self, button_name)
                is_disabled = account_status != 'active'
                button.setEnabled(not is_disabled)
                if is_disabled:
                    button.setStyleSheet(disabled_style)
                else:
                    # Restore original style when enabling
                    if hasattr(self, style_attr):
                        original_style = getattr(self, style_attr)
                        button.setStyleSheet(original_style)
        
        # Renew lease button
        for button_name, style_attr in renew_lease_buttons:
            if hasattr(self, button_name):
                button = getattr(self, button_name)
                is_disabled = not can_renew
                button.setEnabled(not is_disabled)
                if is_disabled:
                    button.setStyleSheet(disabled_style)
                else:
                    # Restore original style when enabling
                    if hasattr(self, style_attr):
                        original_style = getattr(self, style_attr)
                        button.setStyleSheet(original_style)
    
    def add_current_status_section(self, tenant):
        """Add current status section showing overrides, calculations, and important info"""
        from datetime import date
        
        # Create status frame
        status_frame = QFrame()
        status_frame.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        status_frame.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa;
                border: 2px solid #28a745;
                border-radius: 10px;
                margin: 5px;
            }
        """)
        
        status_layout = QVBoxLayout()
        status_layout.setContentsMargins(15, 15, 15, 15)
        status_layout.setSpacing(10)
        
        # Status header
        status_header = QLabel("📊 Current Status & Calculations")
        status_header.setStyleSheet("""
            font-size: 18px; 
            font-weight: bold; 
            color: #28a745; 
            margin-bottom: 10px;
            padding: 8px;
            background-color: #d4edda;
            border-radius: 5px;
            border: 1px solid #c3e6cb;
        """)
        status_layout.addWidget(status_header)
        
        # Create grid for status information
        status_grid = QGridLayout()
        status_grid.setSpacing(8)
        
        # Get current date and calculations
        today = date.today()
        current_month_key = f"{today.year}-{today.month:02d}"
        current_year_key = str(today.year)
        
        # Get effective rent which includes account status check
        effective_rent = self.rent_tracker.get_effective_rent(tenant, today.year, today.month)
        
        # Get override info for display
        monthly_override = None
        yearly_override = None
        if hasattr(tenant, 'monthly_exceptions') and tenant.monthly_exceptions:
            monthly_override = tenant.monthly_exceptions.get(current_month_key)
            if monthly_override is None:
                yearly_override = tenant.monthly_exceptions.get(current_year_key)
        
        # Calculate delinquency and overpayment
        delinquency_balance = getattr(tenant, 'delinquency_balance', 0.0)
        overpayment_credit = getattr(tenant, 'overpayment_credit', 0.0)
        
        row = 0
        
        # Current Month Rent Status
        self.add_status_row(status_grid, row, "📅 Current Month:", f"{today.strftime('%B %Y')}", "#007bff")
        row += 1
        
        # Base rent amount
        self.add_status_row(status_grid, row, "💰 Base Rent:", f"${tenant.rent_amount:.2f}", "#6c757d")
        row += 1
        
        # Override information
        override_info = "None"
        override_color = "#6c757d"
        if monthly_override is not None:
            override_info = f"${monthly_override:.2f} (Monthly Override)"
            override_color = "#fd7e14"
        elif yearly_override is not None:
            override_info = f"${yearly_override:.2f} (Yearly Override)"
            override_color = "#6f42c1"
        
        self.add_status_row(status_grid, row, "🔧 Active Override:", override_info, override_color)
        row += 1
        
        # Effective rent for current month
        effective_color = "#28a745" if effective_rent == tenant.rent_amount else "#fd7e14"
        self.add_status_row(status_grid, row, "💵 Effective Rent:", f"${effective_rent:.2f}", effective_color)
        row += 1
        
        # Balance information
        if delinquency_balance > 0:
            self.add_status_row(status_grid, row, "⚠️ Delinquency:", f"${delinquency_balance:.2f}", "#dc3545")
            row += 1
        
        if overpayment_credit > 0:
            self.add_status_row(status_grid, row, "✅ Overpayment Credit:", f"${overpayment_credit:.2f}", "#28a745")
            row += 1

        # Service Credit balance
        try:
            service_credit = float(getattr(tenant, 'service_credit', 0.0) or 0.0)
        except Exception:
            service_credit = 0.0
        if service_credit > 0:
            self.add_status_row(status_grid, row, "🛠️ Service Credit:", f"${service_credit:.2f}", "#28a745")
            row += 1
        
        # Net amount due calculation
        # Delinquency balance already includes rent calculations, so just subtract overpayment credit
        # Service credit is already factored into delinquency_balance via payment history, so we don't subtract it here
        net_due = delinquency_balance - overpayment_credit
        due_color = "#dc3545" if net_due > 0 else "#28a745" if net_due < 0 else "#6c757d"
        due_text = f"${abs(net_due):.2f}"
        if net_due < 0:
            due_text = f"Credit: ${abs(net_due):.2f}"
        elif net_due == 0:
            due_text = "Paid in Full"
        
        self.add_status_row(status_grid, row, "🎯 Net Amount Due:", due_text, due_color)
        row += 1
        
        # Show all active overrides summary
        if hasattr(tenant, 'monthly_exceptions') and tenant.monthly_exceptions:
            active_overrides = []
            for key, amount in tenant.monthly_exceptions.items():
                if len(key) == 4:  # Yearly override
                    active_overrides.append(f"{key}: ${amount:.2f}")
                else:  # Monthly override
                    try:
                        year, month = key.split('-')
                        month_name = date(int(year), int(month), 1).strftime('%b %Y')
                        active_overrides.append(f"{month_name}: ${amount:.2f}")
                    except:
                        active_overrides.append(f"{key}: ${amount:.2f}")
            
            if active_overrides:
                override_summary = "; ".join(active_overrides[:3])  # Show first 3
                if len(active_overrides) > 3:
                    override_summary += f" (+{len(active_overrides)-3} more)"
                self.add_status_row(status_grid, row, "📋 All Overrides:", override_summary, "#6f42c1")
        
        status_layout.addLayout(status_grid)
        status_frame.setLayout(status_layout)
        
        # Add frame to main layout
        self.layout.addWidget(status_frame)
        
    def add_status_row(self, grid_layout, row, label_text, value_text, color="#000000"):
        """Helper method to add a status row"""
        label = QLabel(label_text)
        label.setStyleSheet(f"""
            font-weight: bold;
            color: {color};
            padding: 5px;
            background-color: white;
            border-radius: 3px;
            min-width: 150px;
        """)
        
        value = QLabel(value_text)
        value.setStyleSheet(f"""
            color: {color};
            font-weight: bold;
            padding: 5px;
            background-color: white;
            border-radius: 3px;
            border-left: 3px solid {color};
        """)
        value.setWordWrap(True)
        
        grid_layout.addWidget(label, row, 0)
        grid_layout.addWidget(value, row, 1)

    def add_payment_section(self, tenant):
        """Add payment tracking section with monthly balance summary"""
        self.layout.addSpacing(5)
        
        # Add monthly balance summary (which includes its own filter controls)
        self.add_balance_summary(tenant)
        
        self.layout.addSpacing(10)
    
    def add_balance_filtering_controls(self):
        """Add filtering controls for monthly balance summary"""
        # Check if filter controls already exist and are visible
        controls_exist_and_visible = False
        if hasattr(self, 'balance_filter_combo') and self.balance_filter_combo is not None:
            try:
                # Check if the combo box is still in the layout and visible
                parent_layout = self.balance_filter_combo.parent()
                if parent_layout is not None and self.balance_filter_combo.isVisible():
                    # Additional check: make sure it's actually in our layout
                    for i in range(self.layout.count()):
                        item = self.layout.itemAt(i)
                        if item and item.widget() == self.balance_filter_combo.parent():
                            controls_exist_and_visible = True
                            break
            except:
                pass
        
        if controls_exist_and_visible:
            return  # Controls already exist and are valid, don't duplicate
        
        print("[DEBUG] Creating new filter controls")
        
        # Get the current number of widgets in layout for positioning
        current_widget_count = self.layout.count()
        print(f"[DEBUG] Current layout widget count: {current_widget_count}")
        
        # Filtering section
        filter_frame = QFrame()
        filter_frame.setObjectName("BalanceFilterFrame")  # Set object name for debugging
        filter_frame.setFrameStyle(QFrame.Shape.StyledPanel)
        filter_frame.setStyleSheet("""
            QFrame#BalanceFilterFrame {
                background-color: #e3f2fd;
                border: 2px solid #1976d2;
                border-radius: 8px;
                padding: 15px;
                margin: 15px 0;
                min-height: 60px;
            }
        """)
        
        filter_layout = QHBoxLayout()
        filter_frame.setLayout(filter_layout)
        
        # Filter type label and dropdown
        filter_label = QLabel('Show:')
        filter_label.setStyleSheet("font-weight: bold; color: #1976d2; margin-right: 10px; font-size: 14px;")
        filter_layout.addWidget(filter_label)
        
        # Filter dropdown
        self.balance_filter_combo = QComboBox()
        self.balance_filter_combo.setObjectName("BalanceFilterCombo")
        self.balance_filter_combo.addItems([
            'Current + Delinquent + Near Future',
            'Current Month Only',
            'Delinquent Months Only',
            'Future Months Only', 
            'Past Months Only',
            'All Relevant Months',
            'Months with Payments Only'
        ])
        self.balance_filter_combo.setStyleSheet("""
            QComboBox#BalanceFilterCombo {
                padding: 8px 12px;
                font-size: 14px;
                font-weight: bold;
                border: 2px solid #1976d2;
                border-radius: 4px;
                background-color: white;
                min-width: 250px;
                color: #1976d2;
            }
            QComboBox#BalanceFilterCombo:hover {
                border-color: #0d47a1;
                background-color: #f3e5f5;
            }
        """)
        self.balance_filter_combo.currentTextChanged.connect(self.on_balance_filter_changed)
        filter_layout.addWidget(self.balance_filter_combo)
        
        # Spacer
        filter_layout.addSpacing(20)
        
        # Limit label and dropdown
        limit_label = QLabel('Limit:')
        limit_label.setStyleSheet("font-weight: bold; color: #1976d2; margin-right: 10px; font-size: 14px;")
        filter_layout.addWidget(limit_label)
        
        # Limit dropdown
        self.balance_limit_combo = QComboBox()
        self.balance_limit_combo.setObjectName("BalanceLimitCombo")
        self.balance_limit_combo.addItems([
            '6 months',
            '3 months',
            '12 months',
            '18 months',
            'No limit'
        ])
        self.balance_limit_combo.setStyleSheet("""
            QComboBox#BalanceLimitCombo {
                padding: 8px 12px;
                font-size: 14px;
                font-weight: bold;
                border: 2px solid #1976d2;
                border-radius: 4px;
                background-color: white;
                min-width: 120px;
                color: #1976d2;
            }
            QComboBox#BalanceLimitCombo:hover {
                border-color: #0d47a1;
                background-color: #f3e5f5;
            }
        """)
        self.balance_limit_combo.currentTextChanged.connect(self.on_balance_filter_changed)
        filter_layout.addWidget(self.balance_limit_combo)
        
        filter_layout.addStretch()
        
        # Refresh button
        refresh_btn = QPushButton('🔄 Refresh')
        refresh_btn.setObjectName("BalanceRefreshButton")
        refresh_btn.clicked.connect(self.refresh_balance_summary)
        refresh_btn.setStyleSheet("""
            QPushButton#BalanceRefreshButton {
                background-color: #4caf50;
                color: white;
                border: none;
                padding: 8px 16px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 6px;
                min-width: 80px;
            }
            QPushButton#BalanceRefreshButton:hover {
                background-color: #45a049;
            }
            QPushButton#BalanceRefreshButton:pressed {
                background-color: #3d8b40;
            }
        """)
        filter_layout.addWidget(refresh_btn)
        
        # Add to the current position in layout (right after the title)
        self.layout.addWidget(filter_frame)
        
        # Force visibility and layout updates
        filter_frame.show()
        filter_frame.setVisible(True)
        filter_frame.update()
        self.balance_filter_combo.show()
        self.balance_limit_combo.show()
        refresh_btn.show()
        
        # Force layout update
        self.layout.update()
        
        # Make sure parent widget updates
        if hasattr(self, 'parent') and self.parent():
            self.parent().update()

    def clear_tenant_widgets(self):
        print(f"[DEBUG] Starting clear_tenant_widgets, layout count: {self.layout.count()}")
        self.tenant_viewed.emit(False)
        
        # Get the tenant selection frame (should be the first widget after defaults_widget)
        tenant_selection_frame = None
        for i in range(self.layout.count()):
            item = self.layout.itemAt(i)
            widget = item.widget()
            print(f"[DEBUG] Item {i}: widget={widget}, type={type(widget)}")
            if widget and widget != self.defaults_widget and hasattr(widget, 'findChild'):
                # Check if this widget contains our tenant dropdown
                if widget.findChild(QComboBox):
                    tenant_selection_frame = widget
                    print(f"[DEBUG] Found tenant selection frame at index {i}: {tenant_selection_frame}")
                    break
        
        print(f"[DEBUG] Tenant selection frame: {tenant_selection_frame}")
        print(f"[DEBUG] Defaults widget: {self.defaults_widget}")
        
        def _clear_layout(layout):
            """Recursively delete all items from a layout, including nested layouts and widgets."""
            try:
                while layout.count():
                    child = layout.takeAt(0)
                    if child is None:
                        continue
                    w = child.widget()
                    sublayout = child.layout()
                    if sublayout is not None:
                        _clear_layout(sublayout)
                        # After clearing, delete the sublayout's parent widget if exists
                        # Note: Qt will own/delete layouts when parent widget is deleted; just ensure it's detached
                    elif w is not None:
                        print(f"[DEBUG] Deleting widget: {w}")
                        w.setParent(None)
                        w.deleteLater()
                    else:
                        # Spacer item or other item type
                        pass
            except Exception as e:
                print(f"[DEBUG] Error in _clear_layout: {e}")

        # Remove all items except tenant selection and defaults
        for i in reversed(range(self.layout.count())):
            item = self.layout.itemAt(i)
            widget = item.widget()
            sublayout = item.layout() if hasattr(item, 'layout') else None
            spacer = item.spacerItem()

            # Always preserve tenant selection frame and defaults widget
            if widget in [self.defaults_widget, tenant_selection_frame]:
                print(f"[DEBUG] Preserving widget at index {i}: {widget}")
                continue

            # If this is a nested layout (no direct widget), clear it recursively
            if sublayout is not None and not widget:
                print(f"[DEBUG] Recursively clearing sublayout at index {i}")
                _clear_layout(sublayout)
                self.layout.removeItem(item)
                continue

            if widget:
                print(f"[DEBUG] Removing widget at index {i}: {widget}")
                # If the widget has its own layout, clear it first
                try:
                    w_layout = widget.layout()
                    if w_layout is not None:
                        _clear_layout(w_layout)
                except Exception as e:
                    print(f"[DEBUG] Error clearing widget layout: {e}")
                self.layout.removeWidget(widget)
                widget.setParent(None)
                widget.deleteLater()
                continue

            if spacer:
                print(f"[DEBUG] Removing spacer at index {i}")
                self.layout.removeItem(item)
                continue

            print(f"[DEBUG] Unknown item at index {i}: {item}")
        
        print(f"[DEBUG] After clearing, layout count: {self.layout.count()}")
        
        # Clear the tenant_info_label reference
        self.tenant_info_label = None
        # Reset container references that will be rebuilt
        self.rent_buttons_container = None
        if hasattr(self, 'payment_table'):
            try:
                self.payment_table = None
            except Exception:
                pass
        
        # Hide defaults widget
        self.defaults_widget.hide()

    def clear_layout(self):
        self.clear_tenant_widgets()
        self.defaults_widget.show()

    def populate_payment_table(self, tenant):
        """Populate the payment table with individual payments"""
        # Get payment history (store original order)
        payment_history = getattr(tenant, 'payment_history', [])
        print(f"[DEBUG] populate_payment_table: tenant={tenant.name}, payment_history length={len(payment_history)}")
        
        # Clear any existing selection
        self.payment_table.clearSelection()
        self.selected_payment_data = None
        self.selected_payment_row = None
        
        # Store payment data with indices for sorting
        self.payment_data = []
        for i, payment in enumerate(payment_history):
            self.payment_data.append({
                'index': i,  # Original insertion order
                'payment': payment,
                'tenant': tenant
            })
        
        print(f"[DEBUG] populate_payment_table: self.payment_data length={len(self.payment_data)}")
        # Render the table
        self.render_payment_table()
    
    def render_payment_table(self):
        """Render the payment table from stored payment data"""
        if not hasattr(self, 'payment_data'):
            print(f"[DEBUG] render_payment_table: no payment_data attribute")
            return
        
        print(f"[DEBUG] render_payment_table: setting row count to {len(self.payment_data)}")
        self.payment_table.setRowCount(len(self.payment_data))
        
        for i, data in enumerate(self.payment_data):
            payment = data['payment']
            tenant = data['tenant']
            
            # Date received
            date_received = payment.get('date', 'Unknown')
            date_item = QTableWidgetItem(date_received)
            # Store sort value for proper date sorting
            date_item.setData(Qt.ItemDataRole.UserRole, date_received)
            self.payment_table.setItem(i, 0, date_item)
            
            # Amount
            amount = payment.get('amount', 0.0)
            amount_item = QTableWidgetItem(f"${amount:.2f}")
            # Store numeric value for proper sorting
            amount_item.setData(Qt.ItemDataRole.UserRole, amount)
            self.payment_table.setItem(i, 1, amount_item)
            
            # Type
            payment_type = payment.get('type', 'Unknown')
            self.payment_table.setItem(i, 2, QTableWidgetItem(payment_type))
            
            # For Month
            payment_month = payment.get('payment_month', 'Unknown')
            month_item = QTableWidgetItem(payment_month)
            # Store sort value for proper month sorting
            month_item.setData(Qt.ItemDataRole.UserRole, payment_month)
            self.payment_table.setItem(i, 3, month_item)
            
            # Status (human-readable)
            status = self.get_human_readable_status(tenant, payment)
            self.payment_table.setItem(i, 4, QTableWidgetItem(status))
            
            # Details
            details = self.get_payment_details(tenant, payment)
            self.payment_table.setItem(i, 5, QTableWidgetItem(details))
            
            # Notes
            notes = payment.get('notes', '')
            self.payment_table.setItem(i, 6, QTableWidgetItem(notes if notes else ''))
    
    def sort_payment_table(self, column):
        """Sort the payment table by the clicked column"""
        if not hasattr(self, 'payment_data') or not self.payment_data:
            return
        
        # Toggle sort order if clicking same column, otherwise default to ascending
        if self.payment_sort_column == column:
            self.payment_sort_ascending = not self.payment_sort_ascending
        else:
            self.payment_sort_column = column
            self.payment_sort_ascending = True
        
        # Define sort keys for each column
        def get_sort_key(data):
            payment = data['payment']
            
            if column == 0:  # Date Received
                date_str = payment.get('date', '0000-00-00')
                return date_str
            elif column == 1:  # Amount
                return payment.get('amount', 0.0)
            elif column == 2:  # Type
                return payment.get('type', '').lower()
            elif column == 3:  # For Month
                return payment.get('payment_month', '0000-00')
            elif column == 4:  # Status
                status = self.get_human_readable_status(data['tenant'], payment)
                return status.lower()
            elif column == 5:  # Details
                details = self.get_payment_details(data['tenant'], payment)
                return details.lower()
            elif column == 6:  # Notes
                return payment.get('notes', '').lower()
            else:
                return data['index']  # Default to original order
        
        # Sort the data
        self.payment_data.sort(key=get_sort_key, reverse=not self.payment_sort_ascending)
        
        # Re-render the table
        self.render_payment_table()
        
        # Update header to show sort indicator
        self.update_sort_indicator(column)
    
    def update_sort_indicator(self, column):
        """Update the table header to show sort direction indicator"""
        headers = ['Date Received', 'Amount', 'Type', 'For Month', 'Status', 'Month Status', 'Notes']
        
        for i, header in enumerate(headers):
            if i == column:
                arrow = ' ▼' if not self.payment_sort_ascending else ' ▲'
                self.payment_table.setHorizontalHeaderItem(i, QTableWidgetItem(header + arrow))
            else:
                self.payment_table.setHorizontalHeaderItem(i, QTableWidgetItem(header))
    
    def add_balance_summary(self, tenant):
        """Add monthly balance summary section with filtering"""
        # Add the balance summary title first
        balance_label = QLabel('📋 Monthly Balance Summary')
        balance_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #0056b3; margin: 10px 0;")
        self.layout.addWidget(balance_label)
        
        # Add filtering controls right after the title
        self.add_balance_filtering_controls()
        
        # Add the balance summary content
        self.add_balance_summary_content(tenant)
        
        # Scroll to show the newly added balance summary section
        # Use a timer to ensure the layout is updated before scrolling
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(100, self.scroll_to_balance_summary)
    
    def scroll_to_balance_summary(self):
        """Scroll the main scroll area to show the balance summary section"""
        try:
            # Ensure the scroll area exists
            if hasattr(self, 'scroll_area') and self.scroll_area:
                # Don't auto-scroll - let user scroll manually to see all content
                print("[DEBUG] Balance summary added - user can scroll to view")
        except Exception as e:
            print(f"[DEBUG] Error with scroll area: {e}")
    
    def add_month_row(self, layout, month_key, summary, tenant, is_delinquent=False):
        """Add a single month row to the layout"""
        from datetime import datetime
        
        # Create month container
        month_frame = QFrame()
        month_frame.setFrameStyle(QFrame.Shape.StyledPanel)
        
        # Style based on delinquent status
        if is_delinquent:
            border_color = "#dc3545"
            bg_color = "#fff5f5"
        else:
            border_color = "#dee2e6"
            bg_color = "#ffffff"
            
        month_frame.setStyleSheet(f"""
            QFrame {{
                border: 1px solid {border_color};
                border-radius: 6px;
                background-color: {bg_color};
                margin: 2px;
                padding: 5px;
            }}
        """)
        
        month_layout = QVBoxLayout(month_frame)
        month_layout.setContentsMargins(8, 6, 8, 6)
        month_layout.setSpacing(3)
        
        # Month header row
        header_layout = QHBoxLayout()
        
        try:
            year, month = month_key.split('-')
            month_date = datetime(int(year), int(month), 1)
            month_display = month_date.strftime('%B %Y')
        except:
            month_display = month_key
            
        month_label = QLabel(month_display)
        month_label.setStyleSheet("font-weight: bold; font-size: 13px; color: #495057;")
        
        # Status with color coding
        status_text = summary['status']
        status_color = self.get_status_color(summary['status'])
        status_label = QLabel(status_text)
        status_label.setStyleSheet(f"color: {status_color}; font-weight: bold; font-size: 12px;")
        
        header_layout.addWidget(month_label)
        header_layout.addStretch()
        header_layout.addWidget(status_label)
        
        month_layout.addLayout(header_layout)
        
        # Details row
        details_layout = QHBoxLayout()
        
        rent_due_label = QLabel(f"Due: ${summary['rent_due']:.2f}")
        rent_due_label.setStyleSheet("color: #6c757d; font-size: 11px;")
        
        paid_label = QLabel(f"Paid: ${summary['total_paid']:.2f}")
        paid_label.setStyleSheet("color: #28a745; font-size: 11px;")
        
        balance_color = "#dc3545" if summary['balance'] > 0 else "#28a745" if summary['balance'] < 0 else "#6c757d"
        balance_text = f"Balance: ${summary['balance']:.2f}"
        if summary['balance'] < 0:
            balance_text = f"Credit: ${abs(summary['balance']):.2f}"
        
        balance_label = QLabel(balance_text)
        balance_label.setStyleSheet(f"color: {balance_color}; font-size: 11px; font-weight: bold;")
        
        details_layout.addWidget(rent_due_label)
        details_layout.addWidget(paid_label)
        details_layout.addStretch()
        details_layout.addWidget(balance_label)
        
        month_layout.addLayout(details_layout)
        
        # Check for pending actions in queue for this month
        try:
            year, month = month_key.split('-')
            pending_actions = self.get_pending_actions_for_month(tenant, int(year), int(month))
            
            if pending_actions:
                queue_layout = QHBoxLayout()
                queue_icon = QLabel("📋")
                queue_icon.setStyleSheet("color: #17a2b8; font-size: 12px;")
                
                queue_texts = []
                for action in pending_actions:
                    action_type = action['type']
                    if action_type == 'rent_change':
                        new_rent = action['action_data'].get('new_rent', 0)
                        queue_texts.append(f"Rent → ${new_rent:.2f}")
                    elif action_type == 'rental_period_change':
                        queue_texts.append("Lease renewal")
                    elif action_type == 'lease_expiry':
                        queue_texts.append("Lease expires")
                    else:
                        queue_texts.append(action['description'])
                
                queue_text = ", ".join(queue_texts)
                queue_label = QLabel(f"Queued: {queue_text}")
                queue_label.setStyleSheet("color: #17a2b8; font-size: 10px; font-style: italic;")
                
                queue_layout.addWidget(queue_icon)
                queue_layout.addWidget(queue_label)
                queue_layout.addStretch()
                
                month_layout.addLayout(queue_layout)
        except:
            pass  # Skip queue info if there's an error
        layout.addWidget(month_frame)
    
    def get_human_readable_status(self, tenant, payment):
        """Get human-readable status for a payment"""
        payment_month = payment.get('payment_month', '')
        if not payment_month:
            return 'Unknown'
        
        try:
            year, month = payment_month.split('-')
            year, month = int(year), int(month)
            
            # Get expected rent for that month
            expected_rent = self.rent_tracker.get_effective_rent(tenant, year, month)
            
            # Get total paid for that month
            total_paid = self.get_total_paid_for_month(tenant, year, month)
            
            if total_paid >= expected_rent:
                if total_paid > expected_rent:
                    return "Overpayment"
                else:
                    return "Paid in Full"
            elif total_paid > 0:
                return "Partial Payment"
            else:
                return "No Payment"
                
        except Exception:
            return 'Unknown'
    
    def get_payment_details(self, tenant, payment):
        """Get detailed information about a payment"""
        payment_month = payment.get('payment_month', '')
        amount = payment.get('amount', 0.0)
        
        if not payment_month:
            return 'No details available'
        
        try:
            year, month = payment_month.split('-')
            year, month = int(year), int(month)
            
            expected_rent = self.rent_tracker.get_effective_rent(tenant, year, month)
            total_paid = self.get_total_paid_for_month(tenant, year, month)
            
            if total_paid > expected_rent:
                overpayment = total_paid - expected_rent
                return f"Overpaid by ${overpayment:.2f}"
            elif total_paid < expected_rent:
                shortage = expected_rent - total_paid
                return f"${shortage:.2f} remaining"
            else:
                return "Fully paid"
                
        except Exception:
            return 'Details unavailable'
    
    def get_total_paid_for_month(self, tenant, year, month):
        """Calculate total paid for a specific month"""
        payment_history = getattr(tenant, 'payment_history', [])
        total = 0.0
        
        print(f"[DEBUG] get_total_paid_for_month: Looking for payments for {year}-{month:02d}")
        print(f"[DEBUG] get_total_paid_for_month: Checking {len(payment_history)} payments")
        
        for i, payment in enumerate(payment_history):
            pay_year = payment.get('year')
            pay_month = payment.get('month')
            pay_amount = payment.get('amount', 0.0)
            payment_month_str = payment.get('payment_month', 'N/A')
            
            print(f"[DEBUG] Payment {i}: year={pay_year}, month={pay_month}, amount=${pay_amount}, payment_month={payment_month_str}")
            
            if pay_year == year and pay_month == month:
                print(f"[DEBUG]   -> MATCH! Adding ${pay_amount}")
                total += pay_amount
            else:
                print(f"[DEBUG]   -> NO MATCH (checking {pay_year}=={year} and {pay_month}=={month})")
        
        print(f"[DEBUG] get_total_paid_for_month({year}-{month:02d}): FINAL TOTAL = ${total:.2f}")
        return total

    def get_pending_actions_for_month(self, tenant, year, month):
        """Get pending actions from the queue for a specific month"""
        if not hasattr(self.rent_tracker, 'action_queue') or not self.rent_tracker.action_queue:
            return []
        
        actions = []
        month_key = f"{year}-{month:02d}"
        
        # Check all pending actions in the queue
        for action in self.rent_tracker.action_queue.get_pending_actions():
            if action.get('tenant_id') == tenant.tenant_id:
                scheduled_date = action.get('scheduled_date', '')
                
                # Check if action is scheduled for this month
                if scheduled_date.startswith(month_key):
                    action_type = action.get('action_type', '')
                    description = action.get('description', '')
                    
                    if action_type in ['rent_change', 'rental_period_change', 'lease_expiry']:
                        actions.append({
                            'type': action_type,
                            'description': description,
                            'scheduled_date': scheduled_date,
                            'action_data': action.get('action_data', {})
                        })
        
        return actions

    def get_monthly_summaries(self, tenant):
        """Get monthly balance summaries for relevant months only"""
        from datetime import date
        summaries = {}
        payment_history = getattr(tenant, 'payment_history', [])
        
        # Get current date for filtering
        today = date.today()
        current_month_tuple = (today.year, today.month)
        
        # Get rental period boundaries
        rental_start_date = None
        rental_end_date = None
        if hasattr(tenant, 'rental_period') and tenant.rental_period:
            try:
                if isinstance(tenant.rental_period, dict):
                    # New format: {"start_date": "2025-09-11", "end_date": "2026-09-11", "lease_type": "Fixed Term"}
                    start_date_str = tenant.rental_period.get('start_date')
                    end_date_str = tenant.rental_period.get('end_date')
                elif isinstance(tenant.rental_period, (list, tuple)) and len(tenant.rental_period) >= 2:
                    # Old format: ["2025-09-11", "2026-09-11"] or ("2025-09-11", "2026-09-11")
                    start_date_str = tenant.rental_period[0]
                    end_date_str = tenant.rental_period[1]
                else:
                    start_date_str = end_date_str = None
                
                if start_date_str and end_date_str:
                    rental_start_date = date.fromisoformat(start_date_str)
                    rental_end_date = date.fromisoformat(end_date_str)
            except:
                pass  # Fall back to no restrictions if parsing fails
        
        # Calculate reasonable range, but respect rental period
        months_to_include = set()
        
        # Add months from payment history (always include if there are payments)
        for payment in payment_history:
            year = payment.get('year')
            month = payment.get('month')
            if year and month:
                months_to_include.add((year, month))
        
        # Add months from tenant's months_to_charge (these are within rental period by definition)
        if hasattr(tenant, 'months_to_charge'):
            for year_month in tenant.months_to_charge:
                year, month = year_month
                month_tuple = (year, month)
                
                # Only include if within reasonable display range
                # Current month +/- 3 months, but respect that we don't want to show too far in future
                if (month_tuple[0] == current_month_tuple[0] and 
                    abs(month_tuple[1] - current_month_tuple[1]) <= 3) or \
                   (abs(month_tuple[0] - current_month_tuple[0]) <= 1):
                    months_to_include.add(month_tuple)  # Add the tuple, not the list
        
        # Add reasonable range around current month, but only within rental period
        current_date = today.replace(day=1)
        
        # Add months back (only within rental period)
        temp_date = current_date
        for _ in range(3):
            if temp_date.month == 1:
                temp_date = temp_date.replace(year=temp_date.year - 1, month=12)
            else:
                temp_date = temp_date.replace(month=temp_date.month - 1)
            
            # Only include if within rental period
            if rental_start_date is None or temp_date >= rental_start_date.replace(day=1):
                months_to_include.add((temp_date.year, temp_date.month))
        
        # Add current month (if within rental period)
        if (rental_start_date is None or current_date >= rental_start_date.replace(day=1)) and \
           (rental_end_date is None or current_date <= rental_end_date.replace(day=1)):
            months_to_include.add((current_date.year, current_date.month))
        
        # Add months forward (only within rental period)
        temp_date = current_date
        for _ in range(3):
            if temp_date.month == 12:
                temp_date = temp_date.replace(year=temp_date.year + 1, month=1)
            else:
                temp_date = temp_date.replace(month=temp_date.month + 1)
            
            # Only include if within rental period
            if rental_end_date is None or temp_date <= rental_end_date.replace(day=1):
                months_to_include.add((temp_date.year, temp_date.month))
        
        # Create summaries for filtered months
        for year, month in sorted(months_to_include, reverse=True):
            month_key = f"{year}-{month:02d}"
            
            expected_rent = self.rent_tracker.get_effective_rent(tenant, year, month)
            total_paid = self.get_total_paid_for_month(tenant, year, month)
            balance = expected_rent - total_paid
            
            # Only include if there's meaningful data or it's a relevant timeframe
            month_tuple = (year, month)
            month_date = date(year, month, 1)
            
            # Check if month is within rental period
            within_rental_period = True
            if rental_start_date and month_date < rental_start_date.replace(day=1):
                within_rental_period = False
            if rental_end_date and month_date > rental_end_date.replace(day=1):
                within_rental_period = False
            
            is_relevant = (
                (total_paid > 0 or balance > 0) and within_rental_period  # Has activity and within period
            ) or (
                within_rental_period and (
                    month_tuple >= current_month_tuple or  # Current or future within period
                    (current_month_tuple[0] - month_tuple[0]) * 12 + (current_month_tuple[1] - month_tuple[1]) <= 3  # Recent past within period
                )
            )
            
            if is_relevant:
                # Check if this month is in the future
                month_tuple = (year, month)
                today = date.today()
                current_month_tuple = (today.year, today.month)
                
                # Calculate next month
                if current_month_tuple[1] == 12:
                    next_month_tuple = (current_month_tuple[0] + 1, 1)
                else:
                    next_month_tuple = (current_month_tuple[0], current_month_tuple[1] + 1)
                
                if month_tuple > current_month_tuple:
                    # Future month
                    if month_tuple == next_month_tuple:
                        status = "Due Soon"
                    else:
                        status = "Not Due"
                elif total_paid >= expected_rent:
                    if total_paid > expected_rent:
                        status = "Overpayment"
                    else:
                        status = "Paid in Full"
                elif total_paid > 0:
                    status = "Partial Payment"
                else:
                    status = "Not Paid"
                
                summaries[month_key] = {
                    'status': status,
                    'rent_due': expected_rent,
                    'total_paid': total_paid,
                    'balance': balance
                }
        
        return summaries
    
    def get_status_color(self, status):
        """Get color for status display"""
        colors = {
            'Paid in Full': '#28a745',
            'Overpayment': '#17a2b8',
            'Partial Payment': '#ffc107',
            'Not Paid': '#dc3545',
            'Due Soon': '#ff9800',
            'Not Due': '#6c757d',
            'Unknown': '#6c757d'
        }
        return colors.get(status, '#6c757d')

    def on_balance_filter_changed(self):
        """Handle balance filter dropdown change"""
        if hasattr(self, 'selected_tenant') and self.selected_tenant:
            self.refresh_balance_summary()
    
    def refresh_balance_summary(self):
        """Refresh the balance summary with current filter"""
        if hasattr(self, 'selected_tenant') and self.selected_tenant:
            # Store the current filter selection
            current_filter = None
            if hasattr(self, 'balance_filter_combo') and self.balance_filter_combo:
                current_filter = self.balance_filter_combo.currentText()
            
            # Clear only the balance summary portion (not the filter controls)
            self.clear_balance_summary_only()
            
            # Re-add just the balance summary
            self.add_balance_summary_content(self.selected_tenant)
            
            # Restore filter selection if it was set
            if current_filter and hasattr(self, 'balance_filter_combo') and self.balance_filter_combo:
                self.balance_filter_combo.setCurrentText(current_filter)
        
        # Also refresh the comprehensive analysis tab if it exists
        if hasattr(self, 'comprehensive_analysis_tab') and self.comprehensive_analysis_tab:
            try:
                self.comprehensive_analysis_tab.refresh_all()
            except Exception as e:
                print(f"[DEBUG] Failed to refresh comprehensive analysis tab: {e}")
    
    def clear_balance_summary_only(self):
        """Clear only the balance summary content, not the filter controls"""
        # Find and remove only the balance summary widgets (labels and scroll areas)
        items_to_remove = []
        for i in range(self.layout.count()):
            item = self.layout.itemAt(i)
            if item and item.widget():
                widget = item.widget()
                # Only remove if it's a balance-related widget but NOT the filter combo
                if (isinstance(widget, QScrollArea) or 
                    (isinstance(widget, QLabel) and 
                     hasattr(widget, 'text') and 
                     '📋 Monthly Balance Summary' in widget.text())):
                    items_to_remove.append((i, widget))
        
        # Remove widgets in reverse order to maintain indices
        for i, widget in reversed(items_to_remove):
            self.layout.removeWidget(widget)
            widget.deleteLater()
    
    def add_balance_summary_content(self, tenant):
        """Add just the balance summary content without the title or filter controls"""
        from datetime import date
        
        # Create scroll area for monthly summaries (no title - that's added by parent method)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFixedHeight(250)  # Fixed height for consistent display
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: 2px solid #dee2e6;
                border-radius: 8px;
                background-color: #f8f9fa;
            }
            QScrollBar:vertical {
                background-color: #f8f9fa;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background-color: #6c757d;
                border-radius: 6px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #495057;
            }
        """)
        
        # Create balance summary widget
        balance_widget = QWidget()
        balance_layout = QVBoxLayout(balance_widget)
        balance_layout.setContentsMargins(15, 10, 15, 10)
        balance_layout.setSpacing(8)
        
        # Get filtered monthly summaries
        monthly_summaries = self.get_filtered_monthly_summaries(tenant)
        
        if not monthly_summaries:
            no_data_label = QLabel("No data matches current filter")
            no_data_label.setStyleSheet("color: #6c757d; font-style: italic; text-align: center; padding: 20px;")
            balance_layout.addWidget(no_data_label)
        else:
            # Get current date for organizing
            today = date.today()
            current_month_tuple = (today.year, today.month)
            
            # Filter and organize summaries
            delinquent_months = []
            current_future_months = []
            
            for month_key, summary in monthly_summaries.items():
                try:
                    year, month = month_key.split('-')
                    year, month = int(year), int(month)
                    month_tuple = (year, month)
                    
                    is_delinquent = False
                    
                    # Check if balance is owed
                    if summary['balance'] > 0:
                        if month_tuple < current_month_tuple:
                            # Past month with balance due = delinquent
                            is_delinquent = True
                        elif month_tuple == current_month_tuple:
                            # Current month - check if rent due date has passed
                            if hasattr(tenant, 'rent_due_date') and tenant.rent_due_date:
                                try:
                                    # Check if today is past the due date in current month
                                    due_day = int(tenant.rent_due_date)
                                    today_day = today.day
                                    
                                    if today_day > due_day:
                                        # Past due date in current month
                                        is_delinquent = True
                                except:
                                    pass
                    
                    if is_delinquent:
                        delinquent_months.append((month_key, summary))
                    else:
                        current_future_months.append((month_key, summary))
                        
                except:
                    # Include if we can't parse the date (safety fallback)
                    current_future_months.append((month_key, summary))
            
            # Add delinquent months section (if any)
            if delinquent_months:
                delinquent_header = QLabel("⚠️ Past Due Months:")
                delinquent_header.setStyleSheet("""
                    font-weight: bold; 
                    color: #dc3545; 
                    font-size: 13px;
                    background-color: #f8d7da;
                    padding: 5px;
                    border-radius: 4px;
                    margin: 5px 0;
                """)
                balance_layout.addWidget(delinquent_header)
                
                for month_key, summary in sorted(delinquent_months):
                    self.add_month_row(balance_layout, month_key, summary, tenant, is_delinquent=True)
                
                # Add separator
                separator = QLabel()
                separator.setFixedHeight(1)
                separator.setStyleSheet("background-color: #dee2e6; margin: 10px 0;")
                balance_layout.addWidget(separator)
            
            # Add current/future months section
            if current_future_months:
                current_header = QLabel("📅 Current & Other Months:")
                current_header.setStyleSheet("""
                    font-weight: bold; 
                    color: #0056b3; 
                    font-size: 13px;
                    background-color: #d1ecf1;
                    padding: 5px;
                    border-radius: 4px;
                    margin: 5px 0;
                """)
                balance_layout.addWidget(current_header)
                
                for month_key, summary in sorted(current_future_months):
                    self.add_month_row(balance_layout, month_key, summary, tenant, is_delinquent=False)
        
        balance_widget.setStyleSheet("""
            QWidget {
                background-color: white; 
                padding: 5px;
            }
        """)
        
        scroll_area.setWidget(balance_widget)
        self.layout.addWidget(scroll_area)
        
        # Add some spacing after the balance summary
        self.layout.addSpacing(10)
    
    def clear_balance_summary(self):
        """Clear the existing balance summary widgets"""
        # Clear the balance filter combo references to allow recreation, but preserve the widgets themselves
        print("[DEBUG] Balance summary widgets cleared")
        
        # Find and remove widgets by looking for specific widget types and styles
        # BUT preserve the filter controls
        items_to_remove = []
        for i in range(self.layout.count()):
            item = self.layout.itemAt(i)
            if item and item.widget():
                widget = item.widget()
                
                # Look for balance summary related widgets but NOT filter controls
                should_remove = False
                
                # Only remove scroll areas (balance content) and balance title labels
                if isinstance(widget, QScrollArea):
                    should_remove = True
                elif isinstance(widget, QLabel):
                    text = widget.text()
                    if '📋 Monthly Balance Summary' in text:
                        should_remove = True
                
                # Do NOT remove filter frame (check object name)
                if hasattr(widget, 'objectName'):
                    obj_name = widget.objectName()
                    if obj_name == "BalanceFilterFrame":
                        should_remove = False  # Preserve filter controls
                
                if should_remove:
                    items_to_remove.append((i, widget))
        
        # Remove widgets in reverse order to maintain indices
        for i, widget in reversed(items_to_remove):
            self.layout.removeWidget(widget)
            widget.deleteLater()
        
        print(f"[DEBUG] Removed {len(items_to_remove)} balance summary widgets, preserved filter controls")
    
    def get_filtered_monthly_summaries(self, tenant):
        """Get monthly summaries based on current filter selection"""
        if not hasattr(self, 'balance_filter_combo'):
            # If no filter combo exists yet, use default filtering
            return self.get_monthly_summaries(tenant)
        
        filter_option = self.balance_filter_combo.currentText()
        limit_option = self.balance_limit_combo.currentText() if hasattr(self, 'balance_limit_combo') else '6 months'
        
        print(f"[DEBUG] Filter option: {filter_option}")
        print(f"[DEBUG] Limit option: {limit_option}")
        
        # Parse limit
        if limit_option == 'No limit':
            month_limit = None
        else:
            month_limit = int(limit_option.split()[0])
        
        all_summaries = self.get_monthly_summaries_all(tenant)  # Get all possible summaries
        print(f"[DEBUG] All summaries found: {list(all_summaries.keys())}")
        
        from datetime import date
        today = date.today()
        current_month_tuple = (today.year, today.month)
        
        filtered_summaries = {}
        
        for month_key, summary in all_summaries.items():
            year, month = map(int, month_key.split('-'))
            month_tuple = (year, month)
            
            include_month = False
            
            if filter_option == 'Current + Delinquent + Near Future':
                # Smart default: Current month, delinquent months, limited future months with balance
                if month_tuple == current_month_tuple:
                    include_month = True
                elif month_tuple < current_month_tuple and summary['balance'] > 0:
                    include_month = True  # Delinquent
                elif month_tuple > current_month_tuple and summary['balance'] > 0:
                    # Future with balance - apply limit
                    if month_limit is None:
                        include_month = True
                    else:
                        months_ahead = (month_tuple[0] - current_month_tuple[0]) * 12 + (month_tuple[1] - current_month_tuple[1])
                        if months_ahead <= month_limit:
                            include_month = True
                    
            elif filter_option == 'Current Month Only':
                if month_tuple == current_month_tuple:
                    include_month = True
                    
            elif filter_option == 'Delinquent Months Only':
                if month_tuple < current_month_tuple and summary['balance'] > 0:
                    include_month = True
                    
            elif filter_option == 'Future Months Only':
                if month_tuple > current_month_tuple:
                    if month_limit is None:
                        include_month = True
                    else:
                        months_ahead = (month_tuple[0] - current_month_tuple[0]) * 12 + (month_tuple[1] - current_month_tuple[1])
                        if months_ahead <= month_limit:
                            include_month = True
                    
            elif filter_option == 'Past Months Only':
                if month_tuple < current_month_tuple:
                    if month_limit is None:
                        include_month = True
                    else:
                        months_behind = (current_month_tuple[0] - month_tuple[0]) * 12 + (current_month_tuple[1] - month_tuple[1])
                        if months_behind <= month_limit:
                            include_month = True
                    
            elif filter_option == 'All Relevant Months':
                if month_limit is None:
                    include_month = True
                else:
                    # Apply limit in both directions from current month
                    months_diff = abs((month_tuple[0] - current_month_tuple[0]) * 12 + (month_tuple[1] - current_month_tuple[1]))
                    if months_diff <= month_limit:
                        include_month = True
                
            elif filter_option == 'Months with Payments Only':
                if summary['total_paid'] > 0:
                    if month_limit is None:
                        include_month = True
                    else:
                        months_diff = abs((month_tuple[0] - current_month_tuple[0]) * 12 + (month_tuple[1] - current_month_tuple[1]))
                        if months_diff <= month_limit:
                            include_month = True
            
            if include_month:
                filtered_summaries[month_key] = summary
                print(f"[DEBUG] {month_key}: Including")
            else:
                print(f"[DEBUG] {month_key}: Excluding")
        
        print(f"[DEBUG] Filtered summaries: {list(filtered_summaries.keys())}")
        return filtered_summaries
    
    def get_monthly_summaries_all(self, tenant):
        """Get all possible monthly summaries without filtering restrictions"""
        from datetime import date
        summaries = {}
        payment_history = getattr(tenant, 'payment_history', [])
        
        # Get broader range for "all" summaries
        today = date.today()
        
        # Get rental period boundaries
        rental_start_date = None
        rental_end_date = None
        if hasattr(tenant, 'rental_period') and tenant.rental_period:
            try:
                if isinstance(tenant.rental_period, dict):
                    start_date_str = tenant.rental_period.get('start_date')
                    end_date_str = tenant.rental_period.get('end_date')
                elif isinstance(tenant.rental_period, (list, tuple)) and len(tenant.rental_period) >= 2:
                    start_date_str = tenant.rental_period[0]
                    end_date_str = tenant.rental_period[1]
                else:
                    start_date_str = end_date_str = None
                
                if start_date_str and end_date_str:
                    rental_start_date = date.fromisoformat(start_date_str)
                    rental_end_date = date.fromisoformat(end_date_str)
            except:
                pass
        
        # Include months from payment history
        months_to_include = set()
        for payment in payment_history:
            year = payment.get('year')
            month = payment.get('month')
            if year and month:
                months_to_include.add((year, month))
        
        # Include months from rental period
        if rental_start_date and rental_end_date:
            current_date = rental_start_date.replace(day=1)
            end_date = rental_end_date.replace(day=1)
            
            while current_date <= end_date:
                months_to_include.add((current_date.year, current_date.month))
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)
        
        # Generate summaries
        for year, month in months_to_include:
            month_key = f"{year}-{month:02d}"
            expected_rent = self.rent_tracker.get_effective_rent(tenant, year, month)
            total_paid = self.get_total_paid_for_month(tenant, year, month)
            balance = expected_rent - total_paid
            
            # Check if this month is in the future
            month_tuple = (year, month)
            today = date.today()
            current_month_tuple = (today.year, today.month)
            
            # Calculate next month
            if current_month_tuple[1] == 12:
                next_month_tuple = (current_month_tuple[0] + 1, 1)
            else:
                next_month_tuple = (current_month_tuple[0], current_month_tuple[1] + 1)
            
            if month_tuple > current_month_tuple:
                # Future month
                if month_tuple == next_month_tuple:
                    status = "Due Soon"
                else:
                    status = "Not Due"
            elif total_paid >= expected_rent:
                if total_paid > expected_rent:
                    status = "Overpayment"
                else:
                    status = "Paid in Full"
            elif total_paid > 0:
                status = "Partial Payment"
            else:
                status = "Not Paid"
            
            summaries[month_key] = {
                'status': status,
                'rent_due': expected_rent,
                'total_paid': total_paid,
                'balance': balance
            }
        
        return summaries

    def export_rent_management_to_csv(self):
        """Export the currently shown tenant's rent management data to Excel or CSV file."""
        if not getattr(self, 'selected_tenant', None):
            QMessageBox.warning(self, 'Export Rent Management', 'Please select a tenant first.')
            return

        tenant = self.selected_tenant

        # Ask user how to sort payments
        sort_dialog = QMessageBox(self)
        sort_dialog.setWindowTitle('Sort Payments By')
        sort_dialog.setText('How would you like to sort the payment history?')
        sort_dialog.setIcon(QMessageBox.Icon.Question)
        
        date_received_btn = sort_dialog.addButton('Date Received', QMessageBox.ButtonRole.AcceptRole)
        for_month_btn = sort_dialog.addButton('Month Payment Is For', QMessageBox.ButtonRole.AcceptRole)
        cancel_btn = sort_dialog.addButton('Cancel', QMessageBox.ButtonRole.RejectRole)
        
        sort_dialog.exec()
        
        clicked_btn = sort_dialog.clickedButton()
        if clicked_btn == cancel_btn or clicked_btn is None:
            return
        
        sort_by = 'date_received' if clicked_btn == date_received_btn else 'payment_month'

        # Create file dialog with both options
        if OPENPYXL_AVAILABLE:
            file_filter = "Excel Files (*.xlsx);;CSV Files (*.csv)"
        else:
            file_filter = "CSV Files (*.csv)"
        
        suggested = f"{tenant.name.replace(' ', '_')}_rent_export"
        path, selected_filter = QFileDialog.getSaveFileName(
            self, 
            'Export Rent Management', 
            suggested,
            file_filter
        )
        
        if not path:
            return

        # Determine format based on selected filter or file extension
        if OPENPYXL_AVAILABLE and 'Excel' in selected_filter:
            if not path.endswith('.xlsx'):
                path = path.rsplit('.', 1)[0] + '.xlsx' if '.' in path else path + '.xlsx'
            self.export_to_excel(tenant, path, sort_by)
        else:
            if not path.endswith('.csv'):
                path = path.rsplit('.', 1)[0] + '.csv' if '.' in path else path + '.csv'
            self.export_to_csv_plain(tenant, path, sort_by)

    def _calculate_overpayment_created(self, tenant, payment):
        """Calculate if this payment created an overpayment and return the amount"""
        try:
            payment_month = payment.get('payment_month', '')
            if not payment_month:
                return 0.0
            
            # Don't count credit usage as creating overpayment
            if payment.get('is_credit_usage', False):
                return 0.0
            
            year, month = payment_month.split('-')
            year, month = int(year), int(month)
            
            # Get expected rent for that month
            expected_rent = self.rent_tracker.get_effective_rent(tenant, year, month)
            
            # Calculate total paid for that month up to and including this payment
            total_paid = 0.0
            payment_date = payment.get('date', '')
            
            for p in getattr(tenant, 'payment_history', []):
                if p.get('payment_month') == payment_month:
                    # Only include payments up to this payment's date
                    if p.get('date', '') <= payment_date:
                        total_paid += p.get('amount', 0.0)
            
            # If total paid exceeds expected rent, calculate overpayment
            if total_paid > expected_rent:
                # Check if previous payments already caused overpayment
                previous_total = total_paid - payment.get('amount', 0.0)
                if previous_total >= expected_rent:
                    # Previous payments already covered it, no new overpayment
                    return 0.0
                else:
                    # This payment created overpayment
                    return total_paid - expected_rent
            
            return 0.0
        except Exception:
            return 0.0
    
    def export_to_excel(self, tenant, path, sort_by='date_received'):
        """Export rent management data to Excel with formatting and colors"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Main"

            # Define colors and styles
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            section_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            section_font = Font(bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            credit_usage_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow for credit usage rows
            
            # Color mapping for statuses
            status_colors = {
                'Paid in Full': "C6EFCE",      # Green
                'Overpayment': "92D050",       # Light Green
                'Partial Payment': "FFEB9C",   # Yellow
                'Not Paid': "FFC7CE",          # Red
                'Due Soon': "FFD580",          # Orange (not red)
                'Not Due': "E2EFDA",           # Light Green (not gray, to indicate no delinquency)
                'Delinquent': "FF0000",        # Dark Red
                'N/A': "CCCCCC"                # Gray
            }
            
            # Color mapping for months (alternating light colors for easy differentiation)
            month_colors = {}
            month_color_cycle = ["E8F4F8", "F0E8F8", "F8F4E8", "E8F8F0", "F8E8E8"]  # Light blue, light purple, light tan, light green, light pink
            month_index = 0
            for payment in getattr(tenant, 'payment_history', []) or []:
                pmonth = payment.get('payment_month', '')
                if pmonth and pmonth not in month_colors:
                    month_colors[pmonth] = month_color_cycle[month_index % len(month_color_cycle)]
                    month_index += 1
            
            # Color mapping for details
            details_colors = {
                'Overpaid': "92D050",           # Light Green
                'remaining': "FFD580",          # Orange
                'Fully paid': "C6EFCE"          # Green
            }

            row = 1
            
            # Add report generation date at the top
            from datetime import datetime
            ws.merge_cells(f'A{row}:H{row}')
            date_cell = ws[f'A{row}']
            date_cell.value = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            date_cell.font = Font(bold=True, size=11)
            date_cell.alignment = Alignment(horizontal='left')
            row += 1
            
            # Add disclaimer at the top
            disclaimer_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws.merge_cells(f'A{row}:H{row}')
            disclaimer_cell = ws[f'A{row}']
            disclaimer_cell.value = "ℹ️ NOTE: Rows highlighted in light yellow are Overpayment Credit or Service Credit usage. These do NOT represent new rent payments and are not included in total payment calculations, as they were already counted when the original overpayment occurred."
            disclaimer_cell.fill = disclaimer_fill
            disclaimer_cell.font = Font(italic=True, size=9)
            disclaimer_cell.alignment = Alignment(wrap_text=True, vertical='center')
            ws.row_dimensions[row].height = 40
            row += 2

            # Helper function to add section header
            def add_section_header(title):
                nonlocal row
                ws.merge_cells(f'A{row}:H{row}')
                cell = ws[f'A{row}']
                cell.value = title
                cell.fill = section_fill
                cell.font = section_font
                row += 1

            # Helper function to add row with formatting
            def add_row(label, value, width=25):
                nonlocal row
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1

            # ===== TENANT INFORMATION SECTION =====
            add_section_header("TENANT INFORMATION")
            add_row("Name", tenant.name)
            add_row("Tenant ID", tenant.tenant_id)
            add_row("Rental Period", self.format_rental_period(tenant.rental_period))
            add_row("Rent Amount", tenant.rent_amount)
            ws[f'B{row-1}'].number_format = '$#,##0.00'
            add_row("Deposit", tenant.deposit_amount)
            ws[f'B{row-1}'].number_format = '$#,##0.00'
            add_row("Due Day", getattr(tenant, 'rent_due_date', '') or 'N/A')
            add_row("Contact", self.format_contact_info(tenant.contact_info))
            if isinstance(getattr(tenant, '_notes_list', None), list):
                notes_text = '; '.join(tenant._notes_list)
            else:
                notes_text = tenant.notes or ''
            add_row("Notes", notes_text)
            row += 1

            # ===== CURRENT STATUS SECTION =====
            add_section_header("CURRENT STATUS & CALCULATIONS")
            from datetime import date
            today = date.today()
            current_month_key = f"{today.year}-{today.month:02d}"
            
            # Get effective rent which includes account status check
            effective_rent = self.rent_tracker.get_effective_rent(tenant, today.year, today.month)
            
            # Get override info for display
            monthly_override = None
            yearly_override = None
            if hasattr(tenant, 'monthly_exceptions') and tenant.monthly_exceptions:
                monthly_override = tenant.monthly_exceptions.get(current_month_key)
                if monthly_override is None:
                    yearly_override = tenant.monthly_exceptions.get(str(today.year))

            delinquency_balance = float(getattr(tenant, 'delinquency_balance', 0.0) or 0.0)
            overpayment_credit = float(getattr(tenant, 'overpayment_credit', 0.0) or 0.0)
            service_credit = float(getattr(tenant, 'service_credit', 0.0) or 0.0)
            net_due = delinquency_balance - overpayment_credit

            add_row("Current Month", today.strftime('%B %Y'))
            add_row("Base Rent", tenant.rent_amount)
            ws[f'B{row-1}'].number_format = '$#,##0.00'
            add_row("Active Override", f"{monthly_override if monthly_override is not None else yearly_override if yearly_override is not None else 'None'}")
            if monthly_override is not None or yearly_override is not None:
                ws[f'B{row-1}'].number_format = '$#,##0.00'
            add_row("Effective Rent", effective_rent)
            ws[f'B{row-1}'].number_format = '$#,##0.00'
            
            add_row("Delinquency Balance", delinquency_balance)
            delinq_cell = ws[f'B{row-1}']
            delinq_cell.number_format = '$#,##0.00'
            if delinquency_balance > 0:
                delinq_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            add_row("Overpayment Credit", overpayment_credit)
            ws[f'B{row-1}'].number_format = '$#,##0.00'
            ws[f'B{row-1}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            add_row("Service Credit", service_credit)
            ws[f'B{row-1}'].number_format = '$#,##0.00'
            
            add_row("Net Amount Due", abs(net_due) if net_due != 0 else 0.00)
            net_cell = ws[f'B{row-1}']
            net_cell.number_format = '$#,##0.00'
            if net_due > 0:
                net_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif net_due < 0:
                net_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            row += 1

            # ===== PAYMENT HISTORY SECTION =====
            add_section_header(f"PAYMENT HISTORY (Sorted by {('Date Received' if sort_by == 'date_received' else 'Payment Month')})")
            
            # Sort payments based on user selection
            payments = getattr(tenant, 'payment_history', []) or []
            if sort_by == 'date_received':
                payments = sorted(
                    payments,
                    key=lambda p: datetime.strptime(p.get('date', '1970-01-01'), '%Y-%m-%d')
                )
            else:  # sort_by == 'payment_month'
                payments = sorted(
                    payments,
                    key=lambda p: (p.get('payment_month', '9999-99'), datetime.strptime(p.get('date', '1970-01-01'), '%Y-%m-%d'))
                )

            # Add column headers
            headers = ['Date Received', 'Amount', 'Type', 'For Month', 'Status', 'Month Status', 'Overpayment Created', 'Notes']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            row += 1

            # Add payment rows
            for payment in payments:
                date_received = payment.get('date', '')
                amount = float(payment.get('amount', 0.0))
                ptype = payment.get('type', '')
                pmonth = payment.get('payment_month', '')
                status = self.get_human_readable_status(tenant, payment)
                details = self.get_payment_details(tenant, payment)
                notes = payment.get('notes', '')
                is_credit_usage = payment.get('is_credit_usage', False)
                overpayment_created = self._calculate_overpayment_created(tenant, payment)

                # Add values
                ws.cell(row=row, column=1).value = datetime.strptime(date_received, '%Y-%m-%d') if date_received else None
                ws.cell(row=row, column=1).number_format = 'MM/DD/YYYY'
                
                ws.cell(row=row, column=2).value = amount
                ws.cell(row=row, column=2).number_format = '$#,##0.00'
                
                ws.cell(row=row, column=3).value = ptype
                ws.cell(row=row, column=4).value = pmonth
                ws.cell(row=row, column=5).value = status
                ws.cell(row=row, column=6).value = details
                ws.cell(row=row, column=7).value = overpayment_created if overpayment_created > 0 else ''
                ws.cell(row=row, column=7).number_format = '$#,##0.00'
                ws.cell(row=row, column=8).value = notes

                # Check if this is a credit usage row (Overpayment Credit or Service Credit)
                is_credit_row = is_credit_usage or 'Overpayment Credit' in ptype or 'Service Credit' in ptype
                
                # Color entire row if it's credit usage
                if is_credit_row:
                    for col in range(1, 9):
                        ws.cell(row=row, column=col).fill = credit_usage_fill
                else:
                    # Apply status color to status cell for non-credit rows
                    status_cell = ws.cell(row=row, column=5)
                    if status in status_colors:
                        status_cell.fill = PatternFill(
                            start_color=status_colors[status],
                            end_color=status_colors[status],
                            fill_type="solid"
                        )
                        if status in ['Delinquent', 'Not Paid']:
                            status_cell.font = Font(bold=True, color="FFFFFF")
                    
                    # Apply month color to "For Month" cell
                    month_cell = ws.cell(row=row, column=4)
                    if pmonth in month_colors:
                        month_cell.fill = PatternFill(
                            start_color=month_colors[pmonth],
                            end_color=month_colors[pmonth],
                            fill_type="solid"
                        )
                    
                    # Apply color to Details cell based on content
                    details_cell = ws.cell(row=row, column=6)
                    if 'Overpaid' in details:
                        details_cell.fill = PatternFill(
                            start_color=details_colors['Overpaid'],
                            end_color=details_colors['Overpaid'],
                            fill_type="solid"
                        )
                    elif 'remaining' in details:
                        details_cell.fill = PatternFill(
                            start_color=details_colors['remaining'],
                            end_color=details_colors['remaining'],
                            fill_type="solid"
                        )
                    elif 'Fully paid' in details:
                        details_cell.fill = PatternFill(
                            start_color=details_colors['Fully paid'],
                            end_color=details_colors['Fully paid'],
                            fill_type="solid"
                        )

                # Apply borders
                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = border
                    ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)

                row += 1
            row += 1

            # ===== MONTHLY BALANCE SUMMARY SECTION =====
            add_section_header("MONTHLY BALANCE SUMMARY (Filtered)")

            # Add column headers
            headers = ['Month', 'Status', 'Rent Due', 'Total Paid', 'Balance']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            row += 1

            # Add summary rows
            summaries = self.get_filtered_monthly_summaries(tenant)
            for month_key, summary in sorted(summaries.items()):
                ws.cell(row=row, column=1).value = month_key
                
                status = summary.get('status', '')
                ws.cell(row=row, column=2).value = status
                
                rent_due = float(summary.get('rent_due', 0.0))
                ws.cell(row=row, column=3).value = rent_due
                ws.cell(row=row, column=3).number_format = '$#,##0.00'
                
                total_paid = float(summary.get('total_paid', 0.0))
                ws.cell(row=row, column=4).value = total_paid
                ws.cell(row=row, column=4).number_format = '$#,##0.00'
                
                balance = float(summary.get('balance', 0.0))
                ws.cell(row=row, column=5).value = balance
                ws.cell(row=row, column=5).number_format = '$#,##0.00'

                # Apply status color and balance coloring
                status_cell = ws.cell(row=row, column=2)
                if status in status_colors:
                    status_cell.fill = PatternFill(
                        start_color=status_colors[status],
                        end_color=status_colors[status],
                        fill_type="solid"
                    )
                    if status in ['Delinquent', 'Not Paid']:
                        status_cell.font = Font(bold=True, color="FFFFFF")

                # Color balance column
                balance_cell = ws.cell(row=row, column=5)
                
                # Set balance based on status
                if status == "Not Due":
                    # Display balance as $0 for "Not Due" entries since they aren't delinquent yet
                    balance_cell.value = 0.00
                    balance_cell.font = Font(bold=True, color="000000")
                else:
                    balance_cell.value = balance
                
                # Apply color based on status
                if status == "Due Soon":
                    # Orange for "Due Soon" - upcoming deadline but not delinquent
                    balance_cell.fill = PatternFill(start_color=status_colors.get(status, "FFD580"), end_color=status_colors.get(status, "FFD580"), fill_type="solid")
                elif status == "Not Due":
                    # Light green for "Not Due" - no delinquency
                    balance_cell.fill = PatternFill(start_color=status_colors.get(status, "E2EFDA"), end_color=status_colors.get(status, "E2EFDA"), fill_type="solid")
                elif balance > 0:
                    # Red for delinquent/unpaid amounts
                    balance_cell.fill = PatternFill(start_color=status_colors.get(status, "FFC7CE"), end_color=status_colors.get(status, "FFC7CE"), fill_type="solid")
                elif balance < 0:
                    balance_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                # Apply borders
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = border
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                row += 1

            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 25
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 25

            # ===== CONFIGURE PRINT SETTINGS FOR MAIN SHEET =====
            # Set explicit print area to ensure all 8 columns are included (for LibreOffice Calc compatibility)
            ws.print_area = f'A1:H{row-1}'
            
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0  # 0 = allow multiple pages vertically
            ws.print_options.horizontalCentered = True
            ws.sheet_properties.pageSetUpPr.fitToPage = True

            # ===== CREATE YEAR-SPECIFIC SHEETS =====
            # Get all unique years from payment history
            from datetime import date as dt_date
            all_years = set()
            
            # Always include current year
            all_years.add(dt_date.today().year)
            
            # Collect years from payment_month (what payment is for)
            for payment in getattr(tenant, 'payment_history', []) or []:
                pmonth = payment.get('payment_month', '')
                if pmonth and '-' in pmonth:
                    year = int(pmonth.split('-')[0])
                    all_years.add(year)
                
                # Also collect years from date_received (when payment was received)
                date_received = payment.get('date', '')
                if date_received and '-' in date_received:
                    year = int(date_received.split('-')[0])
                    all_years.add(year)
            
            # Create a sheet for each year
            for year in sorted(all_years, reverse=True):
                self._create_year_sheet(wb, tenant, year, header_fill, header_font, section_fill, 
                                       section_font, border, credit_usage_fill, status_colors, 
                                       month_colors, details_colors, sort_by)

            wb.save(path)
            QMessageBox.information(self, 'Export Success', f'Exported to:\n{path}')

        except Exception as e:
            QMessageBox.critical(self, 'Export Error', f'Failed to export: {e}')

    def _create_year_sheet(self, wb, tenant, year, header_fill, header_font, section_fill, 
                          section_font, border, credit_usage_fill, status_colors, 
                          month_colors, details_colors, sort_by):
        """Create a worksheet for a specific year with filtered data"""
        from datetime import datetime, date
        
        # Determine if this is the current year
        current_year = date.today().year
        is_current_year = (year == current_year)
        
        # Create new sheet
        ws = wb.create_sheet(title=str(year))
        row = 1
        
        # Add report generation date
        ws.merge_cells(f'A{row}:H{row}')
        date_cell = ws[f'A{row}']
        filter_note = "Payment Month" if is_current_year else "Date Received"
        date_cell.value = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Year: {year} (Filtered by {filter_note})"
        date_cell.font = Font(bold=True, size=11)
        date_cell.alignment = Alignment(horizontal='left')
        row += 1
        
        # Add disclaimer
        disclaimer_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws.merge_cells(f'A{row}:H{row}')
        disclaimer_cell = ws[f'A{row}']
        if is_current_year:
            disclaimer_cell.value = f"ℹ️ NOTE: Current year ({year}) shows payments FOR this year. Rows highlighted in light yellow are credit usage and not included in payment totals."
        else:
            disclaimer_cell.value = f"ℹ️ NOTE: Previous year ({year}) shows payments RECEIVED in {year} (for tax purposes). Rows highlighted in light yellow are credit usage and not included in payment totals."
        disclaimer_cell.fill = disclaimer_fill
        disclaimer_cell.font = Font(italic=True, size=9)
        disclaimer_cell.alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[row].height = 40
        row += 2
        
        # Helper function to add section header
        def add_section_header(title):
            nonlocal row
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws[f'A{row}']
            cell.value = title
            cell.fill = section_fill
            cell.font = section_font
            row += 1
        
        # Helper function to add row with formatting
        def add_row(label, value):
            nonlocal row
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # ===== TENANT INFORMATION =====
        add_section_header(f"TENANT INFORMATION - {year}")
        add_row("Name", tenant.name)
        add_row("Tenant ID", tenant.tenant_id)
        add_row("Rental Period", self.format_rental_period(tenant.rental_period))
        add_row("Base Rent Amount", tenant.rent_amount)
        ws[f'B{row-1}'].number_format = '$#,##0.00'
        row += 1
        
        # ===== PAYMENT HISTORY FOR THIS YEAR =====
        add_section_header(f"PAYMENT HISTORY - {year} (Sorted by {('Date Received' if sort_by == 'date_received' else 'Payment Month')})")
        
        # Filter payments for this year
        year_payments = []
        for payment in getattr(tenant, 'payment_history', []) or []:
            if is_current_year:
                # Current year: filter by payment_month (what the payment is FOR)
                pmonth = payment.get('payment_month', '')
                if pmonth and pmonth.startswith(str(year)):
                    year_payments.append(payment)
            else:
                # Previous years: filter by date_received (for tax purposes)
                date_received = payment.get('date', '')
                if date_received and date_received.startswith(str(year)):
                    year_payments.append(payment)
        
        # Sort payments
        if sort_by == 'date_received':
            year_payments = sorted(
                year_payments,
                key=lambda p: datetime.strptime(p.get('date', '1970-01-01'), '%Y-%m-%d')
            )
        else:
            year_payments = sorted(
                year_payments,
                key=lambda p: (p.get('payment_month', '9999-99'), datetime.strptime(p.get('date', '1970-01-01'), '%Y-%m-%d'))
            )
        
        # Add column headers
        headers = ['Date Received', 'Amount', 'Type', 'For Month', 'Status', 'Month Status', 'Overpayment Created', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1
        
        # Add payment rows
        for payment in year_payments:
            date_received = payment.get('date', '')
            amount = float(payment.get('amount', 0.0))
            ptype = payment.get('type', '')
            pmonth = payment.get('payment_month', '')
            status = self.get_human_readable_status(tenant, payment)
            details = self.get_payment_details(tenant, payment)
            notes = payment.get('notes', '')
            is_credit_usage = payment.get('is_credit_usage', False)
            overpayment_created = self._calculate_overpayment_created(tenant, payment)
            
            # Add values
            ws.cell(row=row, column=1).value = datetime.strptime(date_received, '%Y-%m-%d') if date_received else None
            ws.cell(row=row, column=1).number_format = 'MM/DD/YYYY'
            
            ws.cell(row=row, column=2).value = amount
            ws.cell(row=row, column=2).number_format = '$#,##0.00'
            
            ws.cell(row=row, column=3).value = ptype
            ws.cell(row=row, column=4).value = pmonth
            ws.cell(row=row, column=5).value = status
            ws.cell(row=row, column=6).value = details
            ws.cell(row=row, column=7).value = overpayment_created if overpayment_created > 0 else ''
            ws.cell(row=row, column=7).number_format = '$#,##0.00'
            ws.cell(row=row, column=8).value = notes
            
            # Check if this is a credit usage row
            is_credit_row = is_credit_usage or 'Overpayment Credit' in ptype or 'Service Credit' in ptype
            
            # Color entire row if credit usage
            if is_credit_row:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = credit_usage_fill
            else:
                # Apply status color
                status_cell = ws.cell(row=row, column=5)
                if status in status_colors:
                    status_cell.fill = PatternFill(
                        start_color=status_colors[status],
                        end_color=status_colors[status],
                        fill_type="solid"
                    )
                    if status in ['Delinquent', 'Not Paid']:
                        status_cell.font = Font(bold=True, color="FFFFFF")
                
                # Apply month color
                month_cell = ws.cell(row=row, column=4)
                if pmonth in month_colors:
                    month_cell.fill = PatternFill(
                        start_color=month_colors[pmonth],
                        end_color=month_colors[pmonth],
                        fill_type="solid"
                    )
                
                # Apply details color
                details_cell = ws.cell(row=row, column=6)
                if 'Overpaid' in details:
                    details_cell.fill = PatternFill(
                        start_color=details_colors['Overpaid'],
                        end_color=details_colors['Overpaid'],
                        fill_type="solid"
                    )
                elif 'remaining' in details:
                    details_cell.fill = PatternFill(
                        start_color=details_colors['remaining'],
                        end_color=details_colors['remaining'],
                        fill_type="solid"
                    )
                elif 'Fully paid' in details:
                    details_cell.fill = PatternFill(
                        start_color=details_colors['Fully paid'],
                        end_color=details_colors['Fully paid'],
                        fill_type="solid"
                    )
            
            # Apply borders
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
            
            row += 1
        row += 1
        
        # ===== YEAR SUMMARY TOTALS (FOR TAX PURPOSES) =====
        add_section_header(f"YEAR SUMMARY - {year} (Tax Purposes)")
        
        # Calculate totals - excluding credit usage rows
        total_payment_amount = 0.0
        total_rent_due = 0.0
        
        for payment in year_payments:
            is_credit_usage = payment.get('is_credit_usage', False)
            ptype = payment.get('type', '')
            is_credit_row = is_credit_usage or 'Overpayment Credit' in ptype or 'Service Credit' in ptype
            
            # Only count actual payments (not credit usage) for tax purposes
            if not is_credit_row:
                total_payment_amount += float(payment.get('amount', 0.0))
        
        # Calculate total rent due for the year
        # Use get_monthly_summaries_all() to get unfiltered summaries for accurate calculation
        all_summaries = self.get_monthly_summaries_all(tenant)
        year_summaries = {k: v for k, v in all_summaries.items() if k.startswith(str(year))}
        
        for summary in year_summaries.values():
            rent_due_value = float(summary.get('rent_due', 0.0))
            total_rent_due += rent_due_value
        
        # Display summary totals
        add_row("Total Actual Payments", total_payment_amount)
        ws[f'B{row-1}'].number_format = '$#,##0.00'
        ws[f'B{row-1}'].font = Font(bold=True, size=12)
        ws[f'B{row-1}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        add_row("Total Rent Due", total_rent_due)
        ws[f'B{row-1}'].number_format = '$#,##0.00'
        ws[f'B{row-1}'].font = Font(bold=True, size=12)
        ws[f'B{row-1}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        
        # Calculate and display net difference
        net_difference = total_payment_amount - total_rent_due
        add_row("Net Difference", net_difference)
        ws[f'B{row-1}'].number_format = '$#,##0.00'
        ws[f'B{row-1}'].font = Font(bold=True, size=12)
        if net_difference >= 0:
            ws[f'B{row-1}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            ws[f'B{row-1}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        row += 1
        
        # ===== MONTHLY BALANCE SUMMARY FOR THIS YEAR =====
        add_section_header(f"MONTHLY BALANCE SUMMARY - {year}")
        
        # Add column headers
        headers = ['Month', 'Status', 'Rent Due', 'Total Paid', 'Balance']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1
        
        # Filter summaries for this year using all summaries (not filtered by UI)
        all_summaries = self.get_monthly_summaries_all(tenant)
        year_summaries = {k: v for k, v in all_summaries.items() if k.startswith(str(year))}
        
        for month_key, summary in sorted(year_summaries.items()):
            ws.cell(row=row, column=1).value = month_key
            
            status = summary.get('status', '')
            ws.cell(row=row, column=2).value = status
            
            rent_due = float(summary.get('rent_due', 0.0))
            ws.cell(row=row, column=3).value = rent_due
            ws.cell(row=row, column=3).number_format = '$#,##0.00'
            
            total_paid = float(summary.get('total_paid', 0.0))
            ws.cell(row=row, column=4).value = total_paid
            ws.cell(row=row, column=4).number_format = '$#,##0.00'
            
            balance = float(summary.get('balance', 0.0))
            ws.cell(row=row, column=5).value = balance
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            
            # Apply status color
            status_cell = ws.cell(row=row, column=2)
            if status in status_colors:
                status_cell.fill = PatternFill(
                    start_color=status_colors[status],
                    end_color=status_colors[status],
                    fill_type="solid"
                )
                if status in ['Delinquent', 'Not Paid']:
                    status_cell.font = Font(bold=True, color="FFFFFF")
            
            # Color balance column
            balance_cell = ws.cell(row=row, column=5)
            
            if status == "Not Due":
                balance_cell.value = 0.00
                balance_cell.font = Font(bold=True, color="000000")
            else:
                balance_cell.value = balance
            
            # Apply color based on status
            if status == "Due Soon":
                balance_cell.fill = PatternFill(start_color=status_colors.get(status, "FFD580"), end_color=status_colors.get(status, "FFD580"), fill_type="solid")
            elif status == "Not Due":
                balance_cell.fill = PatternFill(start_color=status_colors.get(status, "E2EFDA"), end_color=status_colors.get(status, "E2EFDA"), fill_type="solid")
            elif balance > 0:
                balance_cell.fill = PatternFill(start_color=status_colors.get(status, "FFC7CE"), end_color=status_colors.get(status, "FFC7CE"), fill_type="solid")
            elif balance < 0:
                balance_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 25
        
        # ===== CONFIGURE PRINT SETTINGS FOR YEAR SHEET =====
        # Set explicit print area to ensure all 8 columns are included (for LibreOffice Calc compatibility)
        ws.print_area = f'A1:H{row-1}'
        
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 0 = allow multiple pages vertically
        ws.print_options.horizontalCentered = True
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    def export_to_csv_plain(self, tenant, path, sort_by='date_received'):
        """Export rent management data to plain CSV (fallback if openpyxl not available)"""
        try:
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Add report generation date at the very top
                from datetime import datetime
                writer.writerow(['Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                writer.writerow([''])
                
                # Add disclaimer at the top
                writer.writerow([''])
                writer.writerow(['NOTE: Rows marked with "Yes" in the "Is Credit Usage" column are Overpayment Credit or Service Credit usage.'])
                writer.writerow(['These do NOT represent new rent payments and are not included in total payment calculations,'])
                writer.writerow(['as they were already counted when the original overpayment occurred.'])
                writer.writerow([''])

                # Section: Tenant Information
                writer.writerow(['Section', 'Tenant Information'])
                writer.writerow(['Name', tenant.name])
                writer.writerow(['Tenant ID', tenant.tenant_id])
                writer.writerow(['Rental Period', self.format_rental_period(tenant.rental_period)])
                writer.writerow(['Rent Amount', f"{tenant.rent_amount:.2f}"])
                writer.writerow(['Deposit', f"{tenant.deposit_amount:.2f}"])
                writer.writerow(['Due Day', getattr(tenant, 'rent_due_date', '') or ''])
                contact = self.format_contact_info(tenant.contact_info)
                writer.writerow(['Contact', contact])
                # Notes may be a list; join with semicolons for CSV row
                if isinstance(getattr(tenant, '_notes_list', None), list):
                    notes_csv = '; '.join(tenant._notes_list)
                else:
                    notes_csv = tenant.notes or ''
                writer.writerow(['Notes', notes_csv])
                writer.writerow([])

                # Section: Current Status & Calculations
                from datetime import date
                today = date.today()
                current_month_key = f"{today.year}-{today.month:02d}"
                monthly_override = None
                yearly_override = None
                effective_rent = tenant.rent_amount
                if hasattr(tenant, 'monthly_exceptions') and tenant.monthly_exceptions:
                    monthly_override = tenant.monthly_exceptions.get(current_month_key)
                    if monthly_override is None:
                        yearly_override = tenant.monthly_exceptions.get(str(today.year))
                    if monthly_override is not None:
                        effective_rent = monthly_override
                    elif yearly_override is not None:
                        effective_rent = yearly_override

                delinquency_balance = float(getattr(tenant, 'delinquency_balance', 0.0) or 0.0)
                overpayment_credit = float(getattr(tenant, 'overpayment_credit', 0.0) or 0.0)
                service_credit = float(getattr(tenant, 'service_credit', 0.0) or 0.0)
                net_due = delinquency_balance - overpayment_credit

                writer.writerow(['Section', 'Current Status'])
                writer.writerow(['Current Month', today.strftime('%B %Y')])
                writer.writerow(['Base Rent', f"{tenant.rent_amount:.2f}"])
                writer.writerow(['Active Override', f"{monthly_override if monthly_override is not None else yearly_override if yearly_override is not None else 'None'}"])
                writer.writerow(['Effective Rent', f"{effective_rent:.2f}"])
                writer.writerow(['Delinquency', f"{delinquency_balance:.2f}"])
                writer.writerow(['Overpayment Credit', f"{overpayment_credit:.2f}"])
                writer.writerow(['Service Credit', f"{service_credit:.2f}"])
                writer.writerow(['Net Amount Due', f"{abs(net_due):.2f}" if net_due != 0 else '0.00'])
                writer.writerow([])

                # Section: Payment History (sorted based on user selection)
                writer.writerow(['Section', f'Payment History (Sorted by {("Date Received" if sort_by == "date_received" else "Payment Month")})'])
                writer.writerow(['Date Received', 'Amount', 'Type', 'For Month', 'Status', 'Month Status', 'Is Credit Usage', 'Overpayment Created', 'Notes'])
                
                # Sort payments based on user selection
                payments = getattr(tenant, 'payment_history', []) or []
                if sort_by == 'date_received':
                    payments = sorted(
                        payments,
                        key=lambda p: datetime.strptime(p.get('date', '1970-01-01'), '%Y-%m-%d')
                    )
                else:  # sort_by == 'payment_month'
                    payments = sorted(
                        payments,
                        key=lambda p: (p.get('payment_month', '9999-99'), datetime.strptime(p.get('date', '1970-01-01'), '%Y-%m-%d'))
                    )
                
                for payment in payments:
                    date_received = payment.get('date', '')
                    amount = f"{payment.get('amount', 0.0):.2f}"
                    ptype = payment.get('type', '')
                    pmonth = payment.get('payment_month', '')
                    status = self.get_human_readable_status(tenant, payment)
                    details = self.get_payment_details(tenant, payment)
                    is_credit_usage = payment.get('is_credit_usage', False)
                    is_credit_row = is_credit_usage or 'Overpayment Credit' in ptype or 'Service Credit' in ptype
                    overpayment_created = self._calculate_overpayment_created(tenant, payment)
                    notes = payment.get('notes', '')
                    writer.writerow([
                        date_received, 
                        amount, 
                        ptype, 
                        pmonth, 
                        status, 
                        details, 
                        'Yes' if is_credit_row else 'No',
                        f"{overpayment_created:.2f}" if overpayment_created > 0 else '',
                        notes
                    ])
                writer.writerow([])

                # Section: Monthly Balance Summary (respect current filter)
                writer.writerow(['Section', 'Monthly Balance Summary (Filtered)'])
                writer.writerow(['Month', 'Status', 'Rent Due', 'Total Paid', 'Balance'])
                summaries = self.get_filtered_monthly_summaries(tenant)
                for month_key, summary in sorted(summaries.items()):
                    writer.writerow([
                        month_key,
                        summary.get('status', ''),
                        f"{summary.get('rent_due', 0.0):.2f}",
                        f"{summary.get('total_paid', 0.0):.2f}",
                        f"{summary.get('balance', 0.0):.2f}"
                    ])

            QMessageBox.information(self, 'Export CSV', f'Exported to:\n{path}')
        except Exception as e:
            QMessageBox.critical(self, 'Export CSV', f'Failed to export: {e}')

    def on_payment_selection_changed(self):
        """Handle payment history row selection changes"""
        print(f"[DEBUG] on_payment_selection_changed called")
        selected_items = self.payment_table.selectedItems()
        print(f"[DEBUG] selected_items count: {len(selected_items)}")
        
        account_status = getattr(self.selected_tenant, 'account_status', 'Active') if self.selected_tenant else 'Active'
        can_modify_payments = account_status.lower() == 'active'
        print(f"[DEBUG] account_status: {account_status}, can_modify_payments: {can_modify_payments}")
        
        if selected_items:
            # A row is selected - enable modify button only if tenant is Active
            row = selected_items[0].row()
            self.selected_payment_row = row
            
            # Get payment data from the selected row based on the new table structure
            date_received = self.payment_table.item(row, 0).text()
            amount_text = self.payment_table.item(row, 1).text()
            payment_type = self.payment_table.item(row, 2).text()
            payment_month = self.payment_table.item(row, 3).text()
            status = self.payment_table.item(row, 4).text()
            details = self.payment_table.item(row, 5).text()
            notes = self.payment_table.item(row, 6).text() if self.payment_table.item(row, 6) else ''
            
            # Parse amount (remove $ and convert to float)
            try:
                amount = float(amount_text.replace('$', ''))
            except (ValueError, AttributeError):
                amount = 0.0
            
            # Get the actual payment record from payment_data
            # The payment_data is already re-ordered by the sorted table, so row index is correct
            actual_payment = {}
            if hasattr(self, 'payment_data') and 0 <= row < len(self.payment_data):
                actual_payment = self.payment_data[row].get('payment', {})
                print(f"[DEBUG] Selected payment from payment_data[{row}]: {actual_payment}")
            elif hasattr(self.selected_tenant, 'payment_history'):
                # Fallback to direct access if payment_data not available
                payment_history = self.selected_tenant.payment_history
                if 0 <= row < len(payment_history):
                    actual_payment = payment_history[row]
                    print(f"[DEBUG] Selected payment from fallback payment_history[{row}]: {actual_payment}")
            
            self.selected_payment_data = {
                'row_index': row,
                'date': date_received,
                'amount': amount,
                'type': payment_type,
                'payment_month': payment_month,
                'status': status,
                'details': details,
                'payment_record': actual_payment
            }
            print(f"[DEBUG] selected_payment_data set: {self.selected_payment_data}")
            
            # Enable modify button only if tenant is Active
            print(f"[DEBUG] Setting modify button enabled to: {can_modify_payments}")
            self.modify_payment_btn.setEnabled(can_modify_payments)
        else:
            # No row selected - disable modify button
            print(f"[DEBUG] No payment selected, disabling modify button")
            self.selected_payment_row = None
            self.selected_payment_data = None
            self.modify_payment_btn.setEnabled(False)
    
    def show_add_payment_dialog(self):
        """Show dialog for adding a new payment"""
        if not self.selected_tenant:
            return
            
        dialog = PaymentDialog(self, mode='add', tenant=self.selected_tenant)
        if dialog.exec() == dialog.DialogCode.Accepted:
            payment_data = dialog.get_payment_data()
            self.process_payment_addition(payment_data)
    
    def show_add_refund_dialog(self):
        """Show dialog for adding a refund/reverse payment"""
        if not self.selected_tenant:
            return
            
        dialog = PaymentDialog(self, mode='refund', tenant=self.selected_tenant)
        if dialog.exec() == dialog.DialogCode.Accepted:
            payment_data = dialog.get_payment_data()
            self.process_refund_addition(payment_data)
    
    def on_modify_payment_button_clicked(self):
        """Wrapper for modify payment button click - with logging"""
        print(f"[DEBUG] on_modify_payment_button_clicked - button clicked!")
        self.show_modify_payment_dialog()
    
    def show_modify_payment_dialog(self):
        """Show dialog for modifying an existing payment"""
        print(f"[DEBUG] show_modify_payment_dialog called")
        print(f"[DEBUG] selected_tenant: {self.selected_tenant}")
        print(f"[DEBUG] selected_payment_data: {self.selected_payment_data}")
        
        if not self.selected_tenant:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, 'No Tenant Selected', 'Please select a tenant first.')
            return
        
        if not self.selected_payment_data:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, 'No Payment Selected', 'Please select a payment from the history to modify.')
            return
        
        # Pass the actual payment record to the dialog
        payment_record = self.selected_payment_data.get('payment_record', {})
        print(f"[DEBUG] payment_record: {payment_record}")
        
        dialog = PaymentDialog(self, mode='modify', payment_data=payment_record, tenant=self.selected_tenant)
        result = dialog.exec()
        print(f"[DEBUG] dialog result: {result}")
        
        if result == dialog.DialogCode.Accepted:
            if hasattr(dialog, 'delete_requested') and dialog.delete_requested:
                self.process_payment_deletion()
            else:
                payment_data = dialog.get_payment_data()
                payment_data['row_index'] = self.selected_payment_data['row_index']
                self.process_payment_modification(payment_data)
    
    def process_refund_addition(self, payment_data):
        """Process adding a refund/reverse payment (negative payment)"""
        try:
            # Extract payment information - amount will be negative
            amount = -abs(payment_data['amount'])  # Ensure negative
            payment_type = payment_data['payment_type']
            payment_month = payment_data['payment_month']
            date_received = payment_data['date_received']
            notes = payment_data.get('notes', '')
            
            # Add refund note to payment type for clarity
            payment_type_display = f"{payment_type} (Refund)"
            
            # Process as regular payment but with negative amount
            success = self.rent_tracker.add_payment(
                tenant_name=self.selected_tenant.name, 
                amount=amount, 
                payment_type=payment_type_display,
                payment_date=date_received,
                payment_month=payment_month,
                notes=notes
            )
            
            if success:
                from PyQt6.QtWidgets import QMessageBox
                QMessageBox.information(self, "Refund Recorded", 
                    f"Refund of ${abs(amount):.2f} via {payment_type} for {payment_month} has been recorded successfully.\n"
                    f"Date: {date_received}\n\n"
                    f"This negative payment has been applied to the tenant's account.")
            else:
                from PyQt6.QtWidgets import QMessageBox
                QMessageBox.critical(self, "Refund Error", "Failed to record refund.")
                return
            
            # Refresh the tenant data
            updated_tenant = self.rent_tracker.get_tenant_by_name(self.selected_tenant.name)
            if updated_tenant:
                print(f"[DEBUG] Refreshing tenant after refund - Delinquency: {updated_tenant.delinquency_balance}, Overpayment: {updated_tenant.overpayment_credit}")
                self.load_tenant(updated_tenant)
            else:
                print("[DEBUG] Could not get updated tenant after refund")
            
            # Refresh dashboard charts to reflect refund
            self.refresh_dashboard_charts()
            
            # Refresh the Tenant Management display
            self.refresh_balance_summary()
                
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Refund Error", f"Failed to record refund: {str(e)}")
    
    def process_payment_addition(self, payment_data):
        """Process adding a new payment"""
        try:
            # Extract payment information
            amount = payment_data['amount']
            payment_type = payment_data['payment_type']
            payment_month = payment_data['payment_month']
            date_received = payment_data['date_received']
            notes = payment_data.get('notes', '')
            
            print(f"[DEBUG] process_payment_addition - amount from dialog: {amount}, type: {payment_type}")
            
            # Handle overpayment credit usage
            if payment_type == 'Overpayment Credit':
                # Validate available credit
                available_credit = getattr(self.selected_tenant, 'overpayment_credit', 0.0)
                print(f"[DEBUG] Overpayment Credit - requested: ${amount:.2f}, available: ${available_credit:.2f}")
                if amount > available_credit:
                    from PyQt6.QtWidgets import QMessageBox
                    QMessageBox.warning(self, "Insufficient Credit", 
                        f"Cannot apply ${amount:.2f} when only ${available_credit:.2f} credit is available.")
                    return
                
                # Reduce overpayment credit
                self.selected_tenant.overpayment_credit -= amount
                print(f"[DEBUG] Reduced overpayment credit by ${amount:.2f}, new balance: ${self.selected_tenant.overpayment_credit:.2f}")
                
                # IMPORTANT: Save the credit reduction BEFORE calling add_payment
                # This ensures add_payment will load the updated tenant with the reduced credit
                # Use a callback to ensure save completes before adding payment
                def after_credit_saved(result):
                    # Add modifier to track this as credit usage
                    payment_type_with_modifier = f"Overpayment Credit (Applied: ${amount:.2f})"
                    
                    # Add payment with special tracking
                    success = self.rent_tracker.add_payment(
                        tenant_name=self.selected_tenant.name, 
                        amount=amount, 
                        payment_type=payment_type_with_modifier,
                        payment_date=date_received,
                        payment_month=payment_month,
                        is_credit_usage=True,  # Special flag for reporting
                        notes=notes
                    )
                    
                    if success:
                        # Save again to ensure all changes are persisted (though add_payment also saves)
                        def after_payment_saved(result):
                            from PyQt6.QtWidgets import QMessageBox
                            QMessageBox.information(self, "Overpayment Credit Applied", 
                                f"${amount:.2f} overpayment credit applied to {payment_month}.\n"
                                f"Remaining credit: ${self.selected_tenant.overpayment_credit:.2f}\n"
                                f"Date applied: {date_received}")
                            # Refresh all UI elements after payment addition
                            self.refresh_ui_after_payment_change()
                        
                        self.save_tenants_async(on_success=after_payment_saved)
                    else:
                        # Restore credit if payment failed
                        self.selected_tenant.overpayment_credit += amount
                        from PyQt6.QtWidgets import QMessageBox
                        QMessageBox.critical(self, "Payment Error", "Failed to apply overpayment credit.")
                
                self.save_tenants_async(on_success=after_credit_saved)
            
            # Handle service credit usage
            elif payment_type == 'Service Credit':
                # Validate available service credit
                available_credit = getattr(self.selected_tenant, 'service_credit', 0.0)
                if amount > available_credit:
                    from PyQt6.QtWidgets import QMessageBox
                    QMessageBox.warning(self, "Insufficient Service Credit", 
                        f"Cannot apply ${amount:.2f} when only ${available_credit:.2f} service credit is available.")
                    return
                
                # Reduce service credit
                self.selected_tenant.service_credit -= amount
                
                # Add to service credit history
                if not hasattr(self.selected_tenant, 'service_credit_history'):
                    self.selected_tenant.service_credit_history = []
                
                credit_record = {
                    'amount': -amount,  # Negative because it's being used
                    'service_type': 'Applied to Rent',
                    'date': date_received.strftime('%Y-%m-%d') if hasattr(date_received, 'strftime') else str(date_received),
                    'notes': f'Applied service credit to {payment_month} rent',
                    'transaction_type': 'applied',
                    'timestamp': datetime.now().isoformat()
                }
                self.selected_tenant.service_credit_history.append(credit_record)
                
                # IMPORTANT: Save the credit reduction BEFORE calling add_payment
                # This ensures add_payment will load the updated tenant with the reduced credit
                # Use a callback to ensure save completes before adding payment
                def after_service_credit_saved(result):
                    # Add modifier to track this as service credit usage
                    payment_type_with_modifier = f"Service Credit (Applied: ${amount:.2f})"
                    
                    # Add payment as regular payment (not credit_usage)
                    # Service credits should act like cash/zelle - they reduce delinquency
                    success = self.rent_tracker.add_payment(
                        tenant_name=self.selected_tenant.name, 
                        amount=amount, 
                        payment_type=payment_type_with_modifier,
                        payment_date=date_received,
                        payment_month=payment_month,
                        is_credit_usage=False,  # Service credits are like regular payments
                        notes=notes
                    )
                    
                    if success:
                        # Save again to ensure all changes are persisted
                        def after_service_payment_saved(result):
                            try:
                                if hasattr(self.selected_tenant, 'add_note'):
                                    self.selected_tenant.add_note(f"Service Credit Applied: ${amount:.2f} to {payment_month} on {date_received}.")
                                else:
                                    existing = getattr(self.selected_tenant, 'notes', '') or ''
                                    line = f"Service Credit Applied: ${amount:.2f} to {payment_month} on {date_received}."
                                    self.selected_tenant.notes = (existing + "\n" + line).strip() if existing else line
                            except Exception:
                                pass
                            
                            from PyQt6.QtWidgets import QMessageBox
                            QMessageBox.information(self, "Service Credit Applied", 
                                f"${amount:.2f} service credit applied to {payment_month}.\n"
                                f"Remaining service credit: ${self.selected_tenant.service_credit:.2f}\n"
                                f"Date applied: {date_received}")
                            # Refresh all UI elements after payment addition
                            self.refresh_ui_after_payment_change()
                        
                        self.save_tenants_async(on_success=after_service_payment_saved)
                    else:
                        # Restore credit if payment failed
                        self.selected_tenant.service_credit += amount
                        # Remove the credit record we just added
                        if self.selected_tenant.service_credit_history:
                            self.selected_tenant.service_credit_history.pop()
                        from PyQt6.QtWidgets import QMessageBox
                        QMessageBox.critical(self, "Payment Error", "Failed to apply service credit.")
                
                self.save_tenants_async(on_success=after_service_credit_saved)
            
            else:
                # Regular payment processing
                success = self.rent_tracker.add_payment(
                    tenant_name=self.selected_tenant.name, 
                    amount=amount, 
                    payment_type=payment_type,
                    payment_date=date_received,
                    payment_month=payment_month,
                    notes=notes
                )
                
                if success:
                    def on_regular_payment_saved(result):
                        from PyQt6.QtWidgets import QMessageBox
                        QMessageBox.information(self, "Payment Added", 
                            f"Payment of ${amount:.2f} via {payment_type} for {payment_month} has been recorded successfully.\n"
                            f"Date received: {date_received}")
                        # Refresh all UI elements after payment addition
                        self.refresh_ui_after_payment_change()
                    
                    self.save_tenants_async(on_success=on_regular_payment_saved)
                else:
                    from PyQt6.QtWidgets import QMessageBox
                    QMessageBox.critical(self, "Payment Error", "Failed to add payment.")
                    return
                
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Payment Error", f"Failed to add payment: {str(e)}")
    
    def process_payment_modification(self, payment_data):
        """Process modifying an existing payment"""
        try:
            # Extract payment information
            amount = payment_data['amount']
            payment_type = payment_data['payment_type']
            payment_month = payment_data['payment_month']
            date_received = payment_data['date_received']
            notes = payment_data.get('notes', '')
            row_index = payment_data.get('row_index', -1)
            
            # Update the payment record in the tenant's payment history
            if hasattr(self.selected_tenant, 'payment_history') and 0 <= row_index < len(self.selected_tenant.payment_history):
                old_payment = self.selected_tenant.payment_history[row_index].copy()
                
                # Parse payment month to get year and month
                try:
                    year, month = payment_month.split('-')
                    year, month = int(year), int(month)
                except (ValueError, AttributeError):
                    year, month = date.today().year, date.today().month
                
                # Update the payment record
                self.selected_tenant.payment_history[row_index].update({
                    'amount': amount,
                    'type': payment_type,
                    'date': date_received,
                    'payment_month': payment_month,
                    'year': year,
                    'month': month,
                    'notes': notes
                })
                
                # Recalculate monthly status for affected months
                self.recalculate_monthly_status(self.selected_tenant, old_payment, self.selected_tenant.payment_history[row_index])
                
                # Save changes with callback to refresh UI after save completes
                def on_modification_saved(result):
                    from PyQt6.QtWidgets import QMessageBox
                    QMessageBox.information(self, "Payment Modified", 
                        f"Payment for {payment_month} has been updated to ${amount:.2f} via {payment_type}.\n"
                        f"Date received: {date_received}")
                    
                    # Refresh all UI elements after payment modification completes
                    self.refresh_ui_after_payment_change()
                
                def on_modification_error(error):
                    from PyQt6.QtWidgets import QMessageBox
                    QMessageBox.critical(self, "Save Error", f"Failed to save modified payment: {str(error)}")
                
                self.save_tenants_async(on_success=on_modification_saved, on_error=on_modification_error)
            else:
                from PyQt6.QtWidgets import QMessageBox
                QMessageBox.warning(self, "Error", "Could not find the payment to modify.")
                
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Modify Error", f"Failed to modify payment: {str(e)}")
    
    def recalculate_monthly_status(self, tenant, old_payment, new_payment):
        """Recalculate monthly status after payment modification"""
        try:
            # Use the existing comprehensive delinquency check method to recalculate all balances
            # This will properly handle overpayments, delinquency, and monthly status for all months
            self.rent_tracker.check_and_update_delinquency()
            
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Recalculation Error", f"Failed to recalculate monthly status: {str(e)}")
    
    def refresh_ui_after_payment_change(self):
        """Refresh all UI elements after a payment has been added, modified, or deleted"""
        try:
            # Force a full delinquency recalculation to ensure all balances are correct
            self.rent_tracker.check_and_update_delinquency()
            
            # Get the updated tenant
            updated_tenant = self.rent_tracker.get_tenant_by_name(self.selected_tenant.name)
            if updated_tenant:
                print(f"[DEBUG] Refreshing UI after payment change - Delinquency: ${updated_tenant.delinquency_balance:.2f}, Overpayment: ${updated_tenant.overpayment_credit:.2f}")
                self.selected_tenant = updated_tenant
                
                # Reload the full tenant display (header info, status, etc.)
                self.load_tenant(updated_tenant)
                
                # Refresh payment table
                self.populate_payment_table(updated_tenant)
                
                # Reset button states
                self.modify_payment_btn.setEnabled(False)
                self.selected_payment_data = None
                self.selected_payment_row = None
                
                # Refresh dashboard charts
                self.refresh_dashboard_charts()
                
                # Refresh the balance summary
                self.refresh_balance_summary()
            else:
                print("[DEBUG] Could not get updated tenant after payment change")
                
        except Exception as e:
            print(f"[ERROR] Failed to refresh UI after payment change: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def process_payment_deletion(self):
        """Process deleting a payment"""
        print(f"[DEBUG] process_payment_deletion called")
        print(f"[DEBUG] selected_payment_data: {self.selected_payment_data}")
        try:
            if not self.selected_tenant or not self.selected_payment_data:
                from PyQt6.QtWidgets import QMessageBox
                print(f"[DEBUG] No tenant or payment data selected")
                QMessageBox.warning(self, "Error", "No payment selected for deletion.")
                return
            
            row_index = self.selected_payment_data.get('row_index', -1)
            print(f"[DEBUG] Row index: {row_index}")
            
            # Get the actual payment to delete
            # Use payment_data structure to handle table sorting correctly
            actual_payment_to_delete = self.selected_payment_data.get('payment_record', {})
            print(f"[DEBUG] Payment to delete from selected_payment_data['payment_record']: {actual_payment_to_delete}")
            print(f"[DEBUG] Payment to delete - date: {actual_payment_to_delete.get('date')}, amount: {actual_payment_to_delete.get('amount')}, month: {actual_payment_to_delete.get('payment_month')}")
            
            if not actual_payment_to_delete:
                from PyQt6.QtWidgets import QMessageBox
                print(f"[DEBUG] No payment record found")
                QMessageBox.warning(self, "Error", "Could not identify the payment to delete.")
                return
            
            # Remove the payment from the tenant's payment history by finding and removing it
            if hasattr(self.selected_tenant, 'payment_history'):
                payment_history = self.selected_tenant.payment_history
                print(f"[DEBUG] Payment history length: {len(payment_history)}")
                print(f"[DEBUG] Full payment_history: {payment_history}")
                # Find and remove the matching payment
                payment_removed = False
                for i, payment in enumerate(payment_history):
                    # Debug each payment
                    print(f"[DEBUG] Comparing payment {i}:")
                    print(f"[DEBUG]   from_history: {payment}")
                    print(f"[DEBUG]   to_delete: {actual_payment_to_delete}")
                    print(f"[DEBUG]   date_match: {payment.get('date')} == {actual_payment_to_delete.get('date')} -> {payment.get('date') == actual_payment_to_delete.get('date')}")
                    print(f"[DEBUG]   amount_match: {payment.get('amount')} == {actual_payment_to_delete.get('amount')} -> {payment.get('amount') == actual_payment_to_delete.get('amount')}")
                    print(f"[DEBUG]   month_match: {payment.get('payment_month')} == {actual_payment_to_delete.get('payment_month')} -> {payment.get('payment_month') == actual_payment_to_delete.get('payment_month')}")
                    # Match by comparing key fields
                    if (payment.get('date') == actual_payment_to_delete.get('date') and
                        payment.get('amount') == actual_payment_to_delete.get('amount') and
                        payment.get('payment_month') == actual_payment_to_delete.get('payment_month')):
                        print(f"[DEBUG] ✓ MATCH FOUND! Deleting payment at index {i}")
                        deleted_payment = payment_history[i].copy()
                        del payment_history[i]
                        payment_removed = True
                        print(f"[DEBUG] Payment found and removed at index {i}")
                        break

                
                if not payment_removed:
                    from PyQt6.QtWidgets import QMessageBox
                    print(f"[DEBUG] Payment not found in history by field matching, attempting fallback by row index {row_index}")
                    # Fallback: try to delete by row index if field matching fails
                    if 0 <= row_index < len(payment_history):
                        deleted_payment = payment_history[row_index].copy()
                        del payment_history[row_index]
                        payment_removed = True
                        print(f"[DEBUG] Payment deleted by fallback row index method")
                    else:
                        QMessageBox.warning(self, "Error", "Could not find the payment to delete.")
                        return
                
                # Recalculate tenant status after payment deletion
                self.recalculate_tenant_status_after_deletion(self.selected_tenant, deleted_payment)
                
                # Save changes with callback to refresh UI after save completes
                def on_deletion_saved(result):
                    from PyQt6.QtWidgets import QMessageBox
                    print(f"[DEBUG] Payment deleted and saved successfully")
                    
                    QMessageBox.information(self, "Payment Deleted", 
                        f"Payment of ${deleted_payment.get('amount', 0):.2f} for {deleted_payment.get('payment_month', 'unknown month')} has been deleted.")
                    
                    # Refresh all UI elements after payment deletion completes
                    self.refresh_ui_after_payment_change()
                
                def on_deletion_error(error):
                    from PyQt6.QtWidgets import QMessageBox
                    QMessageBox.critical(self, "Save Error", f"Failed to save after payment deletion: {str(error)}")
                
                self.save_tenants_async(on_success=on_deletion_saved, on_error=on_deletion_error)
            else:
                from PyQt6.QtWidgets import QMessageBox
                print(f"[DEBUG] Tenant has no payment_history attribute")
                QMessageBox.warning(self, "Error", "Could not access payment history.")
                
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            print(f"[DEBUG] Exception in process_payment_deletion: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Delete Error", f"Failed to delete payment: {str(e)}")

    def recalculate_tenant_status_after_deletion(self, tenant, deleted_payment):
        """Recalculate tenant status after payment deletion"""
        try:
            # Use the existing delinquency check method to recalculate all balances
            # This will properly recalculate delinquency_balance, overpayment_credit, and monthly status
            # Pass the tenant ID to recalculate only this tenant
            self.rent_tracker.check_and_update_delinquency(target_tenant_id=tenant.tenant_id)
            print(f"[DEBUG] Recalculated tenant {tenant.name} after payment deletion")
            print(f"[DEBUG] Updated overpayment_credit: {tenant.overpayment_credit}")
            print(f"[DEBUG] Updated delinquency_balance: {tenant.delinquency_balance}")
            
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            print(f"[DEBUG] Error during recalculation: {e}")
            QMessageBox.critical(self, "Recalculation Error", f"Failed to recalculate tenant status after deletion: {str(e)}")

    def on_payment_type_changed(self, text):
        """Show/hide the 'Other' payment type input field"""
        if text == 'Other':
            self.other_payment_input.show()
        else:
            self.other_payment_input.hide()

    def submit_payment(self):
        if not self.selected_tenant:
            return
            
        amount = self.pay_amount.value()
        if amount <= 0:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Invalid Amount", "Please enter a valid payment amount.")
            return
        
        # Get payment type
        payment_type = self.pay_type.currentText()
        if payment_type == 'Other':
            other_type = self.other_payment_input.text().strip()
            if not other_type:
                from PyQt6.QtWidgets import QMessageBox
                QMessageBox.warning(self, "Invalid Payment Type", "Please specify the payment type for 'Other'.")
                return
            payment_type = f"Other: {other_type}"
            
        try:
            # Add payment to rent tracker (you may need to modify rent_tracker to accept payment type)
            self.rent_tracker.add_payment(self.selected_tenant.name, amount)
            
            # Refresh the tenant data
            updated_tenant = self.rent_tracker.get_tenant_by_name(self.selected_tenant.name)
            if updated_tenant:
                self.load_tenant(updated_tenant)
                
            # Reset payment fields
            self.pay_amount.setValue(0.0)
            self.pay_type.setCurrentIndex(0)  # Reset to 'Cash'
            self.other_payment_input.clear()
            self.other_payment_input.hide()
            
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "Payment Submitted", 
                f"Payment of ${amount:.2f} via {payment_type} has been recorded successfully.")
            
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Payment Error", f"Failed to submit payment: {str(e)}")

    def save_defaults(self):
        due_day_text = self.default_due_day.text().strip()
        try:
            due_day = int(due_day_text)
        except ValueError:
            QMessageBox.warning(self, "Invalid Due Day", "Default due day must be a whole number between 1 and 31.")
            return

        if due_day < 1 or due_day > 31:
            QMessageBox.warning(self, "Invalid Due Day", "Default due day must be between 1 and 31.")
            return

        self.settings.set('default_rent_amount', self.default_rent.value())
        self.settings.set('default_deposit_amount', self.default_deposit.value())
        self.settings.set('default_due_day', str(due_day))
        self.default_due_day.setText(str(due_day))
        QMessageBox.information(self, "Defaults Saved", "Default rent settings were updated successfully.")

    def show_modify_rent_dialog(self):
        """Show dialog to modify base rent amount"""
        if not self.selected_tenant:
            return
        
        dialog = RentModificationDialog(self, self.selected_tenant.rent_amount, self.selected_tenant.name)
        if dialog.exec() == dialog.DialogCode.Accepted:
            rent_data = dialog.get_rent_data()
            self.process_rent_modification(rent_data)
    
    def show_monthly_override_dialog(self):
        """Show dialog for monthly rent override"""
        if not self.selected_tenant:
            return
        
        dialog = MonthlyOverrideDialog(self, self.selected_tenant.rent_amount, self.selected_tenant.name)
        if dialog.exec() == dialog.DialogCode.Accepted:
            override_data = dialog.get_override_data()
            self.process_monthly_override(override_data)
    
    def show_yearly_override_dialog(self):
        """Show dialog for yearly rent override"""
        if not self.selected_tenant:
            return
        
        dialog = YearlyOverrideDialog(self, self.selected_tenant.rent_amount, self.selected_tenant.name)
        if dialog.exec() == dialog.DialogCode.Accepted:
            override_data = dialog.get_override_data()
            self.process_yearly_override(override_data)
    
    def show_service_credit_dialog(self):
        """Show dialog for adding service credits"""
        if not self.selected_tenant:
            QMessageBox.warning(self, 'No Tenant Selected', 'Please select a tenant first.')
            return
        
        try:
            from ui.service_credit_dialog import ServiceCreditDialog
            dialog = ServiceCreditDialog(self, tenant=self.selected_tenant, rent_tracker=self.rent_tracker)
            dialog.credit_added.connect(self.refresh_tenant_display)
            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to open service credit dialog: {str(e)}')
    
    def show_credit_conversion_dialog(self):
        """Show dialog for converting between credit types"""
        if not self.selected_tenant:
            QMessageBox.warning(self, 'No Tenant Selected', 'Please select a tenant first.')
            return
        
        try:
            from ui.credit_conversion_dialog import CreditConversionDialog
            dialog = CreditConversionDialog(self, tenant=self.selected_tenant, rent_tracker=self.rent_tracker)
            dialog.credit_converted.connect(self.refresh_tenant_display)
            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to open credit conversion dialog: {str(e)}')
    
    def process_rent_modification(self, rent_data):
        """Process base rent modification"""
        try:
            old_rent = self.selected_tenant.rent_amount
            new_rent = rent_data['new_rent']
            effective_date = rent_data['effective_date']
            notes = rent_data['notes']
            
            # Update tenant rent amount
            self.selected_tenant.rent_amount = new_rent
            
            # Save changes
            self.save_tenants_async()
            
            # Refresh the display
            self.load_tenant(self.selected_tenant)
            
            # Show confirmation
            QMessageBox.information(self, "Rent Modified", 
                f"Base rent updated from ${old_rent:.2f} to ${new_rent:.2f}\n"
                f"Effective Date: {effective_date}\n"
                f"Notes: {notes if notes else 'None'}")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to modify rent: {str(e)}")
    
    def process_monthly_override(self, override_data):
        """Process monthly rent override"""
        try:
            print(f"DEBUG: override_data received: {override_data}")
            
            month_str = override_data.get('month')  # Format: "YYYY-MM"
            amount = override_data.get('override_amount')
            notes = override_data.get('notes', '')
            remove_requested = override_data.get('remove_requested', False)
            
            print(f"DEBUG: month_str={month_str}, amount={amount}, notes={notes}, remove_requested={remove_requested}")
            
            if not month_str:
                QMessageBox.critical(self, "Error", "No month selected. Please select a month.")
                return
                
            # Parse year and month from the string
            try:
                year, month = month_str.split('-')
                year = int(year)
                month = int(month)
                print(f"DEBUG: Parsed year={year}, month={month}")
            except ValueError as e:
                QMessageBox.critical(self, "Error", f"Invalid month format: {month_str}. Expected format: YYYY-MM")
                return
            
            if remove_requested:
                # Remove override by setting it to None or the base rent
                success = self.rent_tracker.set_monthly_override(
                    tenant_name=self.selected_tenant.name,
                    year=year,
                    month=month,
                    override_amount=None,  # This should remove the override
                    notes=f"Removed override. {notes}" if notes else "Override removed"
                )
                
                if success:
                    QMessageBox.information(self, "Override Removed", 
                        f"Monthly rent override for {month_str} has been removed.")
                else:
                    QMessageBox.critical(self, "Error", "Failed to remove monthly override.")
            else:
                # Add/update override
                success = self.rent_tracker.set_monthly_override(
                    tenant_name=self.selected_tenant.name,
                    year=year,
                    month=month,
                    override_amount=amount,
                    notes=notes
                )
                
                if success:
                    QMessageBox.information(self, "Override Applied", 
                        f"Monthly rent override for {month_str} set to ${amount:.2f}\n"
                        f"Reason: {notes if notes else 'None'}")
                else:
                    QMessageBox.critical(self, "Error", "Failed to apply monthly override.")
            
            # Refresh the display
            if success:
                # Force recalculation of delinquency balances after override change
                self.rent_tracker.check_and_update_delinquency()
                
                # Get updated tenant data
                updated_tenant = self.rent_tracker.get_tenant_by_name(self.selected_tenant.name)
                if updated_tenant:
                    self.load_tenant(updated_tenant)
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to process monthly override: {str(e)}")
    
    def process_yearly_override(self, override_data):
        """Process yearly rent override"""
        try:
            year = override_data['year']
            amount = override_data['override_amount']
            notes = override_data['notes']
            remove_requested = override_data['remove_requested']
            
            if remove_requested:
                # Remove override by setting it to None
                success = self.rent_tracker.set_yearly_override(
                    tenant_name=self.selected_tenant.name,
                    year=year,
                    override_amount=None,  # This should remove the override
                    notes=f"Removed override. {notes}" if notes else "Override removed"
                )
                
                if success:
                    QMessageBox.information(self, "Override Removed", 
                        f"Yearly rent override for {year} has been removed.")
                else:
                    QMessageBox.critical(self, "Error", "Failed to remove yearly override.")
            else:
                # Add/update override
                success = self.rent_tracker.set_yearly_override(
                    tenant_name=self.selected_tenant.name,
                    year=year,
                    override_amount=amount,
                    notes=notes
                )
                
                if success:
                    QMessageBox.information(self, "Override Applied", 
                        f"Yearly rent override for {year} set to ${amount:.2f}\n"
                        f"Reason: {notes if notes else 'None'}")
                else:
                    QMessageBox.critical(self, "Error", "Failed to apply yearly override.")
            
            # Refresh the display
            if success:
                updated_tenant = self.rent_tracker.get_tenant_by_name(self.selected_tenant.name)
                if updated_tenant:
                    self.load_tenant(updated_tenant)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to process yearly override: {str(e)}")

    def show_rent_modification_dialog(self):
        """Show rent modification dialog"""
        if not self.selected_tenant:
            return
            
        dialog = RentModificationDialog(self, self.selected_tenant.rent_amount, self.selected_tenant.name)
        if dialog.exec() == dialog.DialogCode.Accepted:
            rent_data = dialog.get_rent_data()
            self.process_rent_modification(rent_data)
    
    def process_rent_modification(self, rent_data):
        """Process rent modification"""
        try:
            new_rent = rent_data['new_rent']
            effective_date = rent_data['effective_date']
            date_type = rent_data['date_type']
            notes = rent_data['notes']
            
            # Determine if we should backdate
            backdate = (date_type == 'backdate')
            
            success = self.rent_tracker.modify_rent(
                self.selected_tenant.name, new_rent, effective_date, notes, backdate
            )
            
            if success:
                # Refresh the tenant data
                updated_tenant = self.rent_tracker.get_tenant_by_name(self.selected_tenant.name)
                if updated_tenant:
                    self.load_tenant(updated_tenant)
                
                message = f"Rent has been modified to ${new_rent:.2f}"
                if date_type == 'today':
                    message += " effective today."
                elif date_type == 'specific':
                    message += f" effective from {effective_date}."
                elif date_type == 'backdate':
                    message += f" backdated to the beginning of the current month."
                
                QMessageBox.information(self, "Rent Modified", message)
            else:
                QMessageBox.critical(self, "Error", "Failed to modify rent.")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to modify rent: {str(e)}")

    def show_monthly_override_dialog(self):
        """Show monthly override dialog"""
        if not self.selected_tenant:
            print("DEBUG: No tenant selected for monthly override")
            return
            
        print(f"DEBUG: Opening monthly override dialog for tenant: {self.selected_tenant.name}")
        dialog = MonthlyOverrideDialog(self, self.selected_tenant.rent_amount, self.selected_tenant.name)
        if dialog.exec() == dialog.DialogCode.Accepted:
            override_data = dialog.get_override_data()
            print(f"DEBUG: Dialog accepted, calling process_monthly_override with data: {override_data}")
            self.process_monthly_override(override_data)
        else:
            print("DEBUG: Dialog was cancelled or rejected")
    
    def show_renew_lease_dialog(self):
        """Show the lease renewal dialog"""
        if not self.selected_tenant:
            QMessageBox.warning(self, "No Tenant Selected", "Please select a tenant to renew their lease.")
            return
        
        dialog = RenewLeaseDialog(self, self.selected_tenant)
        if dialog.exec() == dialog.DialogCode.Accepted:
            renewal_data = dialog.get_renewal_data()
            self.process_lease_renewal(renewal_data)
    
    def process_lease_renewal(self, renewal_data):
        """Process the lease renewal"""
        try:
            start_date = renewal_data['start_date']
            end_date = renewal_data['end_date']
            period = renewal_data['period']
            unit = renewal_data['unit']
            notes = renewal_data['notes']
            new_rent = renewal_data.get('new_rent')
            queue_rent_change = renewal_data.get('queue_rent_change', False)
            old_rent = renewal_data.get('old_rent', 0.0)
            
            # DON'T update rental period immediately - schedule it instead!
            # The old code that caused the issue:
            # self.selected_tenant.rental_period = [start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')]
            
            # Generate period description for the queue
            period_description = f"{period} {unit.lower()}"
            
            # Schedule the rental period change to happen on the start date
            period_change_action_id = self.action_queue.add_rental_period_change(
                tenant_id=self.selected_tenant.tenant_id,
                new_start_date=start_date.strftime('%Y-%m-%d'),
                new_end_date=end_date.strftime('%Y-%m-%d'),
                effective_date=start_date.strftime('%Y-%m-%d'),
                period_description=period_description,
                notes=f"Lease renewal. {notes}" if notes else "Lease renewal."
            )
            
            # DON'T generate months_to_charge now - let it happen when the action executes
            # The old code that caused issues:
            # new_months = []
            # current_date = start_date
            # while current_date <= end_date:
            #     new_months.append([current_date.year, current_date.month])
            #     ...
            # self.selected_tenant.months_to_charge.extend(new_months)
            
            # Queue rent change if needed (instead of applying immediately)
            rent_change_action_id = None
            if queue_rent_change and new_rent != old_rent:
                rent_change_notes = f"Rent change from lease renewal. {notes}" if notes else "Rent change from lease renewal."
                rent_change_action_id = self.action_queue.add_rent_change(
                    tenant_id=self.selected_tenant.tenant_id,
                    new_rent=new_rent,
                    effective_date=start_date.strftime('%Y-%m-%d'),
                    old_rent=old_rent,
                    notes=rent_change_notes
                )
            
            # Add renewal record for tracking
            if not hasattr(self.selected_tenant, 'lease_renewals'):
                self.selected_tenant.lease_renewals = []
            
            renewal_record = {
                'renewal_date': date.today().strftime('%Y-%m-%d'),
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'period': f"{period} {unit.lower()}",
                'notes': notes,
                'period_change_action_id': period_change_action_id,
                'rent_change_queued': {
                    'old_rent': old_rent,
                    'new_rent': new_rent,
                    'action_id': rent_change_action_id
                } if queue_rent_change else None
            }
            self.selected_tenant.lease_renewals.append(renewal_record)
            
            # Save changes
            self.save_tenants_async()
            
            # Prepare success message
            message = (f"Lease renewal scheduled for {self.selected_tenant.name}\n"
                      f"New period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}\n"
                      f"Duration: {period} {unit.lower()}\n\n"
                      f"The rental period change will take effect on {start_date.strftime('%B %d, %Y')}.")
            
            if queue_rent_change:
                if new_rent > old_rent:
                    change_text = f"increase from ${old_rent:.2f} to ${new_rent:.2f}"
                else:
                    change_text = f"decrease from ${old_rent:.2f} to ${new_rent:.2f}"
                message += f"\n\nRent {change_text} has also been scheduled for {start_date.strftime('%B %d, %Y')}."
            
            message += "\n\nBoth changes are now in the action queue and will be applied automatically on the scheduled date."
            
            # Show success message
            QMessageBox.information(self, "Lease Renewed", message)
            
            # Refresh the tenant data
            updated_tenant = self.rent_tracker.get_tenant_by_name(self.selected_tenant.name)
            if updated_tenant:
                self.load_tenant(updated_tenant)
            
        except Exception as e:
            QMessageBox.critical(self, "Renewal Error", f"Failed to renew lease: {str(e)}")
    
    def show_scheduled_actions(self):
        """Show the scheduled actions dialog for the selected tenant"""
        if not self.selected_tenant:
            QMessageBox.warning(self, "No Tenant Selected", "Please select a tenant to view their scheduled actions.")
            return
        
        dialog = ActionQueueDialog(self, self.action_queue, self.selected_tenant)
        dialog.exec()
    
    def show_schedule_notification_dialog(self):
        """Show the schedule notification dialog for the selected tenant"""
        if not self.selected_tenant:
            QMessageBox.warning(self, "No Tenant Selected", "Please select a tenant to schedule notifications for.")
            return
        
        try:
            from ui.queue_control_dialog import QueueControlDialog
            
            # Create queue control dialog
            dialog = QueueControlDialog(self, action_queue=self.action_queue, rent_tracker=self.rent_tracker)
            
            # Pre-select the current tenant in the dialog
            if hasattr(dialog, 'tenant_combo'):
                tenant_name = f"{self.selected_tenant.name} (ID: {self.selected_tenant.tenant_id})"
                index = dialog.tenant_combo.findText(tenant_name)
                if index >= 0:
                    dialog.tenant_combo.setCurrentIndex(index)
            
            # Connect to refresh when queue is updated
            dialog.queue_updated.connect(self.refresh_monthly_balance_display)
            
            dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open notification dialog: {str(e)}")
            print(f"[ERROR] Failed to open notification dialog: {e}")
                
    def show_query_dialog(self):
        """Show the rent query system dialog"""
        if not self.rent_tracker:
            QMessageBox.warning(self, "Error", "No rent tracker available")
            return
        
        # Pass the currently selected tenant name to the dialog
        default_tenant_name = None
        if self.selected_tenant:
            default_tenant_name = self.selected_tenant.name
        
        dialog = RentQueryDialog(self, self.rent_tracker, default_tenant_name)
        dialog.exec()

    def refresh_tenant_dropdown(self):
        """Refresh the tenant dropdown list"""
        if hasattr(self, 'tenant_dropdown'):
            current_text = self.tenant_dropdown.currentText()
            self.populate_tenant_dropdown()
            
            # If we have a currently selected tenant, make sure it's reflected in dropdown
            if hasattr(self, 'selected_tenant') and self.selected_tenant:
                self.update_dropdown_selection(self.selected_tenant)
            else:
                # Try to maintain the previously selected tenant
                if current_text and current_text not in ["Select a tenant...", "No tenants available", "Error loading tenants"]:
                    index = self.tenant_dropdown.findText(current_text)
                    if index >= 0:
                        self.tenant_dropdown.setCurrentIndex(index)
    
    def backup_tenant(self, tenant):
        """Backup a tenant's data to a JSON file"""
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import json
        from datetime import datetime
        
        try:
            # Get save file location
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Backup Tenant Data",
                f"{tenant.name}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                "JSON Files (*.json);;All Files (*)"
            )
            
            if not file_path:
                return
            
            # Create backup data structure
            backup_data = {
                'backup_date': datetime.now().isoformat(),
                'tenant_name': tenant.name,
                'tenant_data': {
                    'name': tenant.name,
                    'contact_info': tenant.contact_info,
                    'rent_amount': tenant.rent_amount,
                    'deposit_amount': tenant.deposit_amount,
                    'rental_period': tenant.rental_period,
                    'rent_due_date': tenant.rent_due_date,
                    'account_status': tenant.account_status,
                    'delinquency_balance': tenant.delinquency_balance,
                    'overpayment_credit': tenant.overpayment_credit,
                    'service_credit': getattr(tenant, 'service_credit', 0.0),
                    'payment_history': tenant.payment_history if hasattr(tenant, 'payment_history') else [],
                    'notes': tenant.notes,
                    'last_modified': datetime.now().isoformat()
                }
            }
            
            # Write to file
            with open(file_path, 'w') as f:
                json.dump(backup_data, f, indent=2)
            
            QMessageBox.information(self, "Backup Successful", 
                f"Tenant data for {tenant.name} has been backed up to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Backup Error", f"Failed to backup tenant data: {str(e)}")
    
    def delete_tenant(self, tenant):
        """Delete a tenant with confirmation and automatic backup"""
        from PyQt6.QtWidgets import QMessageBox, QFileDialog
        import json
        from datetime import datetime
        import os
        
        try:
            # Validation: Make sure tenant exists before attempting deletion
            existing_tenant = self.rent_tracker.get_tenant_by_name(tenant.name)
            if not existing_tenant:
                QMessageBox.warning(self, "Tenant Not Found", f"Tenant '{tenant.name}' could not be found in the system.")
                return
            
            # First warning
            reply = QMessageBox.warning(self, "Delete Tenant", 
                f"Are you sure you want to delete {tenant.name}?\n\n"
                f"This action CANNOT be undone.\n"
                f"A backup will be created before deletion.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            
            if reply != QMessageBox.StandardButton.Yes:
                return
            
            # Second confirmation with explicit warning
            reply = QMessageBox.warning(self, "Final Confirmation - Type to Confirm", 
                f"This will PERMANENTLY DELETE {tenant.name} and all associated data.\n\n"
                f"Click YES to proceed with deletion.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            
            if reply != QMessageBox.StandardButton.Yes:
                return
            
            # Create automatic backups (both individual and full)
            backups_dir = 'resources/tenant_backups'
            os.makedirs(backups_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Sanitize tenant name for use in filename (replace slashes with underscores)
            safe_tenant_name = tenant.name.replace('/', '_').replace('\\', '_')
            
            # Backup individual tenant
            individual_backup_file = f"{backups_dir}/{safe_tenant_name}_deleted_{timestamp}.json"
            
            individual_backup_data = {
                'backup_date': datetime.now().isoformat(),
                'tenant_name': tenant.name,
                'backup_type': 'automatic_deletion_backup',
                'tenant_data': {
                    'name': tenant.name,
                    'contact_info': tenant.contact_info,
                    'rent_amount': tenant.rent_amount,
                    'deposit_amount': tenant.deposit_amount,
                    'rental_period': tenant.rental_period,
                    'rent_due_date': tenant.rent_due_date,
                    'account_status': tenant.account_status,
                    'delinquency_balance': tenant.delinquency_balance,
                    'overpayment_credit': tenant.overpayment_credit,
                    'service_credit': getattr(tenant, 'service_credit', 0.0),
                    'payment_history': tenant.payment_history if hasattr(tenant, 'payment_history') else [],
                    'notes': tenant.notes
                }
            }
            
            with open(individual_backup_file, 'w') as f:
                json.dump(individual_backup_data, f, indent=2)
            
            # Backup all tenants before deletion (safety measure)
            all_tenants = self.rent_tracker.get_all_tenants()
            all_backup_file = f"{backups_dir}/ALL_TENANTS_backup_before_delete_{timestamp}.json"
            
            all_backup_data = {
                'backup_date': datetime.now().isoformat(),
                'backup_type': 'all_tenants_safety_backup',
                'reason': f'Automatic backup before deleting tenant: {tenant.name}',
                'tenant_count': len(all_tenants),
                'tenants': []
            }
            
            for t in all_tenants:
                all_backup_data['tenants'].append({
                    'name': t.name,
                    'contact_info': t.contact_info,
                    'rent_amount': t.rent_amount,
                    'deposit_amount': t.deposit_amount,
                    'rental_period': t.rental_period,
                    'rent_due_date': t.rent_due_date,
                    'account_status': t.account_status,
                    'delinquency_balance': t.delinquency_balance,
                    'overpayment_credit': t.overpayment_credit,
                    'service_credit': getattr(t, 'service_credit', 0.0),
                    'payment_history': t.payment_history if hasattr(t, 'payment_history') else [],
                    'notes': t.notes
                })
            
            with open(all_backup_file, 'w') as f:
                json.dump(all_backup_data, f, indent=2)
            
            # Now delete from rent tracker
            tenant_id = self.rent_tracker.delete_tenant(tenant.name)
            
            if tenant_id:
                # Delete from database as well
                if hasattr(self, 'rent_db') and self.rent_db:
                    self.rent_db.delete_tenant(tenant_id)
                
                self.save_tenants_async()
                
                QMessageBox.information(self, "Tenant Deleted", 
                    f"{tenant.name} has been permanently deleted.\n\n"
                    f"Individual backup: {individual_backup_file}\n"
                    f"Full system backup: {all_backup_file}")
                
                # Refresh the UI - clear the current view and reset dropdown
                self.tenant_dropdown.setCurrentIndex(0)
                self.refresh_tenant_dropdown()
                self.refresh_tenant_display()
            else:
                QMessageBox.critical(self, "Deletion Failed", f"Failed to delete tenant {tenant.name}.")
            
        except Exception as e:
            QMessageBox.critical(self, "Deletion Error", f"Failed to delete tenant: {str(e)}")

    # ==================== Rental Summaries Dialog Methods ====================
    
    def show_rental_period_summary(self):
        """Show rental period summary dialog"""
        if not self.selected_tenant:
            QMessageBox.warning(self, "No Tenant", "Please select a tenant first.")
            return
        
        try:
            # Get the summary data
            summary = self.rental_summaries.get_rental_period_summary(self.selected_tenant.tenant_id)
            
            if not summary:
                QMessageBox.warning(self, "Summary Error", "Could not generate rental period summary.")
                return
            
            # Create a styled dialog with rich formatting
            from PyQt6.QtWidgets import QDialog, QTextEdit, QVBoxLayout, QHBoxLayout
            from PyQt6.QtGui import QFont, QColor, QTextCursor, QTextCharFormat
            from PyQt6.QtCore import Qt
            
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Rental Period Summary - {self.selected_tenant.name}")
            dialog.setMinimumWidth(800)
            dialog.setMinimumHeight(600)
            
            layout = QVBoxLayout()
            
            # Create formatted text editor
            text_edit = QTextEdit()
            text_edit.setReadOnly(True)
            text_edit.setStyleSheet("""
                QTextEdit {
                    background-color: #ffffff;
                    color: #212529;
                    border: 1px solid #dee2e6;
                    font-family: 'Segoe UI', Arial, sans-serif;
                    font-size: 11px;
                    padding: 10px;
                }
            """)
            
            # Format and set the text with rich formatting
            cursor = text_edit.textCursor()
            self._append_summary_formatted(text_edit, cursor, summary, 'rental_period')
            
            layout.addWidget(text_edit)
            
            # Buttons layout
            button_layout = QHBoxLayout()
            print_btn = QPushButton('Print')
            close_btn = QPushButton('Close')
            
            print_btn.clicked.connect(lambda: self.print_summary(
                self._format_summary_display(summary, 'rental_period'),
                f"Rental Period Summary - {self.selected_tenant.name}"
            ))
            close_btn.clicked.connect(dialog.accept)
            
            button_layout.addStretch()
            button_layout.addWidget(print_btn)
            button_layout.addWidget(close_btn)
            layout.addLayout(button_layout)
            
            dialog.setLayout(layout)
            dialog.exec()
        
        except Exception as e:
            logger.error("RentManagementTab", f"Error showing rental period summary: {str(e)}")
            QMessageBox.critical(self, "Summary Error", f"Failed to generate summary: {str(e)}")
    
    def _append_summary_formatted(self, text_edit, cursor, summary, summary_type):
        """Append formatted summary text to a QTextEdit"""
        from PyQt6.QtGui import QTextCharFormat, QFont, QColor
        
        # Title format
        title_format = QTextCharFormat()
        title_format.setFont(QFont('Arial', 14, QFont.Weight.Bold))
        title_format.setForeground(QColor('#0056b3'))
        
        # Header format
        header_format = QTextCharFormat()
        header_format.setFont(QFont('Arial', 11, QFont.Weight.Bold))
        header_format.setBackground(QColor('#f8f9fa'))
        header_format.setForeground(QColor('#0056b3'))
        
        # Label format (for field names)
        label_format = QTextCharFormat()
        label_format.setFont(QFont('Arial', 10, QFont.Weight.Bold))
        label_format.setForeground(QColor('#495057'))
        
        # Value format
        value_format = QTextCharFormat()
        value_format.setFont(QFont('Arial', 10))
        value_format.setForeground(QColor('#212529'))
        
        # Warning/Delinquent format (red)
        warning_format = QTextCharFormat()
        warning_format.setFont(QFont('Arial', 10, QFont.Weight.Bold))
        warning_format.setForeground(QColor('#dc3545'))
        
        # Success format (green)
        success_format = QTextCharFormat()
        success_format.setFont(QFont('Arial', 10, QFont.Weight.Bold))
        success_format.setForeground(QColor('#28a745'))
        
        # Helper function to add formatted text
        def add_text(fmt, text):
            cursor.setCharFormat(fmt)
            text_edit.insertPlainText(text)
        
        # Title
        add_text(title_format, f"{'='*70}\n")
        if summary_type == 'rental_period':
            add_text(title_format, "RENTAL PERIOD SUMMARY\n")
        elif summary_type == 'monthly':
            add_text(title_format, f"MONTHLY SUMMARY - {summary.get('month_display', 'N/A')}\n")
        elif summary_type == 'yearly':
            add_text(title_format, f"YEARLY SUMMARY - {summary.get('year', 'N/A')}\n")
        
        add_text(title_format, f"{'='*70}\n\n")
        
        # Tenant info
        add_text(label_format, "Tenant: ")
        add_text(value_format, f"{summary.get('tenant_name', 'Unknown')}\n")
        
        add_text(label_format, "Account Status: ")
        status = summary.get('account_status', 'Unknown')
        status_fmt = success_format if status == 'active' else warning_format
        add_text(status_fmt, f"{status}\n\n")
        
        # Lease period (for rental period)
        if summary_type == 'rental_period':
            add_text(label_format, "Lease Period: ")
            add_text(value_format, f"{summary.get('rental_start_date', 'N/A')} to {summary.get('rental_end_date', 'N/A')}\n")
            add_text(label_format, "Lease Days: ")
            add_text(value_format, f"{summary.get('rental_period_days', 0)} days\n")
            add_text(label_format, "Rent Amount: ")
            add_text(value_format, f"${summary.get('rent_amount', 0):.2f}\n\n")
        
        # Financial summary
        add_text(header_format, f"\n{'FINANCIAL SUMMARY':<70}\n")
        add_text(header_format, f"{'='*70}\n")
        
        add_text(label_format, "Total Expected Rent:     ")
        add_text(value_format, f"${summary.get('total_expected_rent', 0):>12,.2f}\n")
        
        add_text(label_format, "Total Paid:              ")
        add_text(value_format, f"${summary.get('total_paid', 0):>12,.2f}\n")
        
        add_text(label_format, "Balance:                 ")
        balance = summary.get('total_balance', 0)
        balance_fmt = warning_format if balance > 0 else success_format
        add_text(balance_fmt, f"${balance:>12,.2f}\n")
        
        add_text(label_format, "Payment Rate:            ")
        payment_rate = summary.get('payment_rate', 0)
        rate_fmt = success_format if payment_rate >= 100 else warning_format if payment_rate < 50 else value_format
        add_text(rate_fmt, f"{payment_rate:>12.1f}%\n\n")
        
        # Additional details
        if summary_type in ['rental_period', 'yearly']:
            add_text(label_format, "Months Paid in Full:     ")
            add_text(success_format, f"{summary.get('months_paid', 0)}\n")
            
            add_text(label_format, "Partial Payments:        ")
            add_text(value_format, f"{summary.get('months_partial', 0)}\n")
            
            add_text(label_format, "Delinquent Months:       ")
            delinquent_count = summary.get('months_delinquent', 0)
            delin_fmt = warning_format if delinquent_count > 0 else success_format
            add_text(delin_fmt, f"{delinquent_count}\n\n")
        
        # Credits and balances
        if summary_type == 'rental_period':
            add_text(header_format, f"\n{'CREDITS & ADJUSTMENTS':<70}\n")
            add_text(header_format, f"{'='*70}\n")
            
            add_text(label_format, "Overpayment Credit:      ")
            overpay = summary.get('overpayment_credit', 0)
            overpay_fmt = success_format if overpay > 0 else value_format
            add_text(overpay_fmt, f"${overpay:>12,.2f}\n")
            
            add_text(label_format, "Service Credit:          ")
            service = summary.get('service_credit', 0)
            service_fmt = success_format if service > 0 else value_format
            add_text(service_fmt, f"${service:>12,.2f}\n")
            
            add_text(label_format, "Delinquency Balance:     ")
            delinq = summary.get('delinquency_balance', 0)
            delinq_fmt = warning_format if delinq > 0 else success_format
            add_text(delinq_fmt, f"${delinq:>12,.2f}\n\n")
        
        # Delinquent months (for rental period)
        if summary_type == 'rental_period' and summary.get('delinquent_months', 0) > 0:
            add_text(header_format, f"\n{'DELINQUENT MONTHS':<70}\n")
            add_text(header_format, f"{'='*70}\n")
            add_text(header_format, f"{'Month':<30} | {'Balance':>12}\n")
            add_text(header_format, f"{'-'*70}\n")
            
            yearly_summaries = summary.get('yearly_summaries', [])
            delinquent_list = []
            
            for yearly in yearly_summaries:
                for monthly in yearly.get('monthly_details', []):
                    if monthly.get('is_delinquent'):
                        delinquent_list.append({
                            'month': monthly.get('month_display', 'Unknown'),
                            'balance': monthly.get('balance', 0)
                        })
            
            if delinquent_list:
                # Display as a formatted table
                for item in delinquent_list:
                    add_text(warning_format, f"  {item['month']:<28} | ")
                    add_text(warning_format, f"${item['balance']:>11,.2f}\n")
            add_text(value_format, "\n")
        
        # Payment history (for rental period)
        if summary_type == 'rental_period':
            # Collect all payments from all years
            all_payments = []
            yearly_summaries = summary.get('yearly_summaries', [])
            
            for yearly in yearly_summaries:
                for monthly in yearly.get('monthly_details', []):
                    for payment in monthly.get('payments', []):
                        # Calculate if this payment created overpayment
                        expected = monthly.get('expected_rent', 0)
                        total_paid_before = sum(p.get('amount', 0) for p in monthly.get('payments', []) if p.get('date', '') < payment.get('date', ''))
                        overpayment = max(0, (total_paid_before + payment.get('amount', 0)) - expected)
                        
                        all_payments.append({
                            **payment,
                            'month': monthly.get('month_display', 'Unknown'),
                            'overpayment': overpayment
                        })
            
            if all_payments:
                add_text(header_format, f"\n{'PAYMENT HISTORY':<70}\n")
                add_text(header_format, f"{'='*70}\n")
                add_text(header_format, f"{'Date':<12} | {'Month':<15} | {'Amount':>12} | {'Type':<12} | {'Overpayment':>12}\n")
                add_text(header_format, f"{'-'*70}\n")
                
                for payment in all_payments:
                    add_text(value_format,
                        f"  {payment.get('date', 'Unknown'):<10} | "
                        f"{payment.get('month', 'Unknown'):<15} | "
                        f"${payment.get('amount', 0):>11,.2f} | "
                        f"{payment.get('type', 'Unknown'):<12} | "
                    )
                    if payment['overpayment'] > 0:
                        add_text(success_format, f"${payment['overpayment']:>11,.2f}\n")
                    else:
                        add_text(value_format, f"${payment['overpayment']:>11,.2f}\n")
                add_text(value_format, "\n")
        
        # Payments received (for monthly summaries)
        if summary_type == 'monthly' and summary.get('payments'):
            add_text(header_format, f"\n{'PAYMENTS RECEIVED':<70}\n")
            add_text(header_format, f"{'='*70}\n")
            add_text(header_format, f"{'Date':<12} | {'Amount':>12} | {'Type':<15}\n")
            add_text(header_format, f"{'-'*70}\n")
            
            for payment in summary.get('payments', []):
                add_text(value_format,
                    f"  {payment.get('date', 'Unknown'):<10} | "
                    f"${payment.get('amount', 0):>11,.2f} | "
                    f"{payment.get('type', 'Unknown'):<15}\n"
                )
            add_text(value_format, "\n")
        
        # Monthly breakdown for yearly (shown as a table)
        if summary_type == 'yearly':
            monthly_details = summary.get('monthly_details', [])
            if monthly_details:
                add_text(header_format, f"\n{'MONTHLY BREAKDOWN':<70}\n")
                add_text(header_format, f"{'='*70}\n")
                add_text(header_format, f"{'Month':<18} | {'Expected':>12} | {'Paid':>12} | {'Balance':>12}\n")
                add_text(header_format, f"{'-'*70}\n")
                
                for month_data in monthly_details:
                    if isinstance(month_data, dict):
                        balance = month_data.get('balance', 0)
                        balance_fmt_inner = warning_format if balance > 0 else success_format
                        
                        add_text(value_format, f"  {month_data.get('month_display', 'N/A'):<16} | ")
                        add_text(value_format, f"${month_data.get('expected_rent', 0):>11,.2f} | ")
                        add_text(value_format, f"${month_data.get('total_paid', 0):>11,.2f} | ")
                        add_text(balance_fmt_inner, f"${balance:>11,.2f}\n")
                add_text(value_format, "\n")
        
        add_text(title_format, f"{'='*70}\n")
    
    def show_monthly_summary_dialog(self):
        """Show dialog to select month and display monthly summary"""
        if not self.selected_tenant:
            QMessageBox.warning(self, "No Tenant", "Please select a tenant first.")
            return
        
        try:
            # Get current date
            today = date.today()
            current_year = today.year
            current_month = today.month
            
            # Create selection dialog
            from PyQt6.QtWidgets import QDialog, QCalendarWidget
            
            dialog = QDialog(self)
            dialog.setWindowTitle("Select Month for Summary")
            dialog.setStyleSheet("""
                QDialog {
                    background-color: #ffffff;
                }
                QLabel {
                    color: #212529;
                    font-weight: bold;
                }
                QComboBox {
                    background-color: #ffffff;
                    color: #212529;
                    border: 1px solid #dee2e6;
                    border-radius: 4px;
                    padding: 5px;
                }
            """)
            
            layout = QVBoxLayout()
            
            # Year selector
            year_label = QLabel("Year:")
            year_combo = QComboBox()
            years = list(range(current_year - 3, current_year + 2))
            year_combo.addItems([str(y) for y in years])
            year_combo.setCurrentText(str(current_year))
            layout.addWidget(year_label)
            layout.addWidget(year_combo)
            
            # Month selector
            month_label = QLabel("Month:")
            month_combo = QComboBox()
            months = ["January", "February", "March", "April", "May", "June",
                     "July", "August", "September", "October", "November", "December"]
            month_combo.addItems(months)
            month_combo.setCurrentIndex(current_month - 1)
            layout.addWidget(month_label)
            layout.addWidget(month_combo)
            
            # Buttons
            button_layout = QHBoxLayout()
            ok_btn = QPushButton("Show Summary")
            cancel_btn = QPushButton("Cancel")
            ok_btn.clicked.connect(dialog.accept)
            cancel_btn.clicked.connect(dialog.reject)
            button_layout.addWidget(ok_btn)
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)
            
            dialog.setLayout(layout)
            
            if dialog.exec() == QDialog.DialogCode.Accepted:
                selected_year = int(year_combo.currentText())
                selected_month = month_combo.currentIndex() + 1
                
                # Get the summary
                summary = self.rental_summaries.get_monthly_summary(
                    self.selected_tenant.tenant_id,
                    selected_year,
                    selected_month
                )
                
                if not summary:
                    QMessageBox.warning(self, "Summary Error", "Could not generate monthly summary.")
                    return
                
                # Create a styled dialog with rich formatting
                from PyQt6.QtWidgets import QDialog, QTextEdit
                
                summary_dialog = QDialog(self)
                summary_dialog.setWindowTitle(f"Monthly Summary - {self.selected_tenant.name} ({months[selected_month-1]} {selected_year})")
                summary_dialog.setMinimumWidth(750)
                summary_dialog.setMinimumHeight(550)
                
                layout = QVBoxLayout()
                
                # Create formatted text editor
                text_edit = QTextEdit()
                text_edit.setReadOnly(True)
                text_edit.setStyleSheet("""
                    QTextEdit {
                        background-color: #ffffff;
                        color: #212529;
                        border: 1px solid #dee2e6;
                        font-family: 'Segoe UI', Arial, sans-serif;
                        font-size: 11px;
                        padding: 10px;
                    }
                """)
                
                cursor = text_edit.textCursor()
                self._append_summary_formatted(text_edit, cursor, summary, 'monthly')
                
                layout.addWidget(text_edit)
                
                # Buttons
                button_layout = QHBoxLayout()
                print_btn = QPushButton('Print')
                close_btn = QPushButton('Close')
                
                print_btn.clicked.connect(lambda: self.print_summary(
                    self._format_summary_display(summary, 'monthly', months[selected_month-1], selected_year),
                    f"Monthly Summary - {self.selected_tenant.name} - {months[selected_month-1]} {selected_year}"
                ))
                close_btn.clicked.connect(summary_dialog.accept)
                
                button_layout.addStretch()
                button_layout.addWidget(print_btn)
                button_layout.addWidget(close_btn)
                layout.addLayout(button_layout)
                
                summary_dialog.setLayout(layout)
                summary_dialog.exec()
        
        except Exception as e:
            logger.error("RentManagementTab", f"Error showing monthly summary: {str(e)}")
            QMessageBox.critical(self, "Summary Error", f"Failed to generate summary: {str(e)}")
    
    def show_yearly_summary_dialog(self):
        """Show dialog to select year and display yearly summary"""
        if not self.selected_tenant:
            QMessageBox.warning(self, "No Tenant", "Please select a tenant first.")
            return
        
        try:
            # Get current year
            current_year = date.today().year
            
            # Create selection dialog
            from PyQt6.QtWidgets import QDialog
            
            dialog = QDialog(self)
            dialog.setWindowTitle("Select Year for Summary")
            dialog.setStyleSheet("""
                QDialog {
                    background-color: #ffffff;
                }
                QLabel {
                    color: #212529;
                    font-weight: bold;
                }
                QComboBox {
                    background-color: #ffffff;
                    color: #212529;
                    border: 1px solid #dee2e6;
                    border-radius: 4px;
                    padding: 5px;
                }
            """)
            
            layout = QVBoxLayout()
            
            # Year selector
            year_label = QLabel("Year:")
            year_combo = QComboBox()
            years = list(range(current_year - 5, current_year + 2))
            year_combo.addItems([str(y) for y in years])
            year_combo.setCurrentText(str(current_year))
            layout.addWidget(year_label)
            layout.addWidget(year_combo)
            
            # Buttons
            button_layout = QHBoxLayout()
            ok_btn = QPushButton("Show Summary")
            cancel_btn = QPushButton("Cancel")
            ok_btn.clicked.connect(dialog.accept)
            cancel_btn.clicked.connect(dialog.reject)
            button_layout.addWidget(ok_btn)
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)
            
            dialog.setLayout(layout)
            
            if dialog.exec() == QDialog.DialogCode.Accepted:
                selected_year = int(year_combo.currentText())
                
                # Get the summary
                summary = self.rental_summaries.get_yearly_summary(
                    self.selected_tenant.tenant_id,
                    selected_year
                )
                
                if not summary:
                    QMessageBox.warning(self, "Summary Error", "Could not generate yearly summary.")
                    return
                
                # Create a styled dialog with rich formatting
                from PyQt6.QtWidgets import QDialog, QTextEdit
                
                summary_dialog = QDialog(self)
                summary_dialog.setWindowTitle(f"Yearly Summary - {self.selected_tenant.name} ({selected_year})")
                summary_dialog.setMinimumWidth(800)
                summary_dialog.setMinimumHeight(600)
                
                layout = QVBoxLayout()
                
                # Create formatted text editor
                text_edit = QTextEdit()
                text_edit.setReadOnly(True)
                text_edit.setStyleSheet("""
                    QTextEdit {
                        background-color: #ffffff;
                        color: #212529;
                        border: 1px solid #dee2e6;
                        font-family: 'Segoe UI', Arial, sans-serif;
                        font-size: 11px;
                        padding: 10px;
                    }
                """)
                
                cursor = text_edit.textCursor()
                self._append_summary_formatted(text_edit, cursor, summary, 'yearly')
                
                # Add monthly breakdown if available
                monthly_details = summary.get('monthly_details', [])
                if monthly_details:
                    from PyQt6.QtGui import QTextCharFormat, QFont, QColor
                    
                    header_format = QTextCharFormat()
                    header_format.setFont(QFont('Arial', 11, QFont.Weight.Bold))
                    header_format.setBackground(QColor('#f8f9fa'))
                    
                    cursor.setCharFormat(header_format)
                    text_edit.insertPlainText(f"\n{'MONTHLY BREAKDOWN':<70}\n")
                    cursor.setCharFormat(header_format)
                    text_edit.insertPlainText(f"{'='*70}\n")
                    
                    value_format = QTextCharFormat()
                    value_format.setFont(QFont('Courier New', 10))
                    
                    for month_data in monthly_details:
                        if isinstance(month_data, dict):
                            cursor.setCharFormat(value_format)
                            text_edit.insertPlainText(
                                f"  {month_data.get('month_display', 'N/A'):20s} | "
                                f"Expected: ${month_data.get('expected_rent', 0):>10,.2f} | "
                                f"Paid: ${month_data.get('total_paid', 0):>10,.2f} | "
                                f"Balance: ${month_data.get('balance', 0):>10,.2f}\n"
                            )
                
                layout.addWidget(text_edit)
                
                # Buttons
                button_layout = QHBoxLayout()
                print_btn = QPushButton('Print')
                close_btn = QPushButton('Close')
                
                print_btn.clicked.connect(lambda: self.print_summary(
                    self._format_summary_display(summary, 'yearly', year=selected_year),
                    f"Yearly Summary - {self.selected_tenant.name} - {selected_year}"
                ))
                close_btn.clicked.connect(summary_dialog.accept)
                
                button_layout.addStretch()
                button_layout.addWidget(print_btn)
                button_layout.addWidget(close_btn)
                layout.addLayout(button_layout)
                
                summary_dialog.setLayout(layout)
                summary_dialog.exec()
        
        except Exception as e:
            logger.error("RentManagementTab", f"Error showing yearly summary: {str(e)}")
            QMessageBox.critical(self, "Summary Error", f"Failed to generate summary: {str(e)}")
    
    def _format_summary_display(self, summary, summary_type, month_name=None, year=None):
        """Format summary data for display"""
        text_lines = []
        
        if not summary:
            return "No summary data available"
        
        if summary_type == 'rental_period':
            text_lines.append("=" * 60)
            text_lines.append("RENTAL PERIOD SUMMARY")
            text_lines.append("=" * 60)
            text_lines.append("")
            text_lines.append(f"Tenant: {summary.get('tenant_name', 'Unknown')}")
            text_lines.append(f"Account Status: {summary.get('account_status', 'Unknown')}")
            text_lines.append("")
            text_lines.append(f"Lease Period: {summary.get('rental_start_date', 'N/A')} to {summary.get('rental_end_date', 'N/A')}")
            text_lines.append(f"Lease Days: {summary.get('rental_period_days', 0)} days")
            text_lines.append("")
            text_lines.append(f"Rent Amount: ${summary.get('rent_amount', 0):.2f}")
            text_lines.append(f"Deposit Amount: ${summary.get('deposit_amount', 0):.2f}")
            text_lines.append("")
            text_lines.append("=" * 60)
            text_lines.append(f"Total Expected Rent: ${summary.get('total_expected_rent', 0):.2f}")
            text_lines.append(f"Total Paid: ${summary.get('total_paid', 0):.2f}")
            text_lines.append(f"Balance: ${summary.get('total_balance', 0):.2f}")
            text_lines.append(f"Payment Rate: {summary.get('payment_rate', 0):.1f}%")
            text_lines.append("")
            text_lines.append(f"Overpayment Credit: ${summary.get('overpayment_credit', 0):.2f}")
            text_lines.append(f"Service Credit: ${summary.get('service_credit', 0):.2f}")
            text_lines.append(f"Delinquency Balance: ${summary.get('delinquency_balance', 0):.2f}")
            text_lines.append(f"Delinquent Months: {summary.get('delinquent_months', 0)}")
            text_lines.append("=" * 60)
        
        elif summary_type == 'monthly':
            text_lines.append("=" * 60)
            text_lines.append(f"MONTHLY SUMMARY - {month_name} {year}")
            text_lines.append("=" * 60)
            text_lines.append("")
            text_lines.append(f"Tenant: {summary.get('tenant_name', 'Unknown')}")
            text_lines.append("")
            text_lines.append(f"Expected Rent: ${summary.get('expected_rent', 0):.2f}")
            text_lines.append(f"Total Paid: ${summary.get('total_paid', 0):.2f}")
            text_lines.append(f"Balance: ${summary.get('balance', 0):.2f}")
            text_lines.append(f"Status: {summary.get('status', 'Unknown')}")
            text_lines.append(f"Delinquent: {'Yes' if summary.get('is_delinquent') else 'No'}")
            text_lines.append("")
            
            # Payment details
            payments = summary.get('payments', [])
            if payments:
                text_lines.append("Payments Received:")
                text_lines.append("-" * 60)
                for payment in payments:
                    if isinstance(payment, dict):
                        text_lines.append(f"  Date: {payment.get('date', 'N/A')}, "
                                        f"Amount: ${payment.get('amount', 0):.2f}, "
                                        f"Type: {payment.get('type', 'N/A')}")
            else:
                text_lines.append("No payments received this month")
            text_lines.append("=" * 60)
        
        elif summary_type == 'yearly':
            text_lines.append("=" * 60)
            text_lines.append(f"YEARLY SUMMARY - {year}")
            text_lines.append("=" * 60)
            text_lines.append("")
            text_lines.append(f"Tenant: {summary.get('tenant_name', 'Unknown')}")
            text_lines.append("")
            text_lines.append(f"Total Expected Rent: ${summary.get('total_expected_rent', 0):.2f}")
            text_lines.append(f"Total Paid: ${summary.get('total_paid', 0):.2f}")
            text_lines.append(f"Total Balance: ${summary.get('total_balance', 0):.2f}")
            text_lines.append(f"Payment Rate: {summary.get('payment_rate', 0):.1f}%")
            text_lines.append("")
            text_lines.append(f"Months Paid: {summary.get('months_paid', 0)}")
            text_lines.append(f"Months Partial: {summary.get('months_partial', 0)}")
            text_lines.append(f"Months Delinquent: {summary.get('months_delinquent', 0)}")
            text_lines.append("")
            
            # Monthly details
            monthly_details = summary.get('monthly_details', [])
            if monthly_details:
                text_lines.append("Monthly Breakdown:")
                text_lines.append("-" * 60)
                for month_data in monthly_details:
                    if isinstance(month_data, dict):
                        text_lines.append(f"  {month_data.get('month_display', 'N/A')}: "
                                        f"Expected: ${month_data.get('expected_rent', 0):.2f}, "
                                        f"Paid: ${month_data.get('total_paid', 0):.2f}, "
                                        f"Balance: ${month_data.get('balance', 0):.2f}")
            text_lines.append("=" * 60)
        
        return "\n".join(text_lines)
    
    def print_summary(self, summary_text, title):
        """Print summary to PDF or printer"""
        try:
            from PyQt6.QtPrintSupport import QPrinter, QPrintDialog
            from PyQt6.QtGui import QTextDocument
            
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            dialog = QPrintDialog(printer, self)
            
            if dialog.exec() == dialog.DialogCode.Accepted:
                doc = QTextDocument()
                doc.setPlainText(summary_text)
                doc.print(printer)
                QMessageBox.information(self, "Print Successful", "Summary printed successfully!")
        
        except Exception as e:
            logger.error("RentManagementTab", f"Error printing summary: {str(e)}")
            QMessageBox.critical(self, "Print Error", f"Failed to print summary: {str(e)}")
